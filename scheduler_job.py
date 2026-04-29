import io
import os
import json
import time
import pyodbc
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
import win32com.client as win32 # Windows specific

# --- CONFIG & PATHS ---
CONFIG_PATH = Path(__file__).parent / "config" / "schedules.json"
LOG_PATH = Path(__file__).parent / "logs" / "scheduler.log"
os.makedirs(Path(__file__).parent / "config", exist_ok=True)
os.makedirs(Path(__file__).parent / "logs", exist_ok=True)

def log_message(msg):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {msg}\n")
    print(f"[{timestamp}] {msg}")

# --- REPLICATED LOGIC FROM APP.PY ---
def get_bidb_connection(database="DWH"):
    conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER=BIDB;DATABASE={database};Trusted_Connection=yes;'
    return pyodbc.connect(conn_str)

def run_query(sql, database="DWH"):
    conn = get_bidb_connection(database)
    try:
        return pd.read_sql(sql, conn)
    finally:
        conn.close()

def to_excel_binary(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sonuçlar')
        worksheet = writer.sheets['Sonuçlar']
        header_fill = PatternFill(start_color='8C28E8', end_color='8C28E8', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, name='Segoe UI')
        header_alignment = Alignment(horizontal='center', vertical='center')
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 50)
    return output.getvalue()

def send_mail_outlook(to, subject, body, attachment_bytes=None, attachment_name="report.xlsx"):
    try:
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.Subject = subject
        mail.HTMLBody = body
        if attachment_bytes:
            import tempfile
            tmp_path = os.path.join(tempfile.gettempdir(), attachment_name)
            with open(tmp_path, "wb") as tp:
                tp.write(attachment_bytes)
            mail.Attachments.Add(tmp_path)
        mail.Send()
        return True, "Başarılı"
    except Exception as e:
        return False, str(e)

def calculate_n_minus_1():
    """Calculates previous month in YYYYMM format."""
    first_day_current = datetime.now().replace(day=1)
    last_day_prev = first_day_current - timedelta(days=1)
    return last_day_prev.strftime("%Y%m")

# --- MAIN JOB ENGINE ---
def run_scheduled_jobs():
    if not CONFIG_PATH.exists():
        log_message("No schedules found.")
        return

    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        schedules = json.load(f)

    if not schedules:
        log_message("Schedule list is empty.")
        return

    # Load templates from fix_queries.json
    queries_path = Path(__file__).parent / "config" / "fix_queries.json"
    if not queries_path.exists():
        log_message("Error: fix_queries.json not found.")
        return
    with open(queries_path, "r", encoding="utf-8") as f:
        templates = json.load(f)

    for job in schedules:
        if not job.get("active", False):
            continue

        tpl_name = job["template_name"]
        recipient = job["recipient"]
        log_message(f"Starting job: {tpl_name} for {recipient}")

        try:
            if tpl_name not in templates:
                log_message(f"Error: Template {tpl_name} not found in fix_queries.json")
                continue

            tpl = templates[tpl_name]
            donem = calculate_n_minus_1()
            sql = tpl["sql_template"].replace("{{donem}}", donem)
            
            log_message(f"Running query for period {donem}...")
            df = run_query(sql)
            
            log_message(f"Generating Excel ({len(df)} rows)...")
            excel_bin = to_excel_binary(df)
            
            subject = f"Otomatik Rapor: {tpl_name} - {donem}"
            body = f"Merhaba,<br><br>Zamanlanmış <b>{tpl_name}</b> raporu <b>{donem}</b> dönemi için otomatik olarak oluşturulmuştur.<br><br>İyi çalışmalar."
            
            log_message(f"Sending mail to {recipient}...")
            ok, msg = send_mail_outlook(recipient, subject, body, excel_bin, f"{tpl_name}_{donem}.xlsx")
            
            if ok:
                log_message(f"Job completed successfully: {tpl_name}")
                job["last_run"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                job["status"] = "Success"
            else:
                log_message(f"Mail failed: {msg}")
                job["status"] = f"Mail Error: {msg}"

        except Exception as e:
            log_message(f"Job failed with exception: {str(e)}")
            job["status"] = f"Exception: {str(e)}"

    # Update last run info
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(schedules, f, indent=4)

if __name__ == "__main__":
    run_scheduled_jobs()
