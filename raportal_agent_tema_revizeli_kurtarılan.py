from pathlib import Path
import asyncio
import concurrent.futures
import html
import traceback
import pandas as pd
import re
import streamlit as st
import types
import io
import pyodbc
import json
import os
import time
import threading
from urllib.parse import quote
import requests
from requests_ntlm import HttpNtlmAuth
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pbi_parser import parse_pbi_file, convert_to_pbit_bytes
from pbi_robot_engine import trigger_pbi_robot_export

# Create temp folder for Robot interaction
if not os.path.exists("temp_pbix"):
    os.makedirs("temp_pbix")

st.set_page_config(page_title="Raportal Recommendation Copilot", layout="wide")

try:
    import google.generativeai as genai
except ImportError:
    genai = None


APP_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
@import url('https://fonts.googleapis.com/icon?family=Material+Icons+Outlined');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

.stApp {
    background-color: #f7f9fc;
}

/* Sidebar Styling - Narrower and more subtle */
section[data-testid="stSidebar"] {
    background-color: #8c28e8 !important;
    background-image: none !important;
    border-right: none !important;
    width: 280px !important;
}

[data-testid="stSidebar"] > div:first-child {
    width: 280px !important;
}

section[data-testid="stSidebar"] .stMarkdown h1, 
section[data-testid="stSidebar"] .stMarkdown h2, 
section[data-testid="stSidebar"] .stMarkdown h3, 
section[data-testid="stSidebar"] .stMarkdown h4,
section[data-testid="stSidebar"] .stMarkdown p,
section[data-testid="stSidebar"] label {
    color: #ffffff !important;
}

/* Specific fix for selectbox and text input visibility in sidebar */
section[data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"],
section[data-testid="stSidebar"] .stTextInput input {
    background-color: rgba(255, 255, 255, 0.1) !important;
    color: white !important;
    border: 1px solid rgba(255, 255, 255, 0.2) !important;
}

section[data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] * {
    color: white !important;
}

/* Fix dropdown menu items (the list when opened) */
div[data-baseweb="popover"] ul {
    background-color: #ffffff !important;
}

div[data-baseweb="popover"] li {
    color: #1a1a1a !important;
}

/* Reset icon font for material-icons class */
.material-icons-outlined {
    font-family: 'Material Icons Outlined' !important;
    font-size: 20px;
    margin-right: 12px;
    vertical-align: middle;
}

.sidebar-brand {
    display: flex;
    align-items: center;
    padding: 1rem 0.75rem 1.5rem 0.75rem;
    font-weight: 800;
    font-size: 1.2rem;
    letter-spacing: -0.01em;
}

.sidebar-menu-item {
    display: flex;
    align-items: center;
    padding: 0.6rem 0.75rem;
    margin: 0.2rem 0.5rem;
    border-radius: 8px;
    font-weight: 600;
    font-size: 0.9rem;
    cursor: pointer;
    transition: all 0.2s;
    text-decoration: none;
    color: rgba(255, 255, 255, 0.8) !important;
}

.sidebar-menu-item.active {
    background-color: rgba(255, 255, 255, 0.15) !important;
    color: #ffffff !important;
}

.sidebar-menu-item:hover {
    background-color: rgba(255, 255, 255, 0.1);
}

/* Streamlit Button Overrides to match sidebar menu */
div.stButton > button.sidebar-btn {
    background-color: transparent !important;
    color: rgba(255, 255, 255, 0.8) !important;
    border: none !important;
    padding: 0.75rem 1rem !important;
    width: 100% !important;
    text-align: left !important;
    display: flex !important;
    align-items: center !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    border-radius: 10px !important;
    margin-bottom: 0.25rem !important;
}

div.stButton > button.sidebar-btn:hover {
    background-color: rgba(255, 255, 255, 0.1) !important;
    color: #ffffff !important;
}

div.stButton > button.sidebar-btn-active {
    background-color: rgba(255, 255, 255, 0.15) !important;
    color: #ffffff !important;
    border: none !important;
    padding: 0.75rem 1rem !important;
    width: 100% !important;
    text-align: left !important;
    display: flex !important;
    align-items: center !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    border-radius: 10px !important;
}

.sql-container {
    background: #1e1e1e;
    color: #d4d4d4;
    padding: 1rem;
    border-radius: 8px;
    font-family: 'Courier New', Courier, monospace;
    font-size: 0.9rem;
    overflow-x: auto;
    border-left: 4px solid #8c28e8;
    margin: 1rem 0;
}

.stHeader {
    background: transparent !important;
}

.main-header {
    display: flex;
    align-items: center;
    padding-bottom: 2rem;
}

.header-logo {
    background: #f3e8ff;
    color: #8c28e8;
    width: 36px;
    height: 36px;
    border-radius: 8px;
    display: flex;
    align-items: center;
    justify-content: center;
    margin-right: 12px;
    font-size: 20px;
}

.header-title {
    font-weight: 700;
    font-size: 1.25rem;
    color: #1a1a1a;
}

.welcome-text h1 {
    font-weight: 800;
    font-size: 2.2rem;
    margin-bottom: 0.5rem;
    color: #1a1a1a;
}

.welcome-text p {
    color: #6b7280;
    font-size: 1.05rem;
    margin-bottom: 2rem;
}

/* Workflow Indicator */
.workflow-container {
    display: flex;
    align-items: center;
    margin-bottom: 2.5rem;
    gap: 12px;
}

.workflow-label {
    font-weight: 700;
    font-size: 0.75rem;
    color: #9ca3af;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-right: 12px;
}

.workflow-step {
    display: flex;
    align-items: center;
    font-size: 0.9rem;
    font-weight: 600;
    color: #a855f7;
}

.workflow-step span {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 20px;
    height: 20px;
    border: 2px solid #a855f7;
    border-radius: 50%;
    margin-right: 8px;
    font-size: 0.7rem;
}

.workflow-arrow {
    color: #d1d5db;
    margin: 0 4px;
    font-family: 'Material Icons Outlined' !important;
    font-size: 18px;
}

/* Navigation Cards */
.nav-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 20px;
    margin-bottom: 2rem;
}

.nav-card {
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 16px;
    padding: 1.5rem;
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
    transition: all 0.3s;
    cursor: pointer;
    position: relative;
}

.nav-card:hover {
    transform: translateY(-4px);
    box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.08);
    border-color: #8c28e8;
}

.card-num {
    position: absolute;
    top: 1rem;
    right: 1rem;
    color: #e5e7eb;
    font-size: 0.75rem;
    font-weight: 700;
}

.card-icon {
    width: 48px;
    height: 48px;
    background: #fdfbff;
    border-radius: 12px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 24px;
    margin-bottom: 1.25rem;
    color: #8c28e8;
}

.card-title {
    font-weight: 700;
    font-size: 1.1rem;
    color: #1a1a1a;
    margin-bottom: 0.5rem;
}

.card-desc {
    font-size: 0.9rem;
    color: #6b7280;
    line-height: 1.5;
}

/* Result Cards */
.result-card {
    background: white;
    border-radius: 16px;
    padding: 1.5rem;
    border: 1px solid #e5e7eb;
    border-left: 6px solid #8c28e8;
    margin-bottom: 1.5rem;
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
}

.result-title {
    font-weight: 700;
    font-size: 1.2rem;
    color: #1a1a1a;
    margin-bottom: 12px;
    display: flex;
    align-items: center;
}

.field-label {
    font-weight: 700;
    font-size: 0.75rem;
    color: #6b7280;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-top: 12px;
    margin-bottom: 4px;
}

.field-value {
    font-size: 0.95rem;
    color: #374151;
    line-height: 1.5;
    margin-bottom: 8px;
}
.badge {
    display: inline-block;
    padding: 4px 12px;
    border-radius: 20px;
    background: #f3e8ff;
    color: #8c28e8;
    font-size: 0.75rem;
    font-weight: 600;
    margin-right: 8px;
}

.badge-pop {
    background: #ecfdf5;
    color: #059669;
}

.result-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.05);
}
.result-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.05);
}

/* Professional Card & SQL Scroll Styles */
.custom-card {
    background: white;
    border-radius: 12px;
    padding: 1.5rem;
    border: 1px solid #e5e7eb;
    margin-bottom: 1.5rem;
    box-shadow: 0 2px 4px rgba(0,0,0,0.02);
}

.sql-scroll-box {
    max-height: 400px;
    overflow-y: auto;
    background-color: #f8f9fa;
    border: 1px solid #e9ecef;
    border-radius: 8px;
    padding: 15px;
    font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
    font-size: 0.85rem;
    line-height: 1.45;
    color: #333;
    white-space: pre-wrap;
    word-break: break-all;
}

div[data-testid="stExpander"] details summary {
    font-weight: 700 !important;
    color: #1a1a1a !important;
}

/* Report Browser Styles */
.browser-container {
    padding: 1rem;
    background: #fdfbff;
    border-radius: 12px;
}

.category-container {
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
    margin-bottom: 20px;
}

.category-chip {
    padding: 6px 16px;
    border-radius: 30px;
    background: white;
    border: 1px solid #e5e7eb;
    color: #6b7280;
    font-size: 0.85rem;
    font-weight: 600;
    cursor: pointer;
    transition: all 0.2s;
}

.category-chip:hover {
    border-color: #8c28e8;
    color: #8c28e8;
}

.category-chip.active {
    background: #8c28e8;
    color: white;
    border-color: #8c28e8;
}

.browser-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(220px, 1fr));
    gap: 15px;
    max-height: 450px;
    overflow-y: auto;
    padding: 10px;
    border: 1px solid #f3e8ff;
    border-radius: 8px;
    background: #fff;
}

.browser-card {
    padding: 12px;
    background: white;
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    cursor: pointer;
    transition: all 0.2s;
    font-size: 0.85rem;
    font-weight: 500;
    display: flex;
    align-items: center;
}

.browser-card:hover {
    border-color: #8c28e8;
    background: #fdfbff;
    transform: translateY(-2px);
}
</style>
"""

st.markdown(APP_CSS, unsafe_allow_html=True)


if genai is None:
    st.error("google-generativeai paketi kurulmamış. Lütfen terminalde şu komutu çalıştırın:\n`pip install google-generativeai`")
    st.stop()


@st.cache_data
def list_models():
    if genai is None:
        raise RuntimeError("google.generativeai paketi yüklenmedi.")

    api_key = st.session_state.get("api_key")
    if api_key:
        genai.configure(api_key=api_key)

    try:
        response = genai.list_models()

        if isinstance(response, types.GeneratorType):
            response = list(response)

        if isinstance(response, dict):
            model_list = response.get("models", [])
            return [m.get("name") for m in model_list if m.get("name")]
        if isinstance(response, list):
            return [
                m.get("name") if isinstance(m, dict)
                else (m.name if hasattr(m, "name") else str(m))
                for m in response
            ]
        return []
    except Exception as e:
        raise RuntimeError(f"Model listesi alınamadı: {e}") from e


@st.cache_data
def generate_ai_recommendation(query: str, reports: list, model_name: str) -> str:
    if model_name is None or model_name.strip() == "":
        model_name = "gemini-pro"

    try:
        model = genai.GenerativeModel(model_name)
        reports_text = "\n".join([
            f"- {r['name']} ({r['topic']}): {r['why']}"
            for r in reports
        ])

        prompt = f"""
        Kullanıcı sorgusu: "{query}"

        Önerilen raporlar:
        {reports_text}

        Lütfen Türkçe olarak, bu raporların kullanıcının ihtiyacını nasıl karşıladığını kısa ve öz bir şekilde açıklayınız (2-3 cümle).
        """

        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"AI açıklama alınamadı: {str(e)}"

def generate_pbi_ai_insight(metadata, model_name):
    """Interpret PBI metadata using AI, including forensically recovered data."""
    if not st.session_state.api_key: return "API Anahtarı eksik."
    try:
        model = genai.GenerativeModel(model_name)
        # Summarize metadata for AI, including new forensic fields
        summary = {
            "type": metadata.get("type"),
            "tables": [t["name"] for t in metadata.get("tables", [])[:40]],
            "measures": [m.get("name") for m in metadata.get("measures", [])[:20]],
            "m_queries": list(metadata.get("m_queries", {}).keys()),
            "forensic_snippets": metadata.get("forensic_matches", [])[:10], # Send some raw snippets for logic detection
            "page_count": len(metadata.get("pages", [])),
            "warnings": metadata.get("warnings", [])
        }
        
        prompt = f"""
        Aşağıdaki Power BI rapor metaverisini (bazıları binary tarama ile kurtarılmıştır) analiz et ve Türkçe yorumla:
        {json.dumps(summary, indent=2)}
        
        Lütfen şunları kapsayan kısa bir özet yap:
        1. Raporun genel yapısı, tabloları ve karmaşıklığı.
        2. Kurtarılan DAX ve M-kod parçacıklarından yola çıkarak raporun ne tür bir mantık (hesaplama) içerdiği.
        3. Model optimizasyonu için teknik öneriler.
        4. Eğer 'warnings' varsa bunların teknik önemini açıkla.
        Kısa, profesyonel ve teknik bir dil kullan.
        """
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"AI yorumu oluşturulamadı: {e}"


# API Key Authentication with Google Gemini
def check_authentication():
    if "api_key" not in st.session_state:
        st.session_state.api_key = None
    if "active_page" not in st.session_state:
        st.session_state.active_page = "Dashboard"

    with st.sidebar:
        st.markdown(
            """
            <div class="sidebar-brand">
                <span class="material-icons-outlined" style="font-size: 1.8rem; margin-right: 12px; color: white;">analytics</span> Raportal Agent
            </div>
            """,
            unsafe_allow_html=True,
        )

        # Functional Sidebar Menu
        pages = [
            {"name": "Dashboard", "icon": "dashboard"},
            {"name": "Tüm Raporlar", "icon": "link"},
            {"name": "Dashboard Ajan", "icon": "smart_toy"},
            {"name": "Rapor İçerik Copilot", "icon": "manage_search"},
            {"name": "PBIT İndir", "icon": "download"},
            {"name": "Fix Sorgular", "icon": "history_edu"},
            {"name": "PBIX Analizi", "icon": "analytics"},
            {"name": "Metadata Listesi", "icon": "reorder"},
            {"name": "Hakkında", "icon": "info"},
        ]

        for p in pages:
            is_active = st.session_state.active_page == p["name"]
            btn_class = "sidebar-btn-active" if is_active else "sidebar-btn"
            if st.button(f"{p['name']}", key=f"nav_{p['name']}", help=p["name"], use_container_width=True, type="secondary"):
                st.session_state.active_page = p["name"]
                st.rerun()

        st.markdown("<br><hr style='border-color: rgba(255,255,255,0.2)'><br>", unsafe_allow_html=True)
        st.markdown("#### Gemini Bağlantısı")

        if not st.session_state.api_key:
            api_key = st.text_input(
                "Google Gemini API Key",
                type="password",
                placeholder="https://aistudio.google.com",
            )

            if st.button("Bağlan", type="primary", use_container_width=True):
                if api_key.strip():
                    try:
                        genai.configure(api_key=api_key)
                        available_models = list_models()
                        if not available_models:
                            st.warning("Model listesi alınamadı. Manuel giriş gerekecek.")
                            st.session_state.api_key = api_key
                            st.session_state.available_models = []
                            st.session_state.selected_model = "gemini-1"
                            st.session_state.manual_model = "gemini-1"
                        else:
                            st.session_state.api_key = api_key
                            st.session_state.available_models = available_models
                            st.session_state.selected_model = available_models[0]
                            st.success("API Key doğrulandı.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ API Key Hatası: {str(e)}")
                else:
                    st.error("❌ Lütfen API Key girin")
        else:
            st.success("✅ API Key Doğrulandı")
            if st.button("Çıkış Yap / Sıfırla", use_container_width=True):
                st.session_state.api_key = None
                st.session_state.available_models = []
                st.rerun()


check_authentication()

if st.session_state.api_key:
    genai.configure(api_key=st.session_state.api_key)

    with st.sidebar:
        st.markdown("### Model Seçimi")
        model_options = st.session_state.get("available_models", [])

        if not model_options:
            model_options = ["gemini-1", "gemini-1.5", "gemini-1.5-pro", "gemini-2-lite"]

        with st.expander("Gelişmiş Seçenekler"):
            manual_model_input = st.text_input(
                "Manuel model adı",
                value=st.session_state.get("manual_model", ""),
                placeholder="Örn. gemini-1.5-pro",
            ).strip()
            if manual_model_input:
                st.session_state.manual_model = manual_model_input
                if manual_model_input not in model_options:
                    model_options.insert(0, manual_model_input)

        if "selected_model" in st.session_state and st.session_state.selected_model in model_options:
            default_index = model_options.index(st.session_state.selected_model)
        else:
            default_index = 0

        selected_model = st.selectbox("Kullanılacak model", model_options, index=default_index)
        st.session_state.selected_model = selected_model
        
        st.markdown("<br><hr style='border-color: rgba(255,255,255,0.2)'><br>", unsafe_allow_html=True)
        st.markdown("#### En Popüler Raporlar")
        
        try:
            # We already have usage_scores cached
            usage_data = load_usage_data()
            top_reports = sorted(usage_data.items(), key=lambda x: x[1], reverse=True)[:5]
            for name, count in top_reports:
                if st.button(f"{name}", key=f"quick_{name}", use_container_width=True):
                    st.session_state.sample_query = name
                    st.rerun()
        except:
            pass


def normalize_text(text: str) -> str:
    text = str(text)
    text = text.replace("\ufeff", "")
    text = text.lower().strip()

    replacements = {
        "ı": "i", "İ": "i",
        "ş": "s", "Ş": "s",
        "ğ": "g", "Ğ": "g",
        "ü": "u", "Ü": "u",
        "ö": "o", "Ö": "o",
        "ç": "c", "Ç": "c",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    # Keep only alphanumeric and spaces for tokenization
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).replace("\ufeff", "").strip() for c in df.columns]
    return df


def find_best_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    normalized_cols = {col: normalize_text(col) for col in df.columns}

    for candidate in candidates:
        candidate_n = normalize_text(candidate)

        for real_col, real_col_n in normalized_cols.items():
            if real_col_n == candidate_n:
                return real_col

        for real_col, real_col_n in normalized_cols.items():
            if candidate_n in real_col_n or real_col_n in candidate_n:
                return real_col

    return None


@st.cache_data
def load_usage_data():
    base_dir = Path(__file__).resolve().parent
    usage_file = base_dir / "raportal_kullanim.csv"
    if not usage_file.exists():
        return {}

    try:
        # Date;User;ReportName;UsageCount;Source1;Source2
        df_usage = pd.read_csv(usage_file, sep=";", encoding="utf-8-sig", header=None,
                               names=["Date", "User", "ReportName", "Usage", "S1", "S2"])
        
        # Aggregate usage per report (taking max usage as it seems cumulative or unique visits)
        usage_scores = df_usage.groupby("ReportName")["Usage"].max().to_dict()
        return usage_scores
    except Exception as e:
        st.warning(f"Kullanım verisi yüklenemedi: {e}")
        return {}


@st.cache_data
def load_metadata():
    base_dir = Path(__file__).resolve().parent

    preferred_names = [
        "db_exported_catalog_v2.csv",
        "db_exported_catalog.csv",
        "raportal_metadata_master.xlsx.csv",
        "pilot_rapor_metadata_clean.csv",
        "pilot_rapor_metadata.csv",
        "pilot_rapor_metadata.xlsx",
        "pilot_rapor_metadata.csv.xlsx",
    ]
    
    usage_scores = load_usage_data()

    file_path = None

    for name in preferred_names:
        p = base_dir / name
        if p.exists():
            file_path = p
            break

    if file_path is None:
        csv_files = list(base_dir.glob("*.csv"))
        xlsx_files = list(base_dir.glob("*.xlsx"))

        if csv_files:
            file_path = csv_files[0]
        elif xlsx_files:
            file_path = xlsx_files[0]
        else:
            raise FileNotFoundError("Klasörde okunacak .csv veya .xlsx dosyası bulunamadı.")

    if file_path.suffix.lower() == ".xlsx":
        df = pd.read_excel(file_path)
    else:
        try:
            df = pd.read_csv(file_path, sep=";", encoding="utf-8-sig")
        except UnicodeDecodeError:
            df = pd.read_csv(file_path, sep=";", encoding="cp1254")

    df = clean_columns(df)
    
    # "Silinen Raporlar" ve "Test" klasörlerini filtrele
    if "Klasör1" in df.columns:
        df = df[~df["Klasör1"].str.contains("Silinen Raporlar", na=False)]
        df = df[~df["Klasör1"].str.contains("test", case=False, na=False)]
    if "Path" in df.columns:
        df = df[~df["Path"].str.contains("Silinen Raporlar", na=False)]
        df = df[~df["Path"].str.contains("test", case=False, na=False)]
        
    # Join Usage Scores
    name_col = find_best_column(df, ["Name", "Rapor Adı", "RaporAdi"])
    if name_col:
        df["UsageScore"] = df[name_col].map(usage_scores).fillna(0)
    else:
        df["UsageScore"] = 0
        
    return df, file_path.name


def prepare_columns(df: pd.DataFrame):
    col_map = {
        "name": find_best_column(df, ["Name", "Rapor Adı", "RaporAdi"]),
        "topic": find_best_column(df, ["Ana Konu"]),
        "desc": find_best_column(df, ["Bu rapor neyi gösteriyor?", "Bu rapor neyi gosteriyor"]),
        "kpi": find_best_column(df, ["Ana KPI'lar", "Ana KPI’lar", "Ana KPIlar"]),
        "similar": find_best_column(df, ["Benzer Raporlar"]),
        "path": find_best_column(df, ["Path", "RaporYolu"]),
        "tip": find_best_column(df, ["Tip", "Type", "Rapor Tipi"]),
    }

    for key, col_name in col_map.items():
        if col_name is None:
            fallback_name = f"__{key}__"
            df[fallback_name] = ""
            col_map[key] = fallback_name

    return df, col_map


RAPORTAL_URL_MAPPING = {
    "K-Board": "https://raportal.kariyer.net/home/powerbi/K-Board?rs:embed=true",
    "Aylık Churn": "https://raportal.kariyer.net/home/powerbi/Satış/Portf%C3%B6y%20Performans/Aylık Churn",
    "Günlük Hedef Dashboard": "https://raportal.kariyer.net/home/powerbi/Gunluk%20Hedef%20Dashboard",
    "Basvuru Cevaplama Raporu (ISO) (R002)": "https://raportal.kariyer.net/home/report/M%C3%BC%C5%9Fteri%20%C4%B0li%C5%9Fkileri/Basvuru%20Cevaplama%20Raporu%20(ISO)%20(R002)"
}


def get_raportal_url(row, col_map):
    name = str(row.get(col_map["name"], ""))
    if name in RAPORTAL_URL_MAPPING:
        return RAPORTAL_URL_MAPPING[name]

    path = str(row.get(col_map["path"], "")).strip("/")
    tip = str(row.get(col_map["tip"], "Report")).lower()

    base_type = "powerbi" if "dashboard" in tip else "report"
    encoded_path = quote(path, safe="/")
    return f"https://raportal.kariyer.net/home/{base_type}/{encoded_path}"


def score_report(row, query: str, col_map: dict) -> int:
    query_n = normalize_text(query)
    tokens = [t for t in query_n.split() if len(t) > 1] # Skip single characters like 'k'
    
    if not tokens:
        tokens = query_n.split()

    score = 0
    name_val = normalize_text(row.get(col_map["name"], ""))
    
    # Exact full match bonus (Highest Priority)
    if query_n == name_val:
        score += 200
    elif query_n in name_val:
        score += 50

    rules = [
        (col_map["name"], 10),
        (col_map["topic"], 8),
        (col_map["kpi"], 5),
        (col_map["desc"], 3),
        (col_map["similar"], 2),
        (col_map["path"], 2),
    ]

    for col, weight in rules:
        value = normalize_text(row.get(col, ""))
        if not value:
            continue

        val_tokens = value.split()
        for token in tokens:
            if token in val_tokens:
                score += weight * 4 # Massive bonus for whole word match
            elif token in value:
                score += weight # Standard weight for substring match

    # Thematic keyword bonuses
    all_text = " ".join([name_val, normalize_text(row.get(col_map["topic"], "")), normalize_text(row.get(col_map["desc"], ""))])
    if "churn" in query_n and "churn" in all_text:
        score += 30
    if "risk" in query_n and ("risk" in all_text or "alarm" in all_text):
        score += 30
    if "kpi" in query_n and ("kpi" in all_text or "dashboard" in all_text):
        score += 30

    # Popularity Factor (Scale score by popularity to break ties and prioritize verified high-usage assets)
    usage = float(row.get("UsageScore", 0))
    if usage > 1000:
        score += 15
    elif usage > 100:
        score += 5

    return score


def search_reports(df: pd.DataFrame, query: str, col_map: dict):
    results = []

    for _, row in df.iterrows():
        score = score_report(row, query, col_map)

        if score > 0:
            results.append({
                "name": row.get(col_map["name"], ""),
                "topic": row.get(col_map["topic"], ""),
                "why": row.get(col_map["desc"], ""),
                "kpis": row.get(col_map["kpi"], ""),
                "similar": row.get(col_map["similar"], ""),
                "path": row.get(col_map["path"], ""),
                "usage_score": row.get("UsageScore", 0),
                "score": score,
            })

    results = sorted(results, key=lambda x: x["score"], reverse=True)
    return results[:3]


def safe_text(value) -> str:
    return html.escape(str(value)) if value not in (None, "") else "-"


def strip_html_tags(text: str) -> str:
    text = re.sub(r"<script[\s\S]*?</script>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"<style[\s\S]*?</style>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_kpi_signals(raw_text: str) -> list[str]:
    keywords = [
        "kpi", "gelir", "ciro", "satis", "hedef", "gercek", "oran", "adet", "maliyet", "kar",
        "churn", "yenileme", "donusum", "basvuru", "ilan", "portfoy", "tahsilat", "arpu"
    ]
    text = raw_text.lower()
    return [k for k in keywords if k in text][:12]


# ── PBIRS REST API ──────────────────────────────────────────────────────────
PBIRS_BASE = "https://raportal.kariyer.net"
PBIRS_API  = f"{PBIRS_BASE}/reports/api/v2.0"

def create_pbirs_session(username: str, password: str, domain: str = "KARIYER") -> requests.Session:
    """NTLM kimlik doğrulamalı PBIRS oturumu oluştur."""
    full_user = username if "\\" in username else f"{domain}\\{username}"
    session = requests.Session()
    session.auth = HttpNtlmAuth(full_user, password, send_cbt=False)
    session.verify = False
    return session


def render_ssrs_snapshot(session: requests.Session, report_path: str) -> bytes | None:
    """SSRS raporunu PNG olarak render et (klasik rendering URL)."""
    clean = report_path.lstrip("/")
    url = f"{PBIRS_BASE}/ReportServer?/{quote(clean)}&rs:Format=IMAGE&rs:Command=Render"
    try:
        resp = session.get(url, timeout=60)
        if resp.status_code == 200 and "image" in resp.headers.get("Content-Type", ""):
            return resp.content
        return None
    except Exception:
        return None


def render_pbi_snapshot(session: requests.Session, item_id: str) -> bytes | None:
    """Power BI raporunu ExportTo API ile PNG olarak al (async polling)."""
    export_url = f"{PBIRS_API}/PowerBIReports({item_id})/ExportTo"
    payload = {"format": "PNG", "powerBIReportConfiguration": {"pages": []}}
    try:
        resp = session.post(export_url, json=payload, timeout=30)
        if resp.status_code not in (202, 200):
            return None
        poll_url = resp.headers.get("Operation-Location") or resp.headers.get("Location")
        if not poll_url:
            return None
        for _ in range(20):
            time.sleep(3)
            pr = session.get(poll_url, timeout=30)
            if pr.status_code == 200:
                data = pr.json()
                status = data.get("status", "")
                if status == "Succeeded":
                    # PBIRS returns file list; fetch first file
                    files = data.get("reportName") or ""
                    file_url = poll_url.rstrip("/") + "/files/0"
                    fr = session.get(file_url, timeout=60)
                    if fr.status_code == 200:
                        return fr.content
                    return None
                elif status == "Failed":
                    return None
        return None
    except Exception:
        return None


def get_report_snapshot(session: requests.Session, report_path: str, item_id: str, type_id) -> bytes | None:
    """Rapor tipine göre doğru rendering metodunu çağır."""
    try:
        t = int(type_id)
    except (ValueError, TypeError):
        t = 0
    if t == 2:    # SSRS
        return render_ssrs_snapshot(session, report_path)
    elif t == 13: # Power BI
        return render_pbi_snapshot(session, str(item_id))
    else:
        # Bilinmeyen tip: önce SSRS dene
        return render_ssrs_snapshot(session, report_path)


def _is_gemini_quota_error(exc: Exception) -> bool:
    return "429" in str(exc) or "quota" in str(exc).lower() or "exceeded" in str(exc).lower()


# ── REPORT PLAYBOOKS ────────────────────────────────────────────────────────
REPORT_PLAYBOOKS = {
    "aylık churn": {
        "business_goal": "Müşteri kaybını izlemek, erken uyarı ve aksiyon mekanizması kurmak.",
        "expected_kpis": [
            "Churn oranı",
            "Pasif dahil churn oranı",
            "Yenileme tipi bazlı churn",
            "Paket bazlı churn",
            "Segment bazlı churn",
            "Gelir bazlı churn",
            "Müşteri bazlı churn skoru",
            "Alarm nedeni",
            "Kullanım oranı",
            "Temas bilgisi",
        ],
        "expected_story": "Üst seviyede churn trendi ve kırılımlar; alt seviyede aksiyon alınabilir müşteri listesi beklenir.",
    },
    "churn": {
        "business_goal": "Yenilememe riskini düşürmek ve elde tutmayı artırmak.",
        "expected_kpis": [
            "Churn oranı",
            "Riskli müşteri adedi",
            "Yenileme olasılığı",
            "Paket/segment kırılımı",
        ],
        "expected_story": "Trend + kırılım + müşteri aksiyon listesi kombinasyonu aranır.",
    },
}


def find_report_playbook(report_name: str) -> dict | None:
    n = normalize_text(report_name)
    for key, cfg in REPORT_PLAYBOOKS.items():
        if normalize_text(key) in n:
            return cfg
    return None


def build_rule_based_commentary(report_name: str, playbook: dict | None, kpi_hits: list[str]) -> str:
    if playbook:
        kpi_text = ", ".join(playbook.get("expected_kpis", [])[:8])
        hit_text = ", ".join(kpi_hits) if kpi_hits else "(canlı sayfada metinsel sinyal bulunamadı)"
        return (
            f"{report_name} raporu için beklenen iş amacı: {playbook.get('business_goal', '-')}. "
            f"Bu raporda öncelikli KPI alanları: {kpi_text}. "
            f"Yorumlama yaklaşımı: önce üst seviyede trendi oku, sonra segment/paket/yenileme tipi kırılımlarına in, "
            f"en son müşteri bazlı alarm listesi üzerinden aksiyon planı üret. "
            f"Canlı içerik sinyalleri: {hit_text}."
        )

    hit_text = ", ".join(kpi_hits) if kpi_hits else "belirgin KPI sinyali yok"
    return (
        f"{report_name} için metadata tabanlı yorum üretildi. "
        f"Önerilen okuma sırası: trend -> kırılım -> müşteri detayı. "
        f"Canlı içerik sinyalleri: {hit_text}."
    )


def analyze_dashboard_image_with_ai(image_bytes: bytes, model_name: str, user_question: str, report_name: str) -> str:
    """Analyze dashboard screenshot with Gemini multimodal model."""
    if not image_bytes:
        return "Görsel verisi boş."

    prompt = f"""
    Sen bir BI danışmanısın. Kullanıcı sorusu: {user_question}
    Rapor adı: {report_name}

    Gönderilen dashboard görselini incele ve Türkçe, profesyonel bir iş yorumu üret.
    Çıktı formatı:
    1) Raporun amacı (1-2 cümle)
    2) Tespit edilen ana KPI'lar (madde madde)
    3) Trend/segment/paket kırılımı bulguları
    4) Operasyonel aksiyon önerileri (kısa maddeler)
    5) Riskler ve takip edilmesi gereken alarm noktaları

    Varsayım uydurma; görselde net olmayan yerde "görselde net değil" de.
    """

    model = genai.GenerativeModel(model_name)
    response = model.generate_content(
        [
            prompt,
            {"mime_type": "image/png", "data": image_bytes},
        ]
    )
    return response.text if hasattr(response, "text") else str(response)


def build_executive_summary(report_name: str, rule_comment: str, ai_comment: str | None = None) -> str:
    parts = [f"Rapor: {report_name}", rule_comment]
    if ai_comment:
        parts.append(ai_comment)
    return "\n\n".join(parts)


def fetch_report_page_ntlm(url: str, username: str, password: str, domain: str = "KARIYER") -> tuple[int, str]:
    username = (username or "").strip()
    if not username:
        raise ValueError("Kullanici adi bos olamaz.")

    full_user = username if "\\" in username else f"{domain}\\{username}"
    session = requests.Session()
    session.auth = HttpNtlmAuth(full_user, password or "", send_cbt=False)
    response = session.get(url, timeout=20, verify=False, allow_redirects=True)
    return response.status_code, response.text


def render_report_content_copilot():
    st.markdown('<div class="section-title">🧠 Rapor İçerik Copilot</div>', unsafe_allow_html=True)
    st.caption(
        "Önce gerçek browser ile raporu arka planda açar, sekmeleri gezer, screenshot alır ve AI'a verir. "
        "Browser başarısız olursa PBIRS export fallback devreye girer."
    )

    # ── Metadata yükle ──────────────────────────────────────────────────────
    try:
        df_raw, _ = load_metadata()
        df_raw, col_map = prepare_columns(df_raw)
    except Exception as e:
        st.error(f"Metadata yüklenemedi: {e}")
        return

    df_raw = df_raw.copy()
    df_raw["URL"] = df_raw.apply(lambda row: get_raportal_url(row, col_map), axis=1)

    # ── Kimlik bilgileri ─────────────────────────────────────────────────────
    with st.expander("🔐 PBIRS Bağlantı Bilgileri", expanded=True):
        st.warning(
            "Şifre yalnızca runtime'da kullanılır; dosyaya yazılmaz. "
            "Girilen kullanıcı/şifre browser oturumuna NTLM cookie bootstrap için de uygulanır."
        )
        c1, c2, c3 = st.columns([2, 2, 1])
        with c1:
            domain_user = st.text_input("Kullanıcı adı", placeholder="esra.akinci")
        with c2:
            domain_password = st.text_input("Şifre", type="password")
        with c3:
            domain = st.text_input("Domain", value="KARIYER")

    # ── Rapor seçimi ─────────────────────────────────────────────────────────
    c4, c5 = st.columns([3, 1])
    with c4:
        report_name = st.selectbox(
            "Rapor Seç",
            options=sorted(df_raw[col_map["name"]].astype(str).unique().tolist()),
        )
    with c5:
        mode = st.selectbox("Mod", ["Tek Rapor", "Toplu Erişim Testi"])

    user_question = st.text_input(
        "AI'ya sor (görsel üzerinden yanıtlar)",
        value="Bu raporda hangi KPI'lar var, hangi kırılımlar verilmiş, ne gibi aksiyonlar alınabilir?",
    )

    use_browser_agent = st.checkbox(
        "Gerçek browser ile dashboard'u arka planda aç (önerilen)",
        value=True,
    )
    allow_export_fallback = st.checkbox(
        "Browser başarısızsa PBIRS export fallback dene",
        value=False,
    )

    # Manüel görsel yükleme — her iki yöntem de başarısız olursa son çare
    uploaded_image = st.file_uploader(
        "Yedek: Kendi aldığın ekran görüntüsünü yükle (sunucuya erişilemeyen durumlarda)",
        type=["png", "jpg", "jpeg"],
    )

    run = st.button("Bağlan ve Görsel Al → AI Yorumla", type="primary")
    if not run:
        return

    need_credentials = (mode == "Toplu Erişim Testi") or allow_export_fallback or (not use_browser_agent)
    pbirs_session = None
    if need_credentials:
        if not domain_user or not domain_password:
            st.error("Bu akış için kullanıcı adı ve şifre zorunlu.")
            return
        pbirs_session = create_pbirs_session(domain_user, domain_password, domain)

    # ══════════════════════════════════════════════════════════════════════════
    # TOPLU ERİŞİM TESTİ
    # ══════════════════════════════════════════════════════════════════════════
    if mode == "Toplu Erişim Testi":
        if pbirs_session is None:
            st.error("Toplu erişim testi için kimlik bilgileri gerekli.")
            return
        rows = []
        progress = st.progress(0, text="Raporlara bağlanılıyor…")
        batch = df_raw.head(30)
        for idx, (_, r) in enumerate(batch.iterrows()):
            try:
                resp = pbirs_session.get(r["URL"], timeout=15)
                code = resp.status_code
            except Exception:
                code = -1
            rows.append({
                "Rapor": r.get(col_map["name"], "-"),
                "Tip": r.get(col_map["tip"], "-"),
                "HTTP": code,
                "Erişim": "✅" if code == 200 else "❌",
                "URL": r["URL"],
            })
            progress.progress((idx + 1) / len(batch), text=f"{idx+1}/{len(batch)} rapor tarandı")

        result_df = pd.DataFrame(rows)
        col_a, col_b, col_c = st.columns(3)
        col_a.metric("Erişilebilen (200)", int((result_df["HTTP"] == 200).sum()))
        col_b.metric("Auth Hatası (401)", int((result_df["HTTP"] == 401).sum()))
        col_c.metric("Erişilemeyen", int((result_df["HTTP"] < 0).sum()))
        st.dataframe(
            result_df,
            use_container_width=True,
            column_config={"URL": st.column_config.LinkColumn("URL", display_text="🔗 Aç")},
        )
        return

    # ══════════════════════════════════════════════════════════════════════════
    # TEK RAPOR — GÖRSEL RENDER + AI ANALİZ
    # ══════════════════════════════════════════════════════════════════════════
    selected = df_raw[df_raw[col_map["name"]].astype(str) == str(report_name)].head(1)
    if selected.empty:
        st.error("Rapor bulunamadı.")
        return

    row = selected.iloc[0]
    target_url = str(row.get("URL", ""))
    report_path = str(row.get(col_map["path"], ""))
    item_id  = str(row.get("ItemID", ""))
    type_id  = row.get(col_map["tip"], 0)
    playbook = find_report_playbook(str(row.get(col_map["name"], "")))

    if pbirs_session is not None:
        with st.spinner("NTLM preflight kontrolü yapılıyor…"):
            try:
                pre = pbirs_session.get(target_url, timeout=20)
                st.caption(f"Preflight HTTP: {pre.status_code}")
                if pre.status_code == 401:
                    st.error(
                        "Kimlik doğrulama başarısız (401). "
                        "Bu durumda export fallback başarısız olur."
                    )
                    if not use_browser_agent:
                        return
            except Exception as pre_exc:
                st.warning(f"Preflight kontrolü tamamlanamadı: {pre_exc}")
    else:
        st.caption("Browser modu: persistent Playwright profile kullanılacak. İlk çalıştırmada açılan browser'da manuel giriş yapabilirsiniz.")

    # ── Rapor metadatası ─────────────────────────────────────────────────────
    st.markdown("---")
    m1, m2, m3 = st.columns(3)
    m1.metric("Tip", "Power BI" if str(type_id) == "13" else ("SSRS" if str(type_id) == "2" else str(type_id)))
    m2.metric("Path", report_path or "-")
    m3.metric("ItemID", item_id[:8] + "…" if len(item_id) > 8 else item_id or "-")

    # ── Öncelikli yol: gerçek browser ile oku ───────────────────────────────
    if use_browser_agent:
        with st.status("🤖 Dashboard browser agent çalışıyor…", expanded=True) as status:
            st.write("Raportal sayfası arka planda gerçek browser ile açılıyor…")
            try:
                outcome = _run_agent_in_thread(target_url, domain_user or "", domain_password or "", domain, st.session_state.get("api_key", ""))
            except Exception as browser_exc:
                outcome = {
                    "ok": False,
                    "analysis": None,
                    "data": None,
                    "error": str(browser_exc),
                }

            if isinstance(outcome, dict) and outcome.get("ok"):
                dashboard_data = outcome.get("data")
                analysis_text = outcome.get("analysis")
                status.update(label="✅ Browser agent tamamlandı", state="complete")
                if outcome.get("strategy"):
                    st.caption(f"Browser açılış stratejisi: {outcome.get('strategy')}")

                if dashboard_data and dashboard_data.pages:
                    st.markdown("### 📸 Dashboard Görselleri")
                    page_names = [p.tab_name for p in dashboard_data.pages]
                    if len(page_names) > 1:
                        ui_tabs = st.tabs(page_names)
                    else:
                        ui_tabs = [st.container()]

                    for ui_tab, page in zip(ui_tabs, dashboard_data.pages):
                        with ui_tab:
                            img_path = Path(page.screenshot_path)
                            if img_path.exists():
                                st.image(str(img_path), caption=f"{page.tab_name}", use_container_width=True)
                            with st.expander("Okunan içerik"):
                                if page.filters:
                                    st.write(f"**Filtreler:** {', '.join(page.filters)}")
                                if page.kpi_values:
                                    st.write(f"**KPI'lar:** {', '.join(page.kpi_values)}")
                                if page.visual_titles:
                                    st.write(f"**Görsel başlıkları:** {', '.join(page.visual_titles)}")
                                if page.visible_text:
                                    st.code(page.visible_text[:2500])

                if analysis_text:
                    st.markdown("### 🤖 AI Analiz")
                    st.markdown(analysis_text)
                return

            status.update(label="Browser agent başarısız, fallback devrede", state="error")
            browser_error_text = None
            if isinstance(outcome, dict):
                browser_error_text = outcome.get("error")
            if not browser_error_text:
                browser_error_text = "Bilinmeyen browser hatası (hata metni boş döndü)."

            st.error("Gerçek browser ile dashboard okunamadı.")
            with st.expander("Browser agent hata detayı", expanded=True):
                st.code(str(browser_error_text))

            if not allow_export_fallback:
                st.info("PBIRS export fallback kapalı. Önce browser hata detayını çözmemiz gerekiyor.")
                return
            st.warning("PBIRS export fallback denenecek.")

    # ── Sunucudan görsel al ───────────────────────────────────────────────────
    snapshot_bytes: bytes | None = None

    if pbirs_session is None:
        st.info("PBIRS export fallback için kimlik bilgileri verilmedi; browser analizi esas alınacak.")
        return

    with st.spinner("PBIRS sunucusundan rapor görseli render ediliyor…"):
        try:
            snapshot_bytes = get_report_snapshot(pbirs_session, report_path, item_id, type_id)
        except Exception as snap_err:
            st.warning(f"Snapshot alınamadı: {snap_err}")

    # Manüel yüklemeyi yedek olarak kullan
    if snapshot_bytes is None and uploaded_image is not None:
        snapshot_bytes = uploaded_image.getvalue()
        st.info("Sunucudan snapshot alınamadı; yüklediğin görsel kullanılıyor.")

    if snapshot_bytes is not None:
        st.markdown("### 📸 Rapor Görseli (sunucudan)")
        st.image(snapshot_bytes, use_container_width=True)
    else:
        st.error(
            "⚠️ Rapor görseli alınamadı.\n\n"
            "**Olası nedenler:**\n"
            "- Kimlik bilgileri hatalı (401)\n"
            "- Report Server RenderingExtension devre dışı\n"
            "- Power BI raporlarda ExportTo henüz tamamlanmadı (async)\n\n"
            "**Çözüm:** Raporun ekran görüntüsünü manuel olarak alıp yukarıdaki "
            "'Yedek' alanına yükleyebilirsin — AI o görseli okuyacak."
        )

    # ── AI görsel analizi ─────────────────────────────────────────────────────
    if snapshot_bytes is not None:
        if not st.session_state.get("api_key"):
            st.warning("AI analizi için soldan Gemini API key bağlayın.")
        else:
            model_name = st.session_state.get("selected_model", "gemini-2.0-flash")
            with st.spinner(f"AI raporu gözüyle okuyor ({model_name})…"):
                try:
                    ai_comment = analyze_dashboard_image_with_ai(
                        snapshot_bytes,
                        model_name,
                        user_question,
                        str(row.get(col_map["name"], report_name)),
                    )
                    st.markdown("### 🤖 AI Görsel Yorumu")
                    st.info(ai_comment)

                    summary = build_executive_summary(
                        str(row.get(col_map["name"], report_name)),
                        build_rule_based_commentary(
                            str(row.get(col_map["name"], "-")), playbook, []
                        ),
                        ai_comment,
                    )
                    with st.expander("📋 Yönetici Özeti"):
                        st.markdown(summary)

                except Exception as ai_err:
                    if _is_gemini_quota_error(ai_err):
                        st.error(
                            "**Gemini ücretsiz kota aşıldı (429).**\n\n"
                            "Yapılabilecekler:\n"
                            "1. [Google AI Studio](https://aistudio.google.com) üzerinden ücretli plana geç\n"
                            "2. Yarın tekrar dene (günlük limit sıfırlanır)\n"
                            "3. Farklı bir API key kullan (soldan 'Çıkış Yap → Sıfırla' ile yeni key gir)"
                        )
                    else:
                        st.error(f"AI analizi başarısız: {ai_err}")
    elif uploaded_image is None:
        # Görsel de yok, AI da çalışamaz — Playbook tabanlı özet göster
        if playbook:
            st.markdown("### Kural Bazlı İş Yorumu (görsel olmadan)")
            st.info(build_rule_based_commentary(
                str(row.get(col_map["name"], "-")), playbook, []
            ))
            st.caption(
                "Not: Bu yorum rapor adı ve iş kurallarına dayanmaktadır. "
                "Gerçek görsel alındığında AI otomatik olarak o görseli okuyacak."
            )


# --- DATABASE HELPERS ---

def get_bidb_connection(database="DWH"):
    """BIDB sunucusuna Windows Authentication ile bağlanmayı dener."""
    drivers = [
        '{ODBC Driver 18 for SQL Server}',
        '{ODBC Driver 17 for SQL Server}',
        '{ODBC Driver 13 for SQL Server}',
        '{SQL Server Native Client 11.0}',
        '{SQL Server}'
    ]
    
    server = "BIDB"
    # User's SSMS settings: Encrypt=Mandatory, TrustServerCertificate=Yes
    base_conn_str = f"Server={server};Database={database};Trusted_Connection=yes;Encrypt=yes;TrustServerCertificate=yes;"
    
    for driver in drivers:
        try:
            conn_str = f"Driver={driver};{base_conn_str}"
            conn = pyodbc.connect(conn_str, timeout=5)
            return conn
        except Exception:
            continue
    return None

def run_query_on_bidb(sql, database="DWH"):
    """Verilen SQL sorgusunu BIDB üzerinde çalıştırır ve DataFrame döndürür."""
    conn = get_bidb_connection(database=database)
    if not conn:
        raise ConnectionError(f"BIDB sunucusuna ({database} veritabanı) bağlanılamadı. Lütfen VPN veya ağ erişiminizi kontrol edin.")
    
    try:
        # Multi-statement support: pd.read_sql returns the first result set.
        # Adding SET NOCOUNT ON is crucial in the SQL template itself.
        df = pd.read_sql(sql, conn)
        
        # Safety: if read_sql returns None or a list (multi-result), normalize to DF
        if df is None:
            return pd.DataFrame()
        if isinstance(df, list):
            return df[-1] if len(df) > 0 else pd.DataFrame()
            
        return df
    except Exception as e:
        # Re-raise with a clear message
        raise RuntimeError(f"Sorgu yürütme hatası: {str(e)}") from e
    finally:
        conn.close()

def to_excel(df):
    """DataFrame'i indirilebilir Excel formatına (bytes) çevirir (Segoe UI Semibold & Purple Header)."""
    output = io.BytesIO()
    
    # Pandas ile Excel'e yaz (openpyxl motoruyla)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sonuclar')
        workbook = writer.book
        worksheet = workbook.active
        
        # Renk ve Font Tanımları
        purple_fill = PatternFill(start_color='8316B5', end_color='8316B5', fill_type='solid')
        header_font = Font(name='Segoe UI Semibold', size=10, bold=True, color='FFFFFF')
        body_font = Font(name='Segoe UI Semibold', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Kenarlık (Opsiyonel ama şık durur)
        thin_side = Side(style='thin', color='DDDDDD')
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Başlık Satırı Formatı (1. satır)
        for cell in worksheet[1]:
            cell.fill = purple_fill
            cell.font = header_font
            cell.alignment = center_align
            
        # Veri Satırları Formatı
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.border = border
        
        # Sütun Genişliklerini Ayarla (Auto-fit)
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = min(adjusted_width, 50) # Limit to 50
            
        # Filtre Ekle
        worksheet.auto_filter.ref = worksheet.dimensions

    return output.getvalue()


# --- UI COMPONENTS & PAGES ---

def load_query_templates():
    tpl_path = Path(__file__).resolve().parent / "query_templates.json"
    if tpl_path.exists():
        with open(tpl_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def render_report_browser(df, col_map):
    """Kategorize edilmiş rapor kataloğu tarayıcısı."""
    st.markdown('<div class="browser-container">', unsafe_allow_html=True)
    
    # Search within browser
    b_search = st.text_input("🔍 Katalogda isim veya içerik ara...", placeholder="Rapor adı veya KPI yazın...", label_visibility="collapsed")
    
    # Categories Chips
    topics = sorted([t for t in df[col_map["topic"]].unique() if t and t != "nan"])
    topics = ["Tümü"] + topics
    
    sel_topic = st.session_state.get("browser_topic", "Tümü")
    
    # Use columns for horizontal chip simulation (or just a selectbox for stability, but we promised chips)
    # Since Steamlit doesn't have native chips, we'll use a segmented control or a small selectbox with a label
    selected_category = st.pills("Kategori Seçin", topics, selection_mode="single", default="Tümü")
    
    # Filtering Logic
    filtered_df = df.copy()
    if selected_category and selected_category != "Tümü":
        filtered_df = filtered_df[filtered_df[col_map["topic"]] == selected_category]
    
    if b_search:
        filtered_df = filtered_df[
            filtered_df[col_map["name"]].str.contains(b_search, case=False, na=False) |
            filtered_df[col_map["desc"]].str.contains(b_search, case=False, na=False)
        ]
    
    st.markdown(f"**{len(filtered_df)}** rapor listeleniyor")
    
    # Display cards as a grid
    # We'll use Streamlit columns to create a grid
    cols_per_row = 3
    for i in range(0, len(filtered_df), cols_per_row):
        batch = filtered_df.iloc[i:i+cols_per_row]
        cols = st.columns(cols_per_row)
        for j, (idx, row) in enumerate(batch.iterrows()):
            with cols[j]:
                r_name = row[col_map["name"]]
                r_desc = row[col_map["desc"]]
                help_text = str(r_desc) if pd.notna(r_desc) else "Açıklama bulunamadı."
                
                # Custom selection button that looks like a card
                if st.button(f"📄 {r_name}", key=f"browser_{idx}", use_container_width=True, help=help_text):
                    st.session_state.sample_query = r_name
                    st.rerun()
    
    st.markdown('</div>', unsafe_allow_html=True)

def render_dashboard_header():
    st.markdown(
        """
        <div class="main-header">
            <div class="header-logo"><span class="material-icons-outlined" style="margin-right: 0;">bolt</span></div>
            <div class="header-title">Raportal AI Assistant</div>
        </div>
        <div class="welcome-text">
            <h1>Hoş Geldiniz 👋</h1>
            <p>Raportal Agent — Metadata destekli ve AI tabanlı akıllı rapor öneri asistanı</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_dashboard():
    render_dashboard_header()
    
    # Workflow Steps
    st.markdown(
        """
        <div class="workflow-container">
            <div class="workflow-label">ANALİZ AKIŞI</div>
            <div class="workflow-step"><span>1</span> İhtiyaç Tanımı</div>
            <div class="workflow-arrow">chevron_right</div>
            <div class="workflow-step"><span>2</span> Metadata Eşleşme</div>
            <div class="workflow-arrow">chevron_right</div>
            <div class="workflow-step"><span>3</span> AI Skorlama</div>
            <div class="workflow-arrow">chevron_right</div>
            <div class="workflow-step"><span>4</span> Rapor Önerisi</div>
            <div class="workflow-arrow">chevron_right</div>
            <div class="workflow-step"><span>5</span> Detaylı Analiz</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not st.session_state.api_key:
        st.info("👋 Sol menüden API anahtarını girerek devam edebilirsin.")
        st.stop()

    # Load Data
    try:
        df, source_name = load_metadata()
        df, col_map = prepare_columns(df)
    except Exception as e:
        st.error(f"Dosya yüklenemedi: {e}")
        st.stop()

    # Popular Reports Bar
    st.markdown('<div class="section-title">En Popüler Dashboardlar</div>', unsafe_allow_html=True)
    pop_cols = st.columns(5)
    top_5 = df.nlargest(5, "UsageScore")

    for i, (idx, prow) in enumerate(top_5.iterrows()):
        with pop_cols[i]:
            report_name = prow[col_map['name']]
            r_url = get_raportal_url(prow, col_map)
            st.link_button(f"📊 {report_name}", r_url, use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)
    selected_model = st.session_state.get("selected_model", "gemini-1")

    if "sample_query" not in st.session_state:
        st.session_state.sample_query = ""

    # --- SEARCH SECTION ---
    st.markdown('<div class="section-title">Rapor Kataloğu & Arama</div>', unsafe_allow_html=True)
    
    # NEW: Expandable Report Browser
    with st.expander("📚 Rapor Kataloğuna Gözat (200+ Rapor)", expanded=False):
        render_report_browser(df, col_map)
    
    st.markdown('<div class="query-card">', unsafe_allow_html=True)

    query = st.text_input(
        "İş ihtiyacını yaz (Veya yukarıdan bir rapor seç)",
        value=st.session_state.sample_query,
        placeholder="Örn: yenileme riski olan firmalari görmek istiyorum",
        help="Yazdığınız metin veya seçtiğiniz rapor başlığı AI tarafından analiz edilecektir."
    )

    search_clicked = st.button("Rapor Öner / Analiz Et", type="primary")
    st.markdown('</div>', unsafe_allow_html=True)

    if search_clicked or st.session_state.sample_query:
        q_to_search = query if query.strip() else st.session_state.sample_query
        if q_to_search.strip():
            matches = search_reports(df, q_to_search, col_map)
            if matches:
                with st.spinner("AI öneri açıklaması hazırlanıyor..."):
                    ai_explanation = generate_ai_recommendation(q_to_search, matches, selected_model)
                st.markdown('<div class="section-title">AI değerlendirmesi</div>', unsafe_allow_html=True)
                st.info(ai_explanation)
                st.markdown('<div class="section-title">Önerilen raporlar</div>', unsafe_allow_html=True)
                for i, m in enumerate(matches, start=1):
                    badge_list = [f'<span class="badge">{safe_text(m["topic"])}</span>']
                    if m.get("usage_score", 0) > 5000:
                        badge_list.append('<span class="badge badge-pop">🔥 En Çok Kullanılan</span>')
                    elif m.get("usage_score", 0) > 1000:
                        badge_list.append('<span class="badge badge-pop">⭐ Popüler</span>')
                    if m.get("score") is not None:
                        badge_list.append(f'<span class="badge">Skor: {safe_text(m["score"])}</span>')

                    badges_html = "".join(badge_list)
                    res_html = f"""
<div class="result-card">
<div class="result-title"><span class="material-icons-outlined" style="color: #8c28e8; margin-right: 8px;">description</span>{i}. {safe_text(m['name'])}</div>
<div style="margin-bottom:10px;">{badges_html}</div>
<div class="field-label">Neden Önerildi?</div><div class="field-value">{safe_text(m['why'])}</div>
<div class="field-label">Ana KPI'lar</div><div class="field-value">{safe_text(m['kpis'])}</div>
<div class="field-label">Rapor Yolu</div><div class="field-value" style="font-family: monospace; font-size: 0.85rem; color: #6b7280;">{safe_text(m['path'])}</div>
</div>
"""
                    st.markdown(res_html, unsafe_allow_html=True)

    st.markdown("<br><hr style='border-color: #e5e7eb;'><br>", unsafe_allow_html=True)
    st.markdown('<div class="section-title">Hızlı Senaryolar</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    scenarios = [
        {"id": "card1", "icon": "📉", "title": "Churn Analizi", "desc": "Müşteri kayıp oranı analizi.", "query": "churn oranini takip etmek istiyorum"},
        {"id": "card2", "icon": "⚠️", "title": "Risk Takibi", "desc": "Yenileme riski olan firmalar.", "query": "yenileme riski olan firmalari gormek istiyorum"},
        {"id": "card3", "icon": "🚀", "title": "Erken Yenileme", "desc": "Fırsat olan müşteriler.", "query": "erken yenileme firsatlarini gormek istiyorum"},
        {"id": "card4", "icon": "➕", "title": "Ek Ürün Fırsatı", "desc": "Çapraz satış analizleri.", "query": "ek urun kullanim ve yenileme durumunu gormek istiyorum"},
        {"id": "card5", "icon": "📊", "title": "Genel Performans", "desc": "Satış verimlilik raporları.", "query": "genel satis ve performans durumunu gormek istiyorum"},
        {"id": "card6", "icon": "📋", "title": "Metadata Listesi", "desc": "Tüm rapor metadataları.", "query": "ilgili raporlardaki metadata listesini gorebilir miyim"},
    ]

    for i, s in enumerate(scenarios):
        target_col = [col1, col2, col3][i % 3]
        with target_col:
            st.markdown(f'<div class="nav-card"><div class="card-num">{i+1}</div><div class="card-icon">{s["icon"]}</div><div class="card-title">{s["title"]}</div><div class="card-desc">{s["desc"]}</div></div>', unsafe_allow_html=True)
            if st.button("Seç", key=s["id"], use_container_width=True):
                st.session_state.sample_query = s["query"]
                st.rerun()

def render_fix_sorgular():
    st.markdown('<div class="section-title">Fix Sorgu Yönetimi</div>', unsafe_allow_html=True)
    
    templates = load_query_templates()
    if not templates:
        st.warning("Henüz yüklü sorgu şablonu bulunamadı.")
        return

    # --- TOP CARD: Selection & Parameters ---
    st.markdown('<div class="custom-card">', unsafe_allow_html=True)
    col_sel1, col_sel2 = st.columns([2, 3])
    with col_sel1:
        selected_tpl_key = st.selectbox("Sorgu Şablonu", options=list(templates.keys()))
        tpl = templates[selected_tpl_key]
    with col_sel2:
        st.caption("Şablon Açıklaması")
        st.info(tpl['description'])

    col_db1, col_db2, col_btn = st.columns([1.5, 1.5, 2])
    with col_db1:
        donem_in = st.text_input("Hedef Dönem", value="202603", help="Örn: 202603")
    with col_db2:
        target_db = st.text_input("Hedef Veritabanı", value="DWH", help="Varsayılan: DWH")
    with col_btn:
        st.write("") # Alignment
        st.write("") 
        run_bidb = st.button("🚀 BIDB Üzerinde Çalıştır (Canlı)", type="primary", use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

    formatted_sql = tpl['sql_template'].replace("{{donem}}", donem_in)
    
    # --- SQL ACCORDION ---
    with st.expander("🔍 Güncel SQL Sorgusunu Görüntüle / Detaylar"):
        st.markdown(f"Bu sorgu şu anda **BIDB.{target_db}** üzerinde çalışacak.")
        # Scrollable SQL box
        st.markdown(f'<div class="sql-scroll-box">{html.escape(formatted_sql)}</div>', unsafe_allow_html=True)
        st.caption("Bu sorguyu SSMS üzerinden manuel olarak da çalıştırabilirsiniz.")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # --- RESULTS AREA ---
    if run_bidb:
        with st.spinner(f"BIDB.{target_db} sunucusuna bağlanılıyor..."):
            try:
                start_time = time.time()
                df_live = run_query_on_bidb(formatted_sql, database=target_db)
                end_time = time.time()
                
                st.success(f"Sorgu başarıyla tamamlandı! ({round(end_time - start_time, 2)} sn)")
                
                # Metrics card
                st.markdown('<div class="custom-card">', unsafe_allow_html=True)
                m1, m2, m3 = st.columns(3)
                m1.metric("Toplam Satır", len(df_live))
                if "FinalCheckGelir" in df_live.columns:
                    m2.metric("Toplam Gelir", f"{df_live['FinalCheckGelir'].sum():,.2f} TL")
                if "TahsilatDurumu" in df_live.columns:
                    rate = (df_live['TahsilatDurumu'].sum() / len(df_live)) * 100 if len(df_live) > 0 else 0
                    m3.metric("Tahsilat Oranı", f"%{rate:,.1f}")
                
                st.dataframe(df_live.head(100), use_container_width=True, height=350)
                
                # Excel Download
                try:
                    excel_data = to_excel(df_live)
                    st.download_button(
                        label="📥 Sonucu Excel Olarak İndir (Segoe Font & Mor Başlık)",
                        data=excel_data,
                        file_name=f"FinalCheck_{donem_in}_{int(time.time())}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                except Exception as ex:
                    st.warning(f"Excel oluşturulamadı: {ex}")
                    csv = df_live.to_csv(index=False).encode('utf-8-sig')
                    st.download_button("📥 CSV İndir", data=csv, file_name=f"FinalCheck_{donem_in}.csv", mime="text/csv")
                
                st.markdown('</div>', unsafe_allow_html=True)

            except Exception as e:
                st.error(f"⚠️ BIDB Hatası: {str(e)}")
                st.info("İpucu: ODBC Driver yanısıra VPN bağlantınızın açık olduğundan emin olun.")

    # --- MANUAL UPLOAD SECTION ---
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown('<div class="section-title">Manuel Veri Kaynağı</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="custom-card">', unsafe_allow_html=True)
    st.markdown("#### Manuel Excel Yükle (Dışarıdan Alınan Dosyalar)")
    uploaded_file = st.file_uploader("Veya SSMS çıktısını (.xlsx) buraya sürükleyin", type=["xlsx", "xls"])
    
    if uploaded_file:
        try:
            df_upload = pd.read_excel(uploaded_file)
            st.success("Dosya başarıyla yüklendi!")
            st.dataframe(df_upload.head(10), use_container_width=True)
            
            if st.button("Veriyi Kaydet & Güncelle", type="primary"):
                save_path = Path(__file__).resolve().parent / "fixed_queries_data.csv"
                df_upload["Donem"] = donem_in
                df_upload["UploadTarihi"] = pd.Timestamp.now()
                
                if save_path.exists():
                    df_old = pd.read_csv(save_path)
                    df_old = df_old[df_old["Donem"].astype(str) != str(donem_in)]
                    df_final = pd.concat([df_old, df_upload], ignore_index=True)
                else:
                    df_final = df_upload
                df_final.to_csv(save_path, index=False, encoding="utf-8-sig")
                st.success(f"{donem_in} dönemi için veriler başarıyla güncellendi.")
        except Exception as e:
            st.error(f"Hata: {e}")
    st.markdown('</div>', unsafe_allow_html=True)

def render_pbix_analyzer():
    st.markdown('<div class="section-title">Power BI Analiz Fabrikası (Batch Engine)</div>', unsafe_allow_html=True)
    st.write("Birden fazla PBIX dosyasını otomatik PBIT'e dönüştürün ve rapor DNA'larını toplu olarak analiz edin.")
    
    # Session state for batch results
    if "batch_results" not in st.session_state:
        st.session_state.batch_results = {} # filename -> metadata_dict

    st.markdown('<div class="custom-card">', unsafe_allow_html=True)
    uploaded_files = st.file_uploader("Power BI Raporlarını Seçin (.pbix, .pbit, .bim)", 
                                     type=["pbix", "pbit", "bim"], 
                                     accept_multiple_files=True)
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_files:
        # Check for new files to clear old batch if needed or just add
        # We'll use a button to trigger the process to avoid heavy UI reruns
        st.markdown('<br>', unsafe_allow_html=True)
        col_btn, col_info = st.columns([1, 2])
        process_trigger = col_btn.button("🚀 Tümünü Dönüştür ve Analiz Et", type="primary", use_container_width=True)
        
        if process_trigger:
            groups = {} # base_name -> list of uploaded_files
            for f in uploaded_files:
                base = f.name.rsplit('.', 1)[0]
                if base not in groups: groups[base] = []
                groups[base].append(f)

            for base_name, files in groups.items():
                if base_name in st.session_state.batch_results: continue
                
                with st.spinner(f"Hibrit Analiz: {base_name}..."):
                    try:
                        master_metadata = None
                        temp_save_path = None
                        
                        # Process files
                        for f in files:
                            # Save to temp for potential Robot usage
                            local_path = os.path.join("temp_pbix", f.name)
                            with open(local_path, "wb") as tp:
                                tp.write(f.getbuffer())
                            
                            if f.name.lower().endswith(".pbit"):
                                master_metadata = parse_pbi_file(f.read(), f.name)
                                master_metadata["source_type"] = "Master (PBIT)"
                            elif f.name.lower().endswith(".pbix"):
                                temp_save_path = os.path.abspath(local_path)

                        pbix_file = next((f for f in files if f.name.lower().endswith(".pbix")), None)
                        if pbix_file and not (master_metadata and master_metadata.get("source_quality") == "High"):
                            # PBIX Scavenge
                            from pbi_parser import convert_to_pbit_bytes
                            pbit_payload = convert_to_pbit_bytes(pbix_file.read())
                            pbix_meta = parse_pbi_file(pbit_payload if pbit_payload else pbix_file.read(), pbix_file.name)
                            
                            if not master_metadata:
                                master_metadata = pbix_meta
                                master_metadata["source_type"] = "Hybrid (PBIX-Forensic)"
                            else:
                                if not master_metadata.get("pages") and pbix_meta.get("pages"):
                                    master_metadata["pages"] = pbix_meta["pages"]
                            
                            master_metadata["pbit_payload"] = pbit_payload
                            master_metadata["local_pbix_path"] = temp_save_path

                        if master_metadata:
                            st.session_state.batch_results[base_name] = master_metadata
                    except Exception as e:
                        st.session_state.batch_results[base_name] = {"status": "Hata", "error": str(e)}
            st.success("Hibrid tarama tamamlandı.")

        # --- HYBRID SUMMARY DASHBOARD ---
        if st.session_state.batch_results:
            st.markdown('<div class="section-title">Hibrid Rapor Envanteri</div>', unsafe_allow_html=True)
            summary_rows = []
            for name, meta in st.session_state.batch_results.items():
                if "error" in meta:
                    summary_rows.append({"Rapor": name, "Durum": "🔴 Hata", "PBIT Gerekli?": "NA"})
                    continue
                
                # Logic for PBIT Necessity
                pbit_req_badge = "✅ HAYIR" if not meta.get("pbit_required") else "⚠️ EVET"
                
                summary_rows.append({
                    "Rapor Adı": name,
                    "Tip": meta.get("source_type", meta.get("type")),
                    "Kalite": meta.get("source_quality", "N/A"),
                    "Tablo": len(meta.get("tables", [])),
                    "Ölçü": len(meta.get("measures", [])),
                    "İlişki": len(meta.get("relationships", [])),
                    "Veri Kaynağı": ", ".join(meta.get("data_sources", []))[:30] + "...",
                    "PBIT Gerekli mi?": pbit_req_badge
                })
            
            st.dataframe(pd.DataFrame(summary_rows), use_container_width=True)

            if st.button("🗑️ Tüm Analizleri Temizle"):
                st.session_state.batch_results = {}
                st.rerun()

            # --- INDIVIDUAL REPORT CARDS ---
            st.markdown('<div class="section-title">Bireysel Rapor DNA Kartları (Hibrid)</div>', unsafe_allow_html=True)
            for fname, metadata in st.session_state.batch_results.items():
                if "error" in metadata:
                    st.error(f"⚠️ {fname}: {metadata['error']}")
                    continue

                card_status = "✅ Tam Veri (DNA)" if metadata.get("source_quality") == "High" else "⚠️ Kısıtlı Veri (Robot Gerekebilir)"
                with st.expander(f"📊 {fname} — {card_status}", expanded=False):
                    # Robot and Download Bar
                    col_robot, col_dl = st.columns([1, 1])
                    
                    if metadata.get("pbit_required") and metadata.get("local_pbix_path"):
                        with col_robot:
                            if st.button(f"🤖 PBI Robotu Başlat (Derin Analiz)", key=f"robot_{fname}", type="primary"):
                                with st.spinner("PBI Robotu devreye giriyor... Power BI Desktop açılıyor..."):
                                    pbit_file = metadata["local_pbix_path"].replace(".pbix", ".pbit")
                                    success, msg = trigger_pbi_robot_export(metadata["local_pbix_path"], pbit_file)
                                    if success:
                                        st.success("PBIT Şablonu başarıyla üretildi! DNA analizi tazeleniyor...")
                                        # Parse the new PBIT
                                        if os.path.exists(pbit_file):
                                            with open(pbit_file, "rb") as pf:
                                                new_meta = parse_pbi_file(pf.read(), f"{fname}.pbit")
                                                # Update state
                                                new_meta["local_pbix_path"] = metadata["local_pbix_path"]
                                                st.session_state.batch_results[fname] = new_meta
                                                st.rerun()
                                    else:
                                        st.error(f"Robot Hatası: {msg}")

                    if metadata.get("pbit_payload"):
                        with col_dl:
                            st.download_button(f"📥 Otomatik Üretilen .pbit İndir", 
                                                 data=metadata["pbit_payload"], 
                                                 file_name=f"{fname}_raportal_template.pbit", 
                                                 mime="application/x-zip-compressed", key=f"dl_{fname}")
                    
                    m1, m2, m3, m4 = st.columns(4)
                    m1.metric("Tip", metadata.get("type", "Bilinmiyor"))
                    m2.metric("Tablo", len(metadata.get("tables", [])))
                    m3.metric("Ölçü", len(metadata.get("measures", [])))
                    m4.metric("Sayfa", len(metadata.get("pages", [])))

                    if metadata.get("forensic_matches"):
                        st.info("🔍 Adli Tarama: Binary bloklardan veri parçacıkları kurtarıldı.")

                    # DNA Tabs for this specific report
                    t_pages, t_tables, t_rels, t_measures, t_m, t_src = st.tabs([
                        "📄 Sayfalar", "📊 Tablalar", "🔗 İlişkiler", "📐 DAX Ölçüleri", "⚡ Power Query", "🔌 Kaynaklar"
                    ])
                    
                    with t_pages:
                        if metadata.get("pages"):
                            for p in metadata["pages"]: st.write(f"- {p['name']}")
                    
                    with t_tables:
                        if metadata.get("tables"):
                            for t in metadata["tables"]:
                                with st.expander(f"📁 {t['name']}"):
                                    st.write(", ".join(t['columns']))
                                    if t.get('partitions'):
                                        st.code(t['partitions'][0]['query'], language="powerquery")

                    with t_rels:
                        if metadata.get("relationships"):
                            st.table(pd.DataFrame(metadata["relationships"]))

                    with t_measures:
                        if metadata.get("measures"):
                            st.dataframe(pd.DataFrame(metadata["measures"]), use_container_width=True)
                        if metadata.get("forensic_matches"):
                            with st.expander("Ham Binary DAX"):
                                for match in metadata["forensic_matches"]: st.code(match)

                    with t_m:
                        if metadata.get("m_queries"):
                            for q_name, q_code in metadata["m_queries"].items():
                                with st.expander(f"⚡ {q_name}"): st.code(q_code, language="powerquery")

                    with t_src:
                        if metadata.get("data_sources"):
                            for ds in metadata["data_sources"]: st.code(ds)
    else:
        # Instructions
        st.markdown("""
        ### Fabrika Nasıl Çalışır?
        1. Listeye istediğiniz kadar `.pbix` dosyası ekleyin.
        2. **'Tümünü Analiz Et'** butonuyla süreci başlatın.
        3. Sistem her dosyayı PBIT'e çevirip DNA analizini yapar.
        4. Sonuçları özet tabloda ve rapor kartlarında inceleyin.
        """)

def render_all_reports():
    st.markdown('<div class="section-title">🔗 Tüm Raportal Bağlantıları</div>', unsafe_allow_html=True)
    st.write("Raportal'daki tüm raporlar ve dashboardlar — tıklanabilir linklerle erişin.")

    try:
        df_raw, source_name = load_metadata()
        df_raw, col_map = prepare_columns(df_raw)
    except Exception as e:
        st.error(f"Dosya yüklenemedi: {e}")
        return

    # Generate URLs for all rows
    df_raw["URL"] = df_raw.apply(lambda row: get_raportal_url(row, col_map), axis=1)

    # Stats
    total = len(df_raw)
    # Tip tespiti: numeric Type (DB) veya string Tip (master)
    tip_col = col_map["tip"]
    type_vals = df_raw[tip_col].dropna().unique().tolist()
    has_numeric = any(str(v).isdigit() for v in type_vals)

    if has_numeric:
        n_ssrs   = len(df_raw[df_raw[tip_col].astype(str) == "2"])
        n_pbi    = len(df_raw[df_raw[tip_col].astype(str) == "13"])
        n_other  = total - n_ssrs - n_pbi
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Toplam", total)
        c2.metric("🗒️ SSRS Raporu", n_ssrs)
        c3.metric("📊 Power BI", n_pbi)
        if n_other:
            c4.metric("Diğer", n_other)
    else:
        n_dashboard = len(df_raw[df_raw[tip_col].str.lower().str.contains("dashboard", na=False)])
        n_report    = len(df_raw[df_raw[tip_col].str.lower().str.contains("report", na=False)])
        n_kpi       = total - n_dashboard - n_report
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Toplam", total)
        c2.metric("📊 Dashboard", n_dashboard)
        c3.metric("🗒️ Report", n_report)
        if n_kpi:
            c4.metric("📌 KPI", n_kpi)

    # Link erişim durumu notu
    st.info("🔐 Linkler canlı ve geçerli — raportal.kariyer.net sunucusu erişilebilir durumda. "
            "Açmak için tarayıcıda KARIYER domain hesabınızla giriş yapmanız yeterli.")

    st.markdown("<br>", unsafe_allow_html=True)

    # Filters
    col_f1, col_f2, col_f3 = st.columns([2, 1, 1])
    with col_f1:
        search_q = st.text_input("🔍 Rapor adı ara...", placeholder="Rapor adı yazın...")
    with col_f2:
        folders = sorted(df_raw["Klasör1"].dropna().unique().tolist()) if "Klasör1" in df_raw.columns else []
        folder_filter = st.selectbox("Klasör", ["Tümü"] + folders)
    with col_f3:
        type_options = sorted(df_raw[col_map["tip"]].dropna().unique().tolist())
        type_filter = st.selectbox("Tip", ["Tümü"] + type_options)

    # Apply filters
    filtered = df_raw.copy()
    if search_q:
        filtered = filtered[filtered[col_map["name"]].str.contains(search_q, case=False, na=False)]
    if folder_filter != "Tümü" and "Klasör1" in filtered.columns:
        filtered = filtered[filtered["Klasör1"] == folder_filter]
    if type_filter != "Tümü":
        filtered = filtered[filtered[col_map["tip"]] == type_filter]

    st.markdown(f"**{len(filtered)}** rapor gösteriliyor")

    # Build display dataframe with rename
    rename_map = {
        col_map["name"]: "Rapor Adı",
        col_map["tip"]: "Tip",
        "UsageScore": "Kullanım",
        "URL": "Raportal Linki",
    }
    if "Klasör1" in filtered.columns:
        rename_map["Klasör1"] = "Klasör"
        display_cols = [col_map["name"], "Klasör1", col_map["tip"], "UsageScore", "URL"]
    else:
        display_cols = [col_map["name"], col_map["path"], col_map["tip"], "UsageScore", "URL"]

    display_cols = [c for c in display_cols if c in filtered.columns]
    display_df = filtered[display_cols].rename(columns=rename_map)
    # Arrow serialization için list içeren kolonları stringe dönüştür
    for col in display_df.columns:
        if display_df[col].apply(lambda x: isinstance(x, list)).any():
            display_df[col] = display_df[col].astype(str)

    st.dataframe(
        display_df,
        use_container_width=True,
        height=520,
        column_config={
            "Raportal Linki": st.column_config.LinkColumn("Raportal Linki", display_text="🔗 Aç"),
            "Kullanım": st.column_config.NumberColumn("Kullanım", format="%d"),
        },
    )

    # CSV download
    csv_data = filtered[display_cols].rename(columns=rename_map).to_csv(
        index=False, sep=";", encoding="utf-8-sig"
    ).encode("utf-8-sig")
    st.download_button(
        "📥 Tüm Linkleri CSV Olarak İndir",
        data=csv_data,
        file_name="raportal_links.csv",
        mime="text/csv",
        use_container_width=True,
    )


def render_metadata_list():
    st.markdown('<div class="section-title">Sistem Metadata Listesi</div>', unsafe_allow_html=True)
    df_all, _ = load_metadata()
    st.dataframe(df_all, use_container_width=True)

def _run_agent_in_thread(url: str, username: str, password: str, domain: str, api_key: str) -> dict:
    """Playwright async kodunu ayrı thread'de çalıştır ve yapılandırılmış sonuç döndür."""
    try:
        from dashboard_agent.browser_agent import RaportalBrowserAgent, AuthError, PageLoadError, PlaywrightNotInstalledError
        from dashboard_agent.dashboard_reader import DashboardReader
        from dashboard_agent.analyzer import analyze_dashboard as da_analyze
        from dashboard_agent.config import ANALYSIS_DIR
    except ImportError as e:
        return {
            "ok": False,
            "analysis": None,
            "data": None,
            "error": f"İçe aktarma hatası: {e}",
        }

    async def _inner():
        agent = RaportalBrowserAgent(username, password, domain)
        try:
            await agent.start()
            reader = DashboardReader(agent)
            data   = await reader.read_dashboard(url)
            # JSON kaydet
            safe_name = re.sub(r"[^\w\u00C0-\u017E\-]", "_", data.report_name).strip()[:80]
            json_path = ANALYSIS_DIR / f"{safe_name}.json"
            json_path.write_text(json.dumps(data.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
            analysis = da_analyze(data, api_key)
            md_path  = ANALYSIS_DIR / f"{safe_name}.md"
            md_path.write_text(analysis, encoding="utf-8")
            return {
                "ok": True,
                "analysis": analysis,
                "data": data,
                "strategy": getattr(agent, "startup_strategy", "unknown"),
                "error": None,
            }
        finally:
            await agent.close()

    result = {"value": None, "error": None}
    def thread_fn():
        try:
            if os.name == "nt":
                loop = asyncio.ProactorEventLoop()
                asyncio.set_event_loop(loop)
                try:
                    result["value"] = loop.run_until_complete(_inner())
                finally:
                    loop.close()
                    asyncio.set_event_loop(None)
            else:
                result["value"] = asyncio.run(_inner())
        except BaseException as e:
            # Bazı runtime/playwright hataları boş mesaj döndürebilir; traceback sakla.
            err_msg = str(e).strip() or repr(e)
            result["error"] = f"{err_msg}\n\n{traceback.format_exc()}"
    t = threading.Thread(target=thread_fn)
    t.start()
    t.join(timeout=240)
    if t.is_alive():
        return {
            "ok": False,
            "analysis": None,
            "data": None,
            "error": "Ajan zaman aşımına uğradı (240 sn). Browser süreci takılmış olabilir.",
        }
    if result["error"] is not None:
        return {
            "ok": False,
            "analysis": None,
            "data": None,
            "error": result["error"],
        }
    if not result["value"]:
        return {
            "ok": False,
            "analysis": None,
            "data": None,
            "error": "Ajan sonuç üretemedi.",
        }
    return result["value"]


def render_dashboard_agent():
    st.markdown('<div class="section-title">🤖 Dashboard Ajan</div>', unsafe_allow_html=True)
    st.caption(
        "Playwright ile raportal.kariyer.net'teki herhangi bir raporun sayfasını gerçek anlamda açar, "
        "her sekmenin screenshot'ını alır, DOM'daki görünür metinleri çeker ve "
        "Gemini ile **Türkçe iş yorumu** üretir. Hiçbir değişiklik yapmaz (read-only)."
    )

    # ── Playwright kurulum kontrolü ────────────────────────────────────────
    try:
        import playwright  # noqa
        pw_ok = True
    except ImportError:
        pw_ok = False

    if not pw_ok:
        st.error(
            "**Playwright kurulu değil.** Terminalde şunu çalıştır:\n\n"
            "```\npip install playwright\npy -m playwright install chromium\n```"
        )
        return

    # ── Kimlik bilgileri ────────────────────────────────────────────────────
    with st.expander("🔐 Bağlantı Bilgileri", expanded=True):
        st.warning("Şifre yalnızca analiz sırasında kullanılır; dosyaya yazılmaz.")
        ca, cb, cc = st.columns([2, 2, 1])
        with ca:
            ag_user   = st.text_input("Kullanıcı adı", placeholder="esra.akinci", key="ag_user")
        with cb:
            ag_pass   = st.text_input("Şifre", type="password", key="ag_pass")
        with cc:
            ag_domain = st.text_input("Domain", value="KARIYER", key="ag_domain")

    # ── Rapor URL girişi ───────────────────────────────────────────────────
    # Metadata'dan hızlı seçim veya manuel URL
    col_sel, col_or = st.columns([3, 1])
    with col_sel:
        try:
            df_meta, col_map = prepare_columns(load_metadata()[0])
            df_meta["URL"] = df_meta.apply(lambda r: get_raportal_url(r, col_map), axis=1)
            report_options = ["(manuel gir)"] + sorted(df_meta[col_map["name"]].astype(str).unique().tolist())
            selected_report = st.selectbox("Katalogdan seç", report_options, key="ag_report_sel")
        except Exception:
            selected_report = "(manuel gir)"
            df_meta = None

    manual_url = st.text_input(
        "Rapor URL'si",
        placeholder="https://raportal.kariyer.net/home/report/Model/...",
        key="ag_url",
    )

    # URL kaynağını belirle
    target_url = manual_url.strip()
    if selected_report and selected_report != "(manuel gir)" and not target_url:
        try:
            row = df_meta[df_meta[col_map["name"]].astype(str) == selected_report].iloc[0]
            target_url = row["URL"]
            st.caption(f"Hedef: {target_url}")
        except Exception:
            pass

    ag_question = st.text_input(
        "AI'ya özel soru (isteğe bağlı)",
        value="Bu raporda hangi KPI'lar var, hangi iş kararlarına destek veriyor?",
        key="ag_question",
    )

    run_btn = st.button("▶ Analizi Başlat", type="primary", key="ag_run")
    if not run_btn:
        st.info(
            "**Nasıl çalışır?**\n\n"
            "1. Kullanıcı adı, şifre ve domain gir\n"
            "2. Katalogdan rapor seç ya da URL'yi manuel yaz\n"
            "3. 'Analizi Başlat' düğmesine bas\n"
            "4. Agent gerçek bir browser açar, sayfayı görür, screenshot alır\n"
            "5. Gemini tüm sekme görsellerini ve DOM metnini okuyarak Türkçe iş yorumu üretir\n\n"
            "**Gereksinimler:** Playwright (`pip install playwright && py -m playwright install chromium`) "
            "ve geçerli Gemini API key (soldan bağlan)"
        )
        return

    if not ag_user or not ag_pass:
        st.error("Kullanıcı adı ve şifre zorunlu.")
        return
    if not target_url:
        st.error("Rapor URL'si girilmedi.")
        return
    if not st.session_state.get("api_key"):
        st.warning("AI yorumu için Gemini API key gerekli. Soldan bağlanın. Screenshot ve DOM verisi yine de alınacak.")

    api_key    = st.session_state.get("api_key", "")
    model_name = st.session_state.get("selected_model", "gemini-2.0-flash")

    # ── Analiz çalıştır ────────────────────────────────────────────────────
    with st.status("🤖 Dashboard Ajan çalışıyor…", expanded=True) as status:
        st.write("Browser başlatılıyor ve NTLM kimlik doğrulaması yapılıyor…")
        try:
            outcome = _run_agent_in_thread(target_url, ag_user, ag_pass, ag_domain, api_key)
        except Exception as exc:
            st.error(f"Ajan çalışırken hata: {exc}")
            status.update(label="Hata oluştu", state="error")
            return

        if not isinstance(outcome, dict):
            st.error("Ajan beklenmeyen formatta sonuç döndürdü.")
            status.update(label="Hata oluştu", state="error")
            return

        if not outcome.get("ok"):
            st.error(outcome.get("error") or "Ajan çalışırken bilinmeyen hata oluştu.")
            status.update(label="Hata oluştu", state="error")
            return

        analysis_text = outcome.get("analysis")
        dashboard_data = outcome.get("data")
        if dashboard_data is None:
            st.error("Ajan dashboard verisini oluşturamadı.")
            status.update(label="Hata oluştu", state="error")
            return
        if outcome.get("strategy"):
            st.caption(f"Açılış stratejisi: {outcome.get('strategy')}")
        status.update(label="✅ Analiz tamamlandı!", state="complete")

    # ── Sonuçları göster ───────────────────────────────────────────────────
    if dashboard_data and dashboard_data.pages:
        st.markdown("---")
        st.subheader(f"📊 {dashboard_data.report_name}")
        st.caption(f"{len(dashboard_data.pages)} sayfa/sekme tarandı")

        # Sekme başına screenshot + DOM özeti
        tab_names = [p.tab_name for p in dashboard_data.pages]
        if len(tab_names) > 1:
            tabs_ui = st.tabs(tab_names)
        else:
            tabs_ui = [st.container()]

        for tab_ui, page in zip(tabs_ui, dashboard_data.pages):
            with tab_ui:
                img_path = Path(page.screenshot_path)
                if img_path.exists():
                    st.image(str(img_path), caption=f"Screenshot — {page.tab_name}", use_container_width=True)
                else:
                    st.warning("Screenshot alınamadı.")
                with st.expander("DOM Özeti (filtreler, KPI'lar, görsel başlıkları)"):
                    if page.filters:
                        st.write("**Filtreler:**", ", ".join(page.filters))
                    if page.kpi_values:
                        st.write("**KPI Değerleri:**", ", ".join(page.kpi_values))
                    if page.visual_titles:
                        st.write("**Görsel Başlıkları:**", ", ".join(page.visual_titles))
                    if page.visible_text:
                        with st.expander("Görünür Metin (DOM)"):
                            st.code(page.visible_text[:3000])

    # ── AI Analiz metni ────────────────────────────────────────────────────
    if analysis_text:
        st.markdown("---")
        st.markdown("## 🤖 AI Analiz Çıktısı")
        st.markdown(analysis_text)

        # İndirme butonu
        st.download_button(
            "📄 Analizi İndir (.md)",
            data=analysis_text.encode("utf-8"),
            file_name=f"{(dashboard_data.report_name if dashboard_data else 'analiz')}_analiz.md",
            mime="text/markdown",
        )


def render_pbit_downloader():
    st.markdown('<div class="section-title">📥 PBIT İndir — PBIRS Toplu İndirici</div>', unsafe_allow_html=True)
    st.caption(
        "raportal.kariyer.net üzerindeki tüm Power BI raporlarını "
        "PBIT (şablon) veya PBIX olarak indirir. "
        "NTLM domain kimlik bilgilerin gerekir."
    )

    # ── Kimlik bilgileri ─────────────────────────────────────────────────────
    with st.expander("🔐 Bağlantı Bilgileri", expanded=True):
        st.warning("Şifre yalnızca indirme isteği sırasında kullanılır; dosyaya kaydedilmez.")
        c1, c2, c3 = st.columns([2, 2, 1])
        with c1:
            dl_user = st.text_input("Kullanıcı adı", placeholder="esra.akinci", key="pbit_user")
        with c2:
            dl_pass = st.text_input("Şifre", type="password", key="pbit_pass")
        with c3:
            dl_domain = st.text_input("Domain", value="KARIYER", key="pbit_domain")

    # ── Seçenekler ───────────────────────────────────────────────────────────
    col_a, col_b, col_c = st.columns([2, 1, 1])
    with col_a:
        dl_limit = st.number_input("Maksimum rapor sayısı (0 = tümü)", min_value=0, value=0, step=5)
    with col_b:
        dry_run = st.checkbox("Kuru çalıştır (indirme yapma, listele)", value=True)
    with col_c:
        prefer_pbit = st.checkbox("PBIT tercih et (yoksa PBIX)", value=True)

    run = st.button("Bağlan ve İndir", type="primary")
    if not run:
        st.info(
            "**Nasıl çalışır?**  \n"
            "1. Kullanıcı adı + şifreni gir  \n"
            "2. Önce *Kuru çalıştır* ile kaç rapor bulunduğunu gör  \n"
            "3. Onay verdikten sonra işareti kaldır ve tekrar bas  \n"
            "4. Dosyalar `pbit_downloads/` klasörüne kaydedilir"
        )
        return

    if not dl_user or not dl_pass:
        st.error("Kullanıcı adı ve şifre zorunlu.")
        return

    # ── PBIRS oturumu aç ─────────────────────────────────────────────────────
    pbirs_session = create_pbirs_session(dl_user, dl_pass, dl_domain)
    PBIRS_API_BASE = "https://raportal.kariyer.net/powerbi/api/v2.0"

    with st.spinner("Sunucuya bağlanılıyor…"):
        try:
            test = pbirs_session.get(f"{PBIRS_API_BASE}/PowerBIReports?$top=1", timeout=15)
            if test.status_code == 401:
                st.error("Kimlik doğrulama başarısız (401). Kullanıcı adı / şifre kontrol edin.")
                return
            elif test.status_code != 200:
                st.error(f"Sunucu hatası: HTTP {test.status_code}")
                return
            st.success(f"✓ Bağlantı başarılı — KARIYER\\{dl_user}")
        except Exception as e:
            st.error(f"Bağlantı hatası: {e}")
            return

    # ── Rapor listesi ─────────────────────────────────────────────────────────
    with st.spinner("Power BI raporları listeleniyor…"):
        try:
            resp = pbirs_session.get(f"{PBIRS_API_BASE}/PowerBIReports?$top=2000", timeout=30)
            reports = resp.json().get("value", [])
        except Exception as e:
            st.error(f"Rapor listesi alınamadı: {e}")
            return

    if not reports:
        st.warning("Hiç Power BI raporu bulunamadı.")
        return

    st.metric("Bulunan Power BI raporu", len(reports))

    if dl_limit and int(dl_limit) > 0:
        reports = reports[: int(dl_limit)]
        st.caption(f"(ilk {len(reports)} rapor işlenecek)")

    # Listeyi tablo olarak göster
    preview_rows = [
        {
            "Rapor Adı": r.get("Name", "-"),
            "Path": r.get("Path", "-"),
            "ID": (r.get("Id") or "")[:8] + "…",
        }
        for r in reports
    ]
    st.dataframe(preview_rows, use_container_width=True)

    if dry_run:
        st.info("Kuru çalıştır modu: dosya indirilmedi. İşareti kaldırıp tekrar bas.")
        return

    # ── Gerçek indirme ────────────────────────────────────────────────────────
    import re as _re
    from pathlib import Path as _Path

    def _safe(name):
        return _re.sub(r'[<>:"/\\|?*]', "_", name).strip()

    out_root = _Path("pbit_downloads")
    out_root.mkdir(parents=True, exist_ok=True)

    progress_bar = st.progress(0, text="İndirme başlıyor…")
    log_area     = st.empty()
    logs         = []
    ok = skip = err = 0

    for i, report in enumerate(reports, 1):
        name  = report.get("Name")  or "unknown"
        path  = report.get("Path")  or ""
        rid   = report.get("Id")    or ""

        rel       = path.lstrip("/").replace("/", "\\") if path else name
        out_dir   = out_root / _Path(rel).parent
        ext       = "pbit" if prefer_pbit else "pbix"
        out_file  = out_dir / f"{_safe(name)}.{ext}"

        progress_bar.progress(i / len(reports), text=f"[{i}/{len(reports)}] {name}")

        if out_file.exists():
            logs.append(f"– {name}: zaten var")
            skip += 1
            log_area.code("\n".join(logs[-30:]))
            continue

        out_dir.mkdir(parents=True, exist_ok=True)
        downloaded = False

        if prefer_pbit:
            # ExportTo PBIT (async)
            try:
                er = pbirs_session.post(
                    f"{PBIRS_API_BASE}/PowerBIReports({rid})/ExportTo",
                    json={"format": "PBIT", "powerBIReportConfiguration": {"pages": []}},
                    timeout=30,
                )
                if er.status_code == 200 and len(er.content) > 100:
                    out_file.write_bytes(er.content)
                    downloaded = True
                elif er.status_code in (202,):
                    poll = er.headers.get("Operation-Location") or er.headers.get("Location")
                    if poll:
                        for _ in range(40):
                            time.sleep(3)
                            pr = pbirs_session.get(poll, timeout=30)
                            if pr.status_code == 200:
                                st_val = pr.json().get("status", "")
                                if st_val == "Succeeded":
                                    fr = pbirs_session.get(poll.rstrip("/") + "/files/0", timeout=120)
                                    if fr.status_code == 200:
                                        out_file.write_bytes(fr.content)
                                        downloaded = True
                                    break
                                elif st_val == "Failed":
                                    break
            except Exception:
                pass

        if not downloaded:
            # PBIX fallback
            pbix_file = out_dir / f"{_safe(name)}.pbix"
            try:
                cr = pbirs_session.get(f"{PBIRS_API_BASE}/PowerBIReports({rid})/Content", timeout=120)
                if cr.status_code == 200 and len(cr.content) > 100:
                    pbix_file.write_bytes(cr.content)
                    downloaded = True
                    out_file = pbix_file
            except Exception:
                pass

        if downloaded:
            ok += 1
            logs.append(f"✓ {name}")
        else:
            err += 1
            logs.append(f"✗ {name}: indirilemedi")

        log_area.code("\n".join(logs[-30:]))

    progress_bar.empty()
    st.success(f"Tamamlandı — ✓ {ok} indirildi · – {skip} atlandı · ✗ {err} hata")
    st.info(f"Dosya konumu: `pbit_downloads/` ({out_root.resolve()})")


def render_about():
    st.markdown('<div class="section-title">Hakkında</div>', unsafe_allow_html=True)
    st.write("Raportal Agent v2.0 - Kariyer.net için özel olarak geliştirilmiştir.")

# --- MAIN PAGE ROUTING ---
if st.session_state.active_page == "Dashboard":
    render_dashboard()
elif st.session_state.active_page == "Tüm Raporlar":
    render_all_reports()
elif st.session_state.active_page == "Rapor İçerik Copilot":
    render_report_content_copilot()
elif st.session_state.active_page == "Dashboard Ajan":
    render_dashboard_agent()
elif st.session_state.active_page == "PBIT İndir":
    render_pbit_downloader()
elif st.session_state.active_page == "Fix Sorgular":
    render_fix_sorgular()
elif st.session_state.active_page == "PBIX Analizi":
    render_pbix_analyzer()
elif st.session_state.active_page == "Metadata Listesi":
    render_metadata_list()
elif st.session_state.active_page == "Hakkında":
    render_about()
