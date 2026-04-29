from pathlib import Path
import asyncio
import concurrent.futures
import html
import traceback
import pandas as pd
import re
import streamlit as st
import types
import io
import pyodbc
import json
import os
import sys
from datetime import datetime, timedelta

from pathlib import Path

# Fix: Ensure the virtual environment's libraries are always preferred
PROJ_ROOT = Path(__file__).parent.parent
VENV_LIB = PROJ_ROOT / ".venv" / "Lib" / "site-packages"
if VENV_LIB.exists() and str(VENV_LIB) not in sys.path:
    sys.path.insert(0, str(VENV_LIB))

import time
import threading
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from urllib.parse import quote
import requests
from requests_ntlm import HttpNtlmAuth
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pbi_parser import parse_pbi_file, convert_to_pbit_bytes
from pbi_robot_engine import trigger_pbi_robot_export
from dashboard_agent.config import GEMINI_MODEL
from dashboard_agent.analyzer import suggest_report_template
from dashboard_agent.data_context import get_real_world_entities
from dashboard_agent.history_manager import save_visionary_request, get_visionary_history, delete_visionary_request
try:
    from raportal_vision import RaportalVisionAgent
    VISION_ENABLED = True
except Exception as e:
    VISION_ENABLED = False
    VISION_ERROR = str(e)

# Create temp folder for Robot interaction
if not os.path.exists("temp_pbix"):
    os.makedirs("temp_pbix")

st.set_page_config(page_title="Raportal Recommendation Copilot", layout="wide")

try:
    import google.generativeai as genai
except ImportError:
    genai = None

def save_schedule(template_name, recipient):
    """Schedules a report for monthly automated delivery."""
    config_dir = Path(__file__).resolve().parent / "config"
    config_dir.mkdir(exist_ok=True)
    path = config_dir / "schedules.json"
    
    schedules = []
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                schedules = json.load(f)
        except: schedules = []
            
    # Check if exists
    for s in schedules:
        if s.get("template_name") == template_name and s.get("recipient") == recipient:
            return False, "Bu zamanlama zaten mevcut."
            
    schedules.append({
        "template_name": template_name,
        "recipient": recipient,
        "active": True,
        "last_run": "Henüz çalışmadı",
        "status": "Bekliyor",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })
    
    with open(path, "w", encoding="utf-8") as f:
        json.dump(schedules, f, indent=4)
    return True, "Zamanlama başarıyla kaydedildi."

# --- DATABASE HELPERS ---
def get_bidb_connection(server="bidb", database="Raportal"):
    """bidb sunucusuna Windows Authentication ile (SSMS Ayarlarıyla) bağlanır."""
    drivers = [
        '{ODBC Driver 18 for SQL Server}',
        '{ODBC Driver 17 for SQL Server}',
        '{SQL Server Native Client 11.0}',
        '{SQL Server}'
    ]
    
    # SSMS Ayarları: Encrypt=Mandatory (yes), Trust Server Certificate=True (yes)
    # Windows Auth: Trusted_Connection=yes
    base_params = f"Server={server};Database={database};Trusted_Connection=yes;Encrypt=yes;TrustServerCertificate=yes;"
    
    for driver in drivers:
        try:
            conn_str = f"Driver={driver};{base_params}"
            return pyodbc.connect(conn_str, timeout=10)
        except Exception:
            # Fallback: Encrypt kapatmayı dene (bazı sürücüler için)
            try:
                conn_str_alt = f"Driver={driver};Server={server};Database={database};Trusted_Connection=yes;Encrypt=no;"
                return pyodbc.connect(conn_str_alt, timeout=10)
            except:
                continue
    return None

def run_query_on_bidb(sql, server="bidb", database="Raportal"):
    """Verilen SQL sorgusunu bidb üzerinde çalıştırır."""
    conn = get_bidb_connection(server=server, database=database)
    if not conn:
        raise ConnectionError(f"[{server}] sunucusuna bağlanılamadı. Lütfen VPN kontrolü yapın.")
    try:
        return pd.read_sql(sql, conn)
    finally:
        conn.close()

def to_excel(df):
    """DataFrame'i Kariyer.net kurumsal renkleriyle (mor başlık) şık bir Excel'e dönüştürür."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sonuçlar')
        
        workbook  = writer.book
        worksheet = writer.sheets['Sonuçlar']
        
        # Stil tanımları
        header_fill = PatternFill(start_color='8C28E8', end_color='8C28E8', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, name='Segoe UI')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Başlık satırını boya ve filtre ekle
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Filtreleri aktif et
        worksheet.auto_filter.ref = worksheet.dimensions
        
        # Sütun genişliklerini içeriğe göre ayarla (Auto-fit)
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            # Minimum 12, maksimum 50 genişlik
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 50)
            
    return output.getvalue()
            
APP_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&display=swap');
@import url('https://fonts.googleapis.com/icon?family=Material+Icons+Outlined');

:root {
    --kariyer-purple: #8c28e8;
    --kariyer-purple-dark: #1e1b4b;
    --kariyer-indigo: #6366f1;
    --kariyer-bg: #f8fafc;
    --sidebar-bg: #1e1b4b;
}

html, body, [class*="css"] {
    font-family: 'Outfit', sans-serif;
}

.stApp {
    background-color: var(--kariyer-bg);
}

/* --- Sidebar Modernization --- */
section[data-testid="stSidebar"] {
    background-color: var(--sidebar-bg) !important;
}

.sidebar-brand {
    padding: 1.5rem 1rem;
    font-size: 1.3rem;
    font-weight: 800;
    color: white;
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 1rem;
}

.brand-icon {
    color: #38bdf8;
    font-size: 1.8rem !important;
}

/* Sidebar Nav Buttons */
section[data-testid="stSidebar"] div.stButton > button {
    background: transparent !important;
    color: rgba(255,255,255,0.6) !important;
    border: none !important;
    padding: 0.7rem 1rem !important;
    border-radius: 12px !important;
    font-weight: 500 !important;
    transition: all 0.2s ease !important;
    width: 100% !important;
}

/* Force inner content to the left */
section[data-testid="stSidebar"] div.stButton > button div[data-testid="stMarkdownContainer"] p {
    text-align: left !important;
    width: 100% !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important;
    gap: 12px !important;
}

/* Handle the button's internal flex container */
section[data-testid="stSidebar"] div.stButton > button > div {
    display: flex !important;
    justify-content: flex-start !important;
    align-items: center !important;
    width: 100% !important;
}

section[data-testid="stSidebar"] div.stButton > button:hover {
    background: rgba(255,255,255,0.05) !important;
    color: white !important;
}

/* Sidebar Active Button */
section[data-testid="stSidebar"] div.stButton > button[kind="primary"] {
    background: linear-gradient(90deg, rgba(140, 40, 232, 0.2) 0%, rgba(99, 102, 241, 0.2) 100%) !important;
    color: white !important;
    font-weight: 700 !important;
    box-shadow: inset 0 0 0 1px rgba(140, 40, 232, 0.4) !important;
}

/* Status Cards in Sidebar */
.gemini-status-card {
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 16px;
    padding: 1rem;
    margin: 1rem 0;
    backdrop-filter: blur(10px);
}

/* --- Page Heroes --- */
.hero-box {
    background: linear-gradient(135deg, #312e81 0%, #1e1b4b 100%);
    border-radius: 24px;
    padding: 2.5rem;
    color: white;
    position: relative;
    overflow: hidden;
    margin-bottom: 2rem;
    box-shadow: 0 20px 40px rgba(0,0,0,0.1);
}

.hero-box-light {
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 24px;
    padding: 2.5rem;
    color: #1e1b4b;
    margin-bottom: 2rem;
    box-shadow: 0 10px 20px rgba(0,0,0,0.02);
}

.hero-title {
    font-size: 2.8rem;
    font-weight: 900;
    letter-spacing: -1.5px;
    margin: 0;
    line-height: 1.1;
}

.hero-desc {
    font-size: 1.1rem;
    opacity: 0.8;
    margin-top: 10px;
    max-width: 600px;
}

/* --- Custom Cards --- */
.feature-card {
    background: white;
    border-radius: 24px;
    padding: 2rem;
    border: 1px solid #f1f5f9;
    box-shadow: 0 10px 25px rgba(0,0,0,0.03);
    height: 100%;
    display: flex;
    flex-direction: column;
}

.feature-icon-box {
    width: 48px;
    height: 48px;
    border-radius: 12px;
    display: flex;
    align-items: center;
    justify-content: center;
    margin-bottom: 1.5rem;
}

/* --- Form & Inputs --- */
.premium-form-box {
    background: white;
    border-radius: 24px;
    padding: 2rem;
    border: 1px solid #f1f5f9;
}

/* --- Buttons --- */
.stButton > button[kind="primary"] {
    background: linear-gradient(90deg, #8c28e8 0%, #6366f1 100%) !important;
    border: none !important;
    padding: 0.6rem 2rem !important;
    border-radius: 12px !important;
    font-weight: 700 !important;
    color: white !important;
    box-shadow: 0 10px 20px rgba(140, 40, 232, 0.2) !important;
    width: 100% !important;
}

/* --- Custom Toggles --- */
div[data-testid="stCheckbox"] {
    background: #f8fafc;
    padding: 10px 15px;
    border-radius: 12px;
    border: 1px solid #e2e8f0;
}

/* --- Insights Sidebar (Info) --- */
.info-sidebar-box {
    background: #f8fafc;
    border-radius: 16px;
    padding: 1.5rem;
    border: 1px solid #e2e8f0;
}

.info-step {
    display: flex;
    gap: 15px;
    margin-bottom: 20px;
}

.step-num {
    background: #eef2ff;
    color: #6366f1;
    width: 32px;
    height: 32px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 800;
    font-size: 0.8rem;
    flex-shrink: 0;
}

/* Status Badges & Pills */
.status-pill {
    padding: 4px 12px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 700;
    display: inline-flex;
    align-items: center;
    gap: 5px;
}
.status-pill-on { background: rgba(16, 185, 129, 0.1); color: #10b981; border: 1px solid rgba(16, 185, 129, 0.2); }
.status-pill-off { background: rgba(239, 68, 68, 0.1); color: #ef4444; border: 1px solid rgba(239, 68, 68, 0.2); }

.server-connect-container-wrapper {
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 16px;
    padding: 1rem;
    margin: 1rem 0;
    backdrop-filter: blur(10px);
}
/* --- Global Streamlit UI Cleanup --- */
[data-testid="stHeader"] {
    display: none;
}
.main .block-container {
    padding-top: 2rem !important;
}

</style>
"""


st.markdown(APP_CSS, unsafe_allow_html=True)


if genai is None:
    st.error("google-generativeai paketi kurulmamış. Lütfen terminalde şu komutu çalıştırın:\n`pip install google-generativeai`")
    st.stop()


@st.cache_data
def list_models():
    if genai is None:
        raise RuntimeError("google.generativeai paketi yüklenmedi.")

    api_key = st.session_state.get("api_key")
    if api_key:
        genai.configure(api_key=api_key)

    try:
        response = genai.list_models()

        if isinstance(response, types.GeneratorType):
            response = list(response)

        if isinstance(response, dict):
            model_list = response.get("models", [])
            return [m.get("name") for m in model_list if m.get("name")]
        if isinstance(response, list):
            return [
                m.get("name") if isinstance(m, dict)
                else (m.name if hasattr(m, "name") else str(m))
                for m in response
            ]
        return []
    except Exception as e:
        raise RuntimeError(f"Model listesi alınamadı: {e}") from e


@st.cache_data
def generate_ai_recommendation(query: str, reports: list, model_name: str) -> str:
    if model_name is None or model_name.strip() == "":
        model_name = "gemini-pro"

    try:
        model = genai.GenerativeModel(model_name)
        reports_text = "\n".join([
            f"- {r['name']} ({r['topic']}): {r['why']}"
            for r in reports
        ])

        prompt = f"""
        Kullanıcı sorgusu: "{query}"

        Önerilen raporlar:
        {reports_text}

        Lütfen Türkçe olarak, bu raporların kullanıcının ihtiyacını nasıl karşıladığını kısa ve öz bir şekilde açıklayınız (2-3 cümle).
        """

        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"AI açıklama alınamadı: {str(e)}"

async def _get_visual_inspiration_task(urls: list[str], username: str, password: str, domain: str) -> list[dict]:
    """Arka planda raporlara gidip ekran görüntüsü ve metin bilgisini alıp döner."""
    from dashboard_agent.browser_agent import RaportalBrowserAgent
    from dashboard_agent.config import SCREENSHOT_DIR
    
    agent = RaportalBrowserAgent(username, password, domain, headless=True)
    results = []
    try:
        await agent.start()
        for i, url in enumerate(urls[:1]): # Hız için şimdilik sadece en iyi eşleşen 1 rapor
            try:
                await agent.navigate(url)
                path = await agent.screenshot(f"vision_ins_{i}")
                text = await agent.get_visible_text()
                results.append({
                    "image_path": str(path),
                    "text": text[:1000]
                })
            except:
                continue
        await agent.close()
        return results
    except Exception:
        if 'agent' in locals(): await agent.close()
        return []

def run_visual_inspiration(urls: list[str], username: str, password: str, domain: str) -> list[dict]:
    """Async ilham tarama görevini sync dünyada çalıştırır."""
    import asyncio
    try:
        return asyncio.run(_get_visual_inspiration_task(urls, username, password, domain))
    except:
        return []

def generate_pbi_ai_insight(metadata, model_name):
    """Interpret PBI metadata using AI, including forensically recovered data."""
    if not st.session_state.api_key: return "API Anahtarı eksik."
    try:
        model = genai.GenerativeModel(model_name)
        # Summarize metadata for AI, including new forensic fields
        summary = {
            "type": metadata.get("type"),
            "tables": [t["name"] for t in metadata.get("tables", [])[:40]],
            "measures": [m.get("name") for m in metadata.get("measures", [])[:20]],
            "m_queries": list(metadata.get("m_queries", {}).keys()),
            "forensic_snippets": metadata.get("forensic_matches", [])[:10], # Send some raw snippets for logic detection
            "page_count": len(metadata.get("pages", [])),
            "warnings": metadata.get("warnings", [])
        }
        
        prompt = f"""
        Aşağıdaki Power BI rapor metaverisini (bazıları binary tarama ile kurtarılmıştır) analiz et ve Türkçe yorumla:
        {json.dumps(summary, indent=2)}
        
        Lütfen şunları kapsayan kısa bir özet yap:
        1. Raporun genel yapısı, tabloları ve karmaşıklığı.
        2. Kurtarılan DAX ve M-kod parçacıklarından yola çıkarak raporun ne tür bir mantık (hesaplama) içerdiği.
        3. Model optimizasyonu için teknik öneriler.
        4. Eğer 'warnings' varsa bunların teknik önemini açıkla.
        Kısa, profesyonel ve teknik bir dil kullan.
        """
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"AI yorumu oluşturulamadı: {e}"


def render_server_status_logic(suffix="pop"):
    # --- SQL CONNECTION POPOVER (Improved Styling) ---
    if "biportal_conn" not in st.session_state:
        st.session_state.biportal_conn = False
    
    is_conn = st.session_state.get("biportal_conn")
    
    # Status Row (Unified to avoid empty bubbles)
    status_pill_html = f'<div class="status-pill status-pill-on">ONLINE</div>' if is_conn else f'<div class="status-pill status-pill-off">OFFLINE</div>'
    
    st.markdown(
        f'<div class="server-connect-container-wrapper">'
        f'<div style="display: flex; align-items: center; justify-content: space-between;">'
        f'<div style="font-size: 0.85rem; font-weight: 700; color: rgba(255,255,255,0.8); display: flex; align-items: center; gap: 8px;">'
        f'<span class="material-icons-outlined" style="font-size: 1.2rem; color: #38bdf8;">storage</span> Sunucu'
        f'</div>'
        f'{status_pill_html}'
        f'</div>'
        f'</div>', 
        unsafe_allow_html=True
    )
    
    st.markdown('<div style="height: 10px;"></div>', unsafe_allow_html=True)
    
    # Action Button (Popover)
    with st.popover("🖥️ Bağlantı Ayarları", use_container_width=True):
        st.markdown("### 🔌 SQL Server Login")
        st.caption("BIDB veritabanına bağlanmak için bilgilerinizi girin.")
        
        sql_user = st.text_input("Windows User", value=st.session_state.get("sb_sql_user", "esra.akinci"), key=f"sb_sql_user_{suffix}")
        sql_pass = st.text_input("Password", value=st.session_state.get("sb_sql_pass", "Ea93934430."), type="password", key=f"sb_sql_pass_{suffix}")
        sql_srv  = st.text_input("Server", value=st.session_state.get("sb_sql_srv", "biportal"), key=f"sb_sql_srv_{suffix}")
        sql_db   = st.text_input("Database", value=st.session_state.get("sb_sql_db", "Raportal"), key=f"sb_sql_db_{suffix}")
        sql_dom  = st.text_input("Domain", value=st.session_state.get("sb_sql_domain", "KARIYER"), key=f"sb_sql_domain_{suffix}")
        
        if st.session_state.biportal_conn:
            if st.button("🛑 Bağlantıyı Kes", use_container_width=True, type="secondary", key=f"sb_sql_disconnect_{suffix}"):
                st.session_state.biportal_conn = False
                if "prof_catalog_df" in st.session_state:
                    del st.session_state.prof_catalog_df
                st.success("Bağlantı kesildi.")
                st.rerun()
        else:
            if st.button("🚀 Bağlantıyı Kur", use_container_width=True, type="primary", key=f"sb_sql_btn_{suffix}"):
                with st.spinner("Bağlanıyor..."):
                    try:
                        st.session_state.sb_sql_user = sql_user
                        st.session_state.sb_sql_pass = sql_pass
                        st.session_state.sb_sql_srv = sql_srv
                        st.session_state.sb_sql_db = sql_db
                        st.session_state.sb_sql_domain = sql_dom
                        
                        conn = get_bidb_connection(server=sql_srv, database=sql_db)
                        if conn:
                            st.session_state.biportal_conn = True
                            st.success("Bağlantı başarılı!")
                            st.rerun()
                        else:
                            st.error("Bağlantı kurulamadı.")
                    except Exception as e:
                        st.error(f"Hata: {e}")
    
    st.markdown('</div>', unsafe_allow_html=True)

# API Authentication with Google Gemini
def check_authentication():
    if "api_key" not in st.session_state:
        st.session_state.api_key = None
    if "active_page" not in st.session_state:
        st.session_state.active_page = "Anasayfa"

    with st.sidebar:
        st.markdown(
            """
            <div class="sidebar-brand">
                <span class="material-icons-outlined brand-icon">hub</span> Raportal Agent
            </div>
            """,
            unsafe_allow_html=True,
        )

        # Functional Sidebar Menu
        pages = [
            {"name": "Anasayfa"},
            {"name": "Raportal Insights Hub"},
            {"name": "Dashboard"},
            {"name": "Fix Sorgular"},
            {"name": "PBIX Analizi"},
            {"name": "Hakkında"},
        ]

        for p in pages:
            is_active = st.session_state.active_page == p["name"]
            btn_key = f"nav_{p['name']}"
            btn_type = "primary" if is_active else "secondary"
            
            # Icon mapping (Using Emojis for better compatibility)
            icon_map = {
                "Anasayfa": "🏠",
                "Raportal Insights Hub": "🧠",
                "Dashboard": "📊",
                "Fix Sorgular": "⚡",
                "PBIX Analizi": "📈",
                "Hakkında": "ℹ️"
            }
            icon = icon_map.get(p["name"], "🔵")
            
            # Render button with icon using HTML for better control if needed, 
            # but standard st.button with type is better for logic.
            # We'll use the CSS to make it look right.
            if st.button(p["name"], key=btn_key, use_container_width=True, type=btn_type, icon=icon):
                st.session_state.active_page = p["name"]
                st.rerun()

        st.markdown("<div style='height: 40px;'></div>", unsafe_allow_html=True)
        
        # --- GEMINI CONNECTION (BEAUTIFIED) ---
        if not st.session_state.get('api_key'):
            st.markdown("""
                <div style="background: linear-gradient(135deg, rgba(140, 40, 232, 0.15) 0%, rgba(99, 102, 241, 0.1) 100%); 
                            border: 1px solid rgba(140, 40, 232, 0.3); border-radius: 20px; padding: 1.5rem; margin: 1rem 0; 
                            box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
                    <div style="display: flex; align-items: center; gap: 10px; color: #c084fc; font-weight: 900; font-size: 1rem; margin-bottom: 10px;">
                        <span class="material-icons-outlined" style="font-size: 1.4rem;">auto_awesome</span>
                        AI Aktivasyonu
                    </div>
                    <div style="font-size: 0.85rem; color: rgba(255,255,255,0.7); margin-bottom: 0; line-height: 1.5;">
                        Platformun akıllı özelliklerini kullanabilmek için <b>Gemini API</b> anahtarınızı aşağıya tanımlayın.
                    </div>
                </div>
            """, unsafe_allow_html=True)
            
            api_key_input = st.text_input(
                "Gemini API Anahtarı",
                type="password",
                placeholder="Anahtarınızı buraya yapıştırın...",
                key="gemini_api_key_widget",
                help="Google AI Studio'dan aldığınız anahtar."
            )
            if api_key_input:
                new_key = api_key_input.strip()
                st.session_state.api_key = new_key
                st.rerun()
        else:
            st.markdown("""
                <div class="gemini-status-card">
                    <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 5px;">
                        <div style="display: flex; align-items: center; gap: 8px; color: #10b981; font-weight: 700; font-size: 0.85rem;">
                            <span class="material-icons-outlined" style="font-size: 1.1rem;">auto_awesome</span>
                            Gemini aktif
                        </div>
                        <span class="material-icons-outlined" style="color: #10b981; font-size: 1.2rem;">check_circle</span>
                    </div>
                    <div style="font-size: 0.7rem; color: rgba(255,255,255,0.4); margin-left: 26px;">API Key doğrulandı</div>
                </div>
            """, unsafe_allow_html=True)
            
            if st.session_state.get("available_models"):
                st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
                current_models = st.session_state.available_models
                default_idx = 0
                if st.session_state.get("selected_model") in current_models:
                    default_idx = current_models.index(st.session_state.selected_model)
                
                st.markdown("<div style='font-size: 0.8rem; font-weight: 700; color: rgba(255,255,255,0.6); margin-bottom: 5px;'>Aktif Model</div>", unsafe_allow_html=True)
                selected_model = st.selectbox(
                    "Model",
                    options=current_models,
                    index=default_idx,
                    key="sb_model_selector_new",
                    label_visibility="collapsed"
                )
                if selected_model != st.session_state.get("selected_model"):
                    st.session_state.selected_model = selected_model
                    st.rerun()

            col_a1, col_a2 = st.columns(2)
            with col_a1:
                if st.button("🔍 Modelleri Tara", use_container_width=True, key="scan_models_btn"):
                    with st.spinner("Taranıyor..."):
                        try:
                            genai.configure(api_key=st.session_state.api_key)
                            models = []
                            for m in genai.list_models():
                                if 'generateContent' in m.supported_generation_methods:
                                    models.append(m.name.replace('models/', ''))
                            st.session_state.available_models = sorted(models)
                            st.success(f"{len(models)} model bulundu.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Hata: {e}")
            with col_a2:
                if st.button("🛑 Sıfırla", use_container_width=True, key="reset_api_btn"):
                    st.session_state.api_key = None
                    st.session_state.available_models = []
                    st.rerun()

        st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
        # --- SIMULATION MODE ---
        st.markdown("""
            <div style="background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.1); border-radius: 12px; padding: 10px;">
                <div style="font-size: 0.75rem; color: #94a3b8; margin-bottom: 5px; font-weight: 700;">🔧 TEKNİK ÖNİZLEME</div>
            </div>
        """, unsafe_allow_html=True)
        st.checkbox("🚀 Simülasyon Modu", value=st.session_state.get("sim_mode", False), key="sim_mode_toggle", help="API hatası durumunda gerçekçi örneklerle çalışmanızı sağlar.")
        st.session_state.sim_mode = st.session_state.sim_mode_toggle

        st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
        # --- SQL CONNECTION SECTION ---
        render_server_status_logic(suffix="sidebar")





check_authentication()

if st.session_state.api_key:
    # Auto-clean the key every rerun for safety
    st.session_state.api_key = st.session_state.api_key.strip()
    genai.configure(api_key=st.session_state.api_key)
    
    # Auto-fetch models if not present
    if not st.session_state.get("available_models"):
        try:
            models = []
            for m in genai.list_models():
                if 'generateContent' in m.supported_generation_methods:
                    models.append(m.name.replace('models/', ''))
            st.session_state.available_models = sorted(models)
            if GEMINI_MODEL not in st.session_state.available_models:
                st.session_state.available_models.append(GEMINI_MODEL)
        except Exception:
            st.session_state.available_models = [GEMINI_MODEL]


def normalize_text(text: str) -> str:
    text = str(text)
    text = text.replace("\ufeff", "")
    text = text.lower().strip()

    replacements = {
        "ı": "i", "İ": "i",
        "ş": "s", "Ş": "s",
        "ğ": "g", "Ğ": "g",
        "ü": "u", "Ü": "u",
        "ö": "o", "Ö": "o",
        "ç": "c", "Ç": "c",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    # Keep only alphanumeric and spaces for tokenization
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).replace("\ufeff", "").strip() for c in df.columns]
    return df


def find_best_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    normalized_cols = {col: normalize_text(col) for col in df.columns}

    for candidate in candidates:
        candidate_n = normalize_text(candidate)

        for real_col, real_col_n in normalized_cols.items():
            if real_col_n == candidate_n:
                return real_col

        for real_col, real_col_n in normalized_cols.items():
            if candidate_n in real_col_n or real_col_n in candidate_n:
                return real_col

    return None


@st.cache_data
def load_usage_data():
    base_dir = Path(__file__).resolve().parent
    usage_file = base_dir / "raportal_kullanim.csv"
    if not usage_file.exists():
        return {}

    try:
        # Date;User;ReportName;UsageCount;Source1;Source2
        df_usage = pd.read_csv(usage_file, sep=";", encoding="utf-8-sig", header=None,
                               names=["Date", "User", "ReportName", "Usage", "S1", "S2"])
        
        # Aggregate usage per report (taking max usage as it seems cumulative or unique visits)
        usage_scores = df_usage.groupby("ReportName")["Usage"].max().to_dict()
        return usage_scores
    except Exception as e:
        st.warning(f"Kullanım verisi yüklenemedi: {e}")
        return {}


@st.cache_data
def load_metadata():
    base_dir = Path(__file__).resolve().parent

    preferred_names = [
        "db_exported_catalog_v2.csv",
        "db_exported_catalog.csv",
        "raportal_metadata_master.xlsx.csv",
        "pilot_rapor_metadata_clean.csv",
        "pilot_rapor_metadata.csv",
        "pilot_rapor_metadata.xlsx",
        "pilot_rapor_metadata.csv.xlsx",
    ]
    
    usage_scores = load_usage_data()

    file_path = None

    for name in preferred_names:
        p = base_dir / name
        if p.exists():
            file_path = p
            break

    if file_path is None:
        csv_files = list(base_dir.glob("*.csv"))
        xlsx_files = list(base_dir.glob("*.xlsx"))

        if csv_files:
            file_path = csv_files[0]
        elif xlsx_files:
            file_path = xlsx_files[0]
        else:
            raise FileNotFoundError("Klasörde okunacak .csv veya .xlsx dosyası bulunamadı.")

    if file_path.suffix.lower() == ".xlsx":
        df = pd.read_excel(file_path)
    else:
        try:
            df = pd.read_csv(file_path, sep=";", encoding="utf-8-sig")
        except UnicodeDecodeError:
            df = pd.read_csv(file_path, sep=";", encoding="cp1254")

    df = clean_columns(df)
    
    # "Silinen Raporlar" ve "Test" klasörlerini filtrele
    if "Klasör1" in df.columns:
        df = df[~df["Klasör1"].str.contains("Silinen Raporlar", na=False)]
        df = df[~df["Klasör1"].str.contains("test", case=False, na=False)]
    if "Path" in df.columns:
        df = df[~df["Path"].str.contains("Silinen Raporlar", na=False)]
        df = df[~df["Path"].str.contains("test", case=False, na=False)]
        
    # Join Usage Scores
    name_col = find_best_column(df, ["Name", "Rapor Adı", "RaporAdi"])
    if name_col:
        df["UsageScore"] = df[name_col].map(usage_scores).fillna(0)
    else:
        df["UsageScore"] = 0
        
    return df, file_path.name


def prepare_columns(df: pd.DataFrame):
    col_map = {
        "name": find_best_column(df, ["Name", "Rapor Adı", "RaporAdi"]),
        "topic": find_best_column(df, ["Ana Konu"]),
        "desc": find_best_column(df, ["Bu rapor neyi gösteriyor?", "Bu rapor neyi gosteriyor"]),
        "kpi": find_best_column(df, ["Ana KPI'lar", "Ana KPI’lar", "Ana KPIlar"]),
        "similar": find_best_column(df, ["Benzer Raporlar"]),
        "path": find_best_column(df, ["Path", "RaporYolu"]),
        "tip": find_best_column(df, ["Tip", "Type", "Rapor Tipi"]),
    }

    for key, col_name in col_map.items():
        if col_name is None:
            fallback_name = f"__{key}__"
            df[fallback_name] = ""
            col_map[key] = fallback_name
    
    return df, col_map


RAPORTAL_URL_MAPPING = {
    "K-Board": "https://raportal.kariyer.net/home/powerbi/K-Board?rs:embed=true",
    "Aylık Churn": "https://raportal.kariyer.net/home/powerbi/Satış/Portf%C3%B6y%20Performans/Aylık Churn",
    "Günlük Hedef Dashboard": "https://raportal.kariyer.net/home/powerbi/Gunluk%20Hedef%20Dashboard",
    "Basvuru Cevaplama Raporu (ISO) (R002)": "https://raportal.kariyer.net/home/report/M%C3%BC%C5%9Fteri%20%C4%B0li%C5%9Fkileri/Basvuru%20Cevaplama%20Raporu%20(ISO)%20(R002)"
}


def get_raportal_url(row, col_map):
    name = str(row.get(col_map["name"], ""))
    if name in RAPORTAL_URL_MAPPING:
        return RAPORTAL_URL_MAPPING[name]

    path = str(row.get(col_map["path"], "")).strip("/")
    tip = str(row.get(col_map["tip"], "Report")).lower()

    base_type = "powerbi" if "dashboard" in tip else "report"
    encoded_path = quote(path, safe="/")
    return f"https://raportal.kariyer.net/home/{base_type}/{encoded_path}"


def score_report(row, query: str, col_map: dict) -> int:
    query_n = normalize_text(query)
    tokens = [t for t in query_n.split() if len(t) > 1] # Skip single characters like 'k'
    
    if not tokens:
        tokens = query_n.split()

    score = 0
    name_val = normalize_text(row.get(col_map["name"], ""))
    
    # Exact full match bonus (Highest Priority)
    if query_n == name_val:
        score += 200
    elif query_n in name_val:
        score += 50

    rules = [
        (col_map["name"], 10),
        (col_map["topic"], 8),
        (col_map["kpi"], 5),
        (col_map["desc"], 3),
        (col_map["similar"], 2),
        (col_map["path"], 2),
    ]

    for col, weight in rules:
        value = normalize_text(row.get(col, ""))
        if not value:
            continue

        val_tokens = value.split()
        for token in tokens:
            if token in val_tokens:
                score += weight * 4 # Massive bonus for whole word match
            elif token in value:
                score += weight # Standard weight for substring match

    # Thematic keyword bonuses
    all_text = " ".join([name_val, normalize_text(row.get(col_map["topic"], "")), normalize_text(row.get(col_map["desc"], ""))])
    if "churn" in query_n and "churn" in all_text:
        score += 30
    if "risk" in query_n and ("risk" in all_text or "alarm" in all_text):
        score += 30
    if "kpi" in query_n and ("kpi" in all_text or "dashboard" in all_text):
        score += 30

    # Popularity Factor (Scale score by popularity to break ties and prioritize verified high-usage assets)
    usage = float(row.get("UsageScore", 0))
    if usage > 1000:
        score += 15
    elif usage > 100:
        score += 5

    return score


def search_reports(df: pd.DataFrame, query: str, col_map: dict):
    results = []

    for _, row in df.iterrows():
        score = score_report(row, query, col_map)

        if score > 0:
            results.append({
                "name": row.get(col_map["name"], ""),
                "topic": row.get(col_map["topic"], ""),
                "why": row.get(col_map["desc"], ""),
                "kpis": row.get(col_map["kpi"], ""),
                "similar": row.get(col_map["similar"], ""),
                "path": row.get(col_map["path"], ""),
                "usage_score": row.get("UsageScore", 0),
                "score": score,
            })

    results = sorted(results, key=lambda x: x["score"], reverse=True)
    return results[:3]


def safe_text(value) -> str:
    return html.escape(str(value)) if value not in (None, "") else "-"


def strip_html_tags(text: str) -> str:
    text = re.sub(r"<script[\s\S]*?</script>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"<style[\s\S]*?</style>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_kpi_signals(raw_text: str) -> list[str]:
    keywords = [
        "kpi", "gelir", "ciro", "satis", "hedef", "gercek", "oran", "adet", "maliyet", "kar",
        "churn", "yenileme", "donusum", "basvuru", "ilan", "portfoy", "tahsilat", "arpu"
    ]
    text = raw_text.lower()
    return [k for k in keywords if k in text][:12]


# ── PBIRS REST API ──────────────────────────────────────────────────────────
PBIRS_BASE = "https://raportal.kariyer.net"
PBIRS_API  = f"{PBIRS_BASE}/reports/api/v2.0"

def create_pbirs_session(username: str, password: str, domain: str = "KARIYER") -> requests.Session:
    """NTLM kimlik doğrulamalı PBIRS oturumu oluştur."""
    full_user = username if "\\" in username else f"{domain}\\{username}"
    session = requests.Session()
    session.auth = HttpNtlmAuth(full_user, password, send_cbt=False)
    session.verify = False
    return session


def render_ssrs_snapshot(session: requests.Session, report_path: str) -> bytes | None:
    """SSRS raporunu PNG olarak render et (klasik rendering URL)."""
    clean = report_path.lstrip("/")
    url = f"{PBIRS_BASE}/ReportServer?/{quote(clean)}&rs:Format=IMAGE&rs:Command=Render"
    try:
        resp = session.get(url, timeout=60)
        if resp.status_code == 200 and "image" in resp.headers.get("Content-Type", ""):
            return resp.content
        return None
    except Exception:
        return None


def render_pbi_snapshot(session: requests.Session, item_id: str) -> bytes | None:
    """Power BI raporunu ExportTo API ile PNG olarak al (async polling)."""
    export_url = f"{PBIRS_API}/PowerBIReports({item_id})/ExportTo"
    payload = {"format": "PNG", "powerBIReportConfiguration": {"pages": []}}
    try:
        resp = session.post(export_url, json=payload, timeout=30)
        if resp.status_code not in (202, 200):
            return None
        poll_url = resp.headers.get("Operation-Location") or resp.headers.get("Location")
        if not poll_url:
            return None
        for _ in range(20):
            time.sleep(3)
            pr = session.get(poll_url, timeout=30)
            if pr.status_code == 200:
                data = pr.json()
                status = data.get("status", "")
                if status == "Succeeded":
                    # PBIRS returns file list; fetch first file
                    files = data.get("reportName") or ""
                    file_url = poll_url.rstrip("/") + "/files/0"
                    fr = session.get(file_url, timeout=60)
                    if fr.status_code == 200:
                        return fr.content
                    return None
                elif status == "Failed":
                    return None
        return None
    except Exception:
        return None


def get_report_snapshot(session: requests.Session, report_path: str, item_id: str, type_id) -> bytes | None:
    """Rapor tipine göre doğru rendering metodunu çağır."""
    try:
        t = int(type_id)
    except (ValueError, TypeError):
        t = 0
    if t == 2:    # SSRS
        return render_ssrs_snapshot(session, report_path)
    elif t == 13: # Power BI
        return render_pbi_snapshot(session, str(item_id))
    else:
        # Bilinmeyen tip: önce SSRS dene
        return render_ssrs_snapshot(session, report_path)


def _is_gemini_quota_error(exc: Exception) -> bool:
    return "429" in str(exc) or "quota" in str(exc).lower() or "exceeded" in str(exc).lower()


# ── REPORT PLAYBOOKS ────────────────────────────────────────────────────────
REPORT_PLAYBOOKS = {
    "aylık churn": {
        "business_goal": "Müşteri kaybını izlemek, erken uyarı ve aksiyon mekanizması kurmak.",
        "expected_kpis": [
            "Churn oranı",
            "Pasif dahil churn oranı",
            "Yenileme tipi bazlı churn",
            "Paket bazlı churn",
            "Segment bazlı churn",
            "Gelir bazlı churn",
            "Müşteri bazlı churn skoru",
            "Alarm nedeni",
            "Kullanım oranı",
            "Temas bilgisi",
        ],
        "expected_story": "Üst seviyede churn trendi ve kırılımlar; alt seviyede aksiyon alınabilir müşteri listesi beklenir.",
    },
    "churn": {
        "business_goal": "Yenilememe riskini düşürmek ve elde tutmayı artırmak.",
        "expected_kpis": [
            "Churn oranı",
            "Riskli müşteri adedi",
            "Yenileme olasılığı",
            "Paket/segment kırılımı",
        ],
        "expected_story": "Trend + kırılım + müşteri aksiyon listesi kombinasyonu aranır.",
    },
}


def find_report_playbook(report_name: str) -> dict | None:
    n = normalize_text(report_name)
    for key, cfg in REPORT_PLAYBOOKS.items():
        if normalize_text(key) in n:
            return cfg
    return None


def build_rule_based_commentary(report_name: str, playbook: dict | None, kpi_hits: list[str]) -> str:
    if playbook:
        kpi_text = ", ".join(playbook.get("expected_kpis", [])[:8])
        hit_text = ", ".join(kpi_hits) if kpi_hits else "(canlı sayfada metinsel sinyal bulunamadı)"
        return (
            f"{report_name} raporu için beklenen iş amacı: {playbook.get('business_goal', '-')}. "
            f"Bu raporda öncelikli KPI alanları: {kpi_text}. "
            f"Yorumlama yaklaşımı: önce üst seviyede trendi oku, sonra segment/paket/yenileme tipi kırılımlarına in, "
            f"en son müşteri bazlı alarm listesi üzerinden aksiyon planı üret. "
            f"Canlı içerik sinyalleri: {hit_text}."
        )

    hit_text = ", ".join(kpi_hits) if kpi_hits else "belirgin KPI sinyali yok"
    return (
        f"{report_name} için metadata tabanlı yorum üretildi. "
        f"Önerilen okuma sırası: trend -> kırılım -> müşteri detayı. "
        f"Canlı içerik sinyalleri: {hit_text}."
    )


def analyze_dashboard_image_with_ai(image_bytes: bytes, model_name: str, user_question: str, report_name: str) -> str:
    """Analyze dashboard screenshot with Gemini multimodal model."""
    if not image_bytes:
        return "Görsel verisi boş."

    prompt = f"""
    Sen bir BI danışmanısın. Kullanıcı sorusu: {user_question}
    Rapor adı: {report_name}

    Gönderilen dashboard görselini incele ve Türkçe, profesyonel bir iş yorumu üret.
    Çıktı formatı:
    1) Raporun amacı (1-2 cümle)
    2) Tespit edilen ana KPI'lar (madde madde)
    3) Trend/segment/paket kırılımı bulguları
    4) Operasyonel aksiyon önerileri (kısa maddeler)
    5) Riskler ve takip edilmesi gereken alarm noktaları

    Varsayım uydurma; görselde net olmayan yerde "görselde net değil" de.
    """

    model = genai.GenerativeModel(model_name)
    response = model.generate_content(
        [
            prompt,
            {"mime_type": "image/png", "data": image_bytes},
        ]
    )
    return response.text if hasattr(response, "text") else str(response)


def build_executive_summary(report_name: str, rule_comment: str, ai_comment: str | None = None) -> str:
    parts = [f"Rapor: {report_name}", rule_comment]
    if ai_comment:
        parts.append(ai_comment)
    return "\n\n".join(parts)


def fetch_report_page_ntlm(url: str, username: str, password: str, domain: str = "KARIYER") -> tuple[int, str]:
    username = (username or "").strip()
    if not username:
        raise ValueError("Kullanici adi bos olamaz.")

    full_user = username if "\\" in username else f"{domain}\\{username}"
    session = requests.Session()
    session.auth = HttpNtlmAuth(full_user, password or "", send_cbt=False)
    response = session.get(url, timeout=20, verify=False, allow_redirects=True)
    return response.status_code, response.text


def render_report_content_copilot():
    st.markdown('<div class="section-title">🧠 Rapor İçerik Copilot</div>', unsafe_allow_html=True)
    st.caption(
        "Önce gerçek browser ile raporu arka planda açar, sekmeleri gezer, screenshot alır ve AI'a verir. "
        "Browser başarısız olursa PBIRS export fallback devreye girer."
    )

    # ── Metadata yükle ──────────────────────────────────────────────────────
    try:
        df_raw, _ = load_metadata()
        df_raw, col_map = prepare_columns(df_raw)
    except Exception as e:
        st.error(f"Metadata yüklenemedi: {e}")
        return

    df_raw = df_raw.copy()
    df_raw["URL"] = df_raw.apply(lambda row: get_raportal_url(row, col_map), axis=1)

    # ── Kimlik bilgileri ─────────────────────────────────────────────────────
    with st.expander("🔐 PBIRS Bağlantı Bilgileri", expanded=True):
        st.warning(
            "Şifre yalnızca runtime'da kullanılır; dosyaya yazılmaz. "
            "Girilen kullanıcı/şifre browser oturumuna NTLM cookie bootstrap için de uygulanır."
        )
        c1, c2, c3 = st.columns([2, 2, 1])
        with c1:
            domain_user = st.text_input("Kullanıcı adı", placeholder="esra.akinci")
        with c2:
            domain_password = st.text_input("Şifre", type="password")
        with c3:
            domain = st.text_input("Domain", value="KARIYER")

    # ── Rapor seçimi ─────────────────────────────────────────────────────────
    c4, c5 = st.columns([3, 1])
    with c4:
        report_name = st.selectbox(
            "Rapor Seç",
            options=sorted(df_raw[col_map["name"]].astype(str).unique().tolist()),
        )
    with c5:
        mode = st.selectbox("Mod", ["Tek Rapor", "Toplu Erişim Testi"])

    user_question = st.text_input(
        "AI'ya sor (görsel üzerinden yanıtlar)",
        value="Bu raporda hangi KPI'lar var, hangi kırılımlar verilmiş, ne gibi aksiyonlar alınabilir?",
    )

    use_browser_agent = st.checkbox(
        "Gerçek browser ile dashboard'u arka planda aç (önerilen)",
        value=True,
    )
    allow_export_fallback = st.checkbox(
        "Browser başarısızsa PBIRS export fallback dene",
        value=False,
    )

    # Manüel görsel yükleme — her iki yöntem de başarısız olursa son çare
    uploaded_image = st.file_uploader(
        "Yedek: Kendi aldığın ekran görüntüsünü yükle (sunucuya erişilemeyen durumlarda)",
        type=["png", "jpg", "jpeg"],
    )

    run = st.button("Bağlan ve Görsel Al → AI Yorumla", type="primary")
    if not run:
        return

    need_credentials = (mode == "Toplu Erişim Testi") or allow_export_fallback or (not use_browser_agent)
    pbirs_session = None
    if need_credentials:
        if not domain_user or not domain_password:
            st.error("Bu akış için kullanıcı adı ve şifre zorunlu.")
            return
        pbirs_session = create_pbirs_session(domain_user, domain_password, domain)

    # ══════════════════════════════════════════════════════════════════════════
    # TOPLU ERİŞİM TESTİ
    # ══════════════════════════════════════════════════════════════════════════
    if mode == "Toplu Erişim Testi":
        if pbirs_session is None:
            st.error("Toplu erişim testi için kimlik bilgileri gerekli.")
            return
        rows = []
        progress = st.progress(0, text="Raporlara bağlanılıyor…")
        batch = df_raw.head(30)
        for idx, (_, r) in enumerate(batch.iterrows()):
            try:
                resp = pbirs_session.get(r["URL"], timeout=15)
                code = resp.status_code
            except Exception:
                code = -1
            rows.append({
                "Rapor": r.get(col_map["name"], "-"),
                "Tip": r.get(col_map["tip"], "-"),
                "HTTP": code,
                "Erişim": "✅" if code == 200 else "❌",
                "URL": r["URL"],
            })
            progress.progress((idx + 1) / len(batch), text=f"{idx+1}/{len(batch)} rapor tarandı")

        result_df = pd.DataFrame(rows)
        col_a, col_b, col_c = st.columns(3)
        col_a.metric("Erişilebilen (200)", int((result_df["HTTP"] == 200).sum()))
        col_b.metric("Auth Hatası (401)", int((result_df["HTTP"] == 401).sum()))
        col_c.metric("Erişilemeyen", int((result_df["HTTP"] < 0).sum()))
        st.dataframe(
            result_df,
            use_container_width=True,
            column_config={"URL": st.column_config.LinkColumn("URL", display_text="🔗 Aç")},
        )
        return

    # ══════════════════════════════════════════════════════════════════════════
    # TEK RAPOR — GÖRSEL RENDER + AI ANALİZ
    # ══════════════════════════════════════════════════════════════════════════
    selected = df_raw[df_raw[col_map["name"]].astype(str) == str(report_name)].head(1)
    if selected.empty:
        st.error("Rapor bulunamadı.")
        return

    row = selected.iloc[0]
    target_url = str(row.get("URL", ""))
    report_path = str(row.get(col_map["path"], ""))
    item_id  = str(row.get("ItemID", ""))
    type_id  = row.get(col_map["tip"], 0)
    playbook = find_report_playbook(str(row.get(col_map["name"], "")))

    if pbirs_session is not None:
        with st.spinner("NTLM preflight kontrolü yapılıyor…"):
            try:
                pre = pbirs_session.get(target_url, timeout=20)
                st.caption(f"Preflight HTTP: {pre.status_code}")
                if pre.status_code == 401:
                    st.error(
                        "Kimlik doğrulama başarısız (401). "
                        "Bu durumda export fallback başarısız olur."
                    )
                    if not use_browser_agent:
                        return
            except Exception as pre_exc:
                st.warning(f"Preflight kontrolü tamamlanamadı: {pre_exc}")
    else:
        st.caption("Browser modu: persistent Playwright profile kullanılacak. İlk çalıştırmada açılan browser'da manuel giriş yapabilirsiniz.")

    # ── Rapor metadatası ─────────────────────────────────────────────────────
    st.markdown("---")
    m1, m2, m3 = st.columns(3)
    m1.metric("Tip", "Power BI" if str(type_id) == "13" else ("SSRS" if str(type_id) == "2" else str(type_id)))
    m2.metric("Path", report_path or "-")
    m3.metric("ItemID", item_id[:8] + "…" if len(item_id) > 8 else item_id or "-")

    # ── Öncelikli yol: gerçek browser ile oku ───────────────────────────────
    if use_browser_agent:
        with st.status("🤖 Dashboard browser agent çalışıyor…", expanded=True) as status:
            st.write("Raportal sayfası arka planda gerçek browser ile açılıyor…")
            try:
                outcome = _run_agent_in_thread(target_url, domain_user or "", domain_password or "", domain, st.session_state.get("api_key", ""))
            except Exception as browser_exc:
                outcome = {
                    "ok": False,
                    "analysis": None,
                    "data": None,
                    "error": str(browser_exc),
                }

            if isinstance(outcome, dict) and outcome.get("ok"):
                dashboard_data = outcome.get("data")
                analysis_text = outcome.get("analysis")
                status.update(label="✅ Browser agent tamamlandı", state="complete")
                if outcome.get("strategy"):
                    st.caption(f"Browser açılış stratejisi: {outcome.get('strategy')}")

                if dashboard_data and dashboard_data.pages:
                    st.markdown("### 📸 Dashboard Görselleri")
                    page_names = [p.tab_name for p in dashboard_data.pages]
                    if len(page_names) > 1:
                        ui_tabs = st.tabs(page_names)
                    else:
                        ui_tabs = [st.container()]

                    for ui_tab, page in zip(ui_tabs, dashboard_data.pages):
                        with ui_tab:
                            img_path = Path(page.screenshot_path)
                            if img_path.exists():
                                st.image(str(img_path), caption=f"{page.tab_name}", use_container_width=True)
                            with st.expander("Okunan içerik"):
                                if page.filters:
                                    st.write(f"**Filtreler:** {', '.join(page.filters)}")
                                if page.kpi_values:
                                    st.write(f"**KPI'lar:** {', '.join(page.kpi_values)}")
                                if page.visual_titles:
                                    st.write(f"**Görsel başlıkları:** {', '.join(page.visual_titles)}")
                                if page.visible_text:
                                    st.code(page.visible_text[:2500])

                if analysis_text:
                    st.markdown("### 🤖 AI Analiz")
                    st.markdown(analysis_text)
                return

            status.update(label="Browser agent başarısız, fallback devrede", state="error")
            browser_error_text = None
            if isinstance(outcome, dict):
                browser_error_text = outcome.get("error")
            if not browser_error_text:
                browser_error_text = "Bilinmeyen browser hatası (hata metni boş döndü)."

            st.error("Gerçek browser ile dashboard okunamadı.")
            with st.expander("Browser agent hata detayı", expanded=True):
                st.code(str(browser_error_text))

            if not allow_export_fallback:
                st.info("PBIRS export fallback kapalı. Önce browser hata detayını çözmemiz gerekiyor.")
                return
            st.warning("PBIRS export fallback denenecek.")

    # ── Sunucudan görsel al ───────────────────────────────────────────────────
    snapshot_bytes: bytes | None = None

    if pbirs_session is None:
        st.info("PBIRS export fallback için kimlik bilgileri verilmedi; browser analizi esas alınacak.")
        return

    with st.spinner("PBIRS sunucusundan rapor görseli render ediliyor…"):
        try:
            snapshot_bytes = get_report_snapshot(pbirs_session, report_path, item_id, type_id)
        except Exception as snap_err:
            st.warning(f"Snapshot alınamadı: {snap_err}")

    # Manüel yüklemeyi yedek olarak kullan
    if snapshot_bytes is None and uploaded_image is not None:
        snapshot_bytes = uploaded_image.getvalue()
        st.info("Sunucudan snapshot alınamadı; yüklediğin görsel kullanılıyor.")

    if snapshot_bytes is not None:
        st.markdown("### 📸 Rapor Görseli (sunucudan)")
        st.image(snapshot_bytes, use_container_width=True)
    else:
        st.error(
            "⚠️ Rapor görseli alınamadı.\n\n"
            "**Olası nedenler:**\n"
            "- Kimlik bilgileri hatalı (401)\n"
            "- Report Server RenderingExtension devre dışı\n"
            "- Power BI raporlarda ExportTo henüz tamamlanmadı (async)\n\n"
            "**Çözüm:** Raporun ekran görüntüsünü manuel olarak alıp yukarıdaki "
            "'Yedek' alanına yükleyebilirsin — AI o görseli okuyacak."
        )

    # ── AI görsel analizi ─────────────────────────────────────────────────────
    if snapshot_bytes is not None:
        if not st.session_state.get("api_key"):
            st.warning("AI analizi için soldan Gemini API key bağlayın.")
        else:
            model_name = st.session_state.get("selected_model", GEMINI_MODEL)
            with st.spinner(f"AI raporu gözüyle okuyor ({model_name})…"):
                try:
                    ai_comment = analyze_dashboard_image_with_ai(
                        snapshot_bytes,
                        model_name,
                        user_question,
                        str(row.get(col_map["name"], report_name)),
                    )
                    st.markdown("### 🤖 AI Görsel Yorumu")
                    st.info(ai_comment)

                    summary = build_executive_summary(
                        str(row.get(col_map["name"], report_name)),
                        build_rule_based_commentary(
                            str(row.get(col_map["name"], "-")), playbook, []
                        ),
                        ai_comment,
                    )
                    with st.expander("📋 Yönetici Özeti"):
                        st.markdown(summary)

                except Exception as ai_err:
                    if _is_gemini_quota_error(ai_err):
                        st.error(
                            "**Gemini ücretsiz kota aşıldı (429).**\n\n"
                            "Yapılabilecekler:\n"
                            "1. [Google AI Studio](https://aistudio.google.com) üzerinden ücretli plana geç\n"
                            "2. Yarın tekrar dene (günlük limit sıfırlanır)\n"
                            "3. Farklı bir API key kullan (soldan 'Çıkış Yap → Sıfırla' ile yeni key gir)"
                        )
                    else:
                        st.error(f"AI analizi başarısız: {ai_err}")
    elif uploaded_image is None:
        # Görsel de yok, AI da çalışamaz — Playbook tabanlı özet göster
        if playbook:
            st.markdown("### Kural Bazlı İş Yorumu (görsel olmadan)")
            st.info(build_rule_based_commentary(
                str(row.get(col_map["name"], "-")), playbook, []
            ))
            st.caption(
                "Not: Bu yorum rapor adı ve iş kurallarına dayanmaktadır. "
                "Gerçek görsel alındığında AI otomatik olarak o görseli okuyacak."
            )



        
        # Sütun Genişliklerini Ayarla (Auto-fit)
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = min(adjusted_width, 50) # Limit to 50
            
        # Filtre Ekle
        worksheet.auto_filter.ref = worksheet.dimensions

    return output.getvalue()


# --- UI COMPONENTS & PAGES ---

def load_query_templates():
    tpl_path = Path(__file__).resolve().parent / "query_templates.json"
    if tpl_path.exists():
        with open(tpl_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def render_report_browser(df, col_map):
    """Kategorize edilmiş rapor kataloğu tarayıcısı."""
    st.markdown('<div class="browser-container">', unsafe_allow_html=True)
    
    # Search within browser
    b_search = st.text_input("🔍 Katalogda isim veya içerik ara...", placeholder="Rapor adı veya KPI yazın...", label_visibility="collapsed")
    
    # Categories Chips
    topics = sorted([t for t in df[col_map["topic"]].unique() if t and t != "nan"])
    topics = ["Tümü"] + topics
    
    sel_topic = st.session_state.get("browser_topic", "Tümü")
    
    # Use columns for horizontal chip simulation (or just a selectbox for stability, but we promised chips)
    # Since Steamlit doesn't have native chips, we'll use a segmented control or a small selectbox with a label
    selected_category = st.pills("Kategori Seçin", topics, selection_mode="single", default="Tümü")
    
    # Filtering Logic
    filtered_df = df.copy()
    if selected_category and selected_category != "Tümü":
        filtered_df = filtered_df[filtered_df[col_map["topic"]] == selected_category]
    
    if b_search:
        filtered_df = filtered_df[
            filtered_df[col_map["name"]].str.contains(b_search, case=False, na=False) |
            filtered_df[col_map["desc"]].str.contains(b_search, case=False, na=False)
        ]
    
    st.markdown(f"**{len(filtered_df)}** rapor listeleniyor")
    
    # Display cards as a grid
    # We'll use Streamlit columns to create a grid
    cols_per_row = 3
    for i in range(0, len(filtered_df), cols_per_row):
        batch = filtered_df.iloc[i:i+cols_per_row]
        cols = st.columns(cols_per_row)
        for j, (idx, row) in enumerate(batch.iterrows()):
            with cols[j]:
                r_name = row[col_map["name"]]
                r_desc = row[col_map["desc"]]
                help_text = str(r_desc) if pd.notna(r_desc) else "Açıklama bulunamadı."
                
                # Custom selection button that looks like a card
                if st.button(f"📄 {r_name}", key=f"browser_{idx}", use_container_width=True, help=help_text):
                    st.session_state.sample_query = r_name
                    st.rerun()
    
    st.markdown('</div>', unsafe_allow_html=True)

def render_home():
    render_premium_header(
        title_main="Raportal Agent'a",
        title_highlight="Hoş Geldiniz!",
        subtitle="Yapay zeka destekli raporlama ekosistemimize hoş geldiniz.",
        icon="👋",
        icon_color="white",
        gradient_start="#1e1b4b",
        gradient_end="#4c1d95",
        highlight_gradient="linear-gradient(90deg, #d8b4fe 0%, #f472b6 100%)",
        tags=[
            ("auto_awesome", "AI Destekli", "#d8b4fe"),
            ("database", "Veri Odaklı", "#818cf8")
        ]
    )

    # Feature Cards
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.html(f"""
<div class="feature-card">
    <div class="feature-icon-box" style="background: #eef2ff; color: #6366f1;">
        <span class="material-icons-outlined">grid_view</span>
    </div>
    <h3 style="margin: 0; color: #1e1b4b; font-size: 1.3rem; font-weight: 800;">Dashboard</h3>
    <p style="margin: 15px 0; color: #64748b; font-size: 0.9rem; line-height: 1.6; flex-grow: 1;">
        Rapor performanslarını izle, trendleri keşfet ve veriye dayalı kararlar al.
    </p>
</div>
""")
        if st.button("Dashboard'a Git →", key="go_dash", use_container_width=True):
            st.session_state.active_page = "Dashboard"
            st.rerun()

    with col2:
        st.html(f"""
<div class="feature-card">
    <div class="feature-icon-box" style="background: #fdf2f8; color: #db2777;">
        <span class="material-icons-outlined">insights</span>
    </div>
    <h3 style="margin: 0; color: #1e1b4b; font-size: 1.3rem; font-weight: 800;">Raportal Insights Hub</h3>
    <p style="margin: 15px 0; color: #64748b; font-size: 0.9rem; line-height: 1.6; flex-grow: 1;">
        Akıllı analizlerle içgörüler üret, rapor DNA’sını keşfet ve öngörüleri yakala.
    </p>
</div>
""")
        if st.button("Insights Hub'a Git →", key="go_hub", use_container_width=True):
            st.session_state.active_page = "Raportal Insights Hub"
            st.rerun()

    with col3:
        st.markdown(f"""
            <div class="feature-card">
                <div class="feature-icon-box" style="background: #f0fdf4; color: #16a34a;">
                    <span class="material-icons-outlined">search</span>
                </div>
                <h3 style="margin: 0; color: #1e1b4b; font-size: 1.3rem; font-weight: 800;">Fix Sorgular</h3>
                <p style="margin: 15px 0; color: #64748b; font-size: 0.9rem; line-height: 1.6; flex-grow: 1;">
                    SQL süreçlerini otomatikleştirin, hataları çöz ve sonuçları anında e-posta ile al.
                </p>
            </div>
        """, unsafe_allow_html=True)
        if st.button("Fix Sorgulara Git →", key="go_fix", use_container_width=True):
            st.session_state.active_page = "Fix Sorgular"
            st.rerun()

    # Bottom Status Bar
    st.html("""
<div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 16px; padding: 1.2rem; margin-top: 30px; display: flex; align-items: center; gap: 20px;">
    <div style="background: #eef2ff; padding: 10px; border-radius: 12px;">
        <span class="material-icons-outlined" style="color: #6366f1;">auto_awesome</span>
    </div>
    <div>
        <div style="font-weight: 800; color: #1e1b4b; font-size: 0.95rem;">Akıllı. Hızlı. Güvenilir.</div>
        <div style="color: #64748b; font-size: 0.85rem;">Raportal Agent, kurumunuzun raporlama süreçlerini hızlandırır, hataları azaltır ve stratejik içgörüler sunar.</div>
    </div>
    <div style="margin-left: auto; opacity: 0.4;">
        <span class="material-icons-outlined" style="font-size: 3rem;">security</span>
    </div>
</div>
""")

def render_premium_header(title_main, title_highlight, subtitle, icon, icon_color, gradient_start, gradient_end, highlight_gradient, tags=None):
    """Premium header component shared across modules."""
    tag_html = ""
    if tags:
        for t_icon, t_text, t_color in tags:
            tag_html += f"""
            <div style="background: rgba(255,255,255,0.05); padding: 8px 12px; border-radius: 10px; border: 1px solid rgba(255,255,255,0.1); display: flex; align-items: center; gap: 8px;">
                <span class="material-icons-outlined" style="color: {t_color}; font-size: 1.1rem;">{t_icon}</span>
                <span style="font-size: 0.8rem; font-weight: 600; color: white;">{t_text}</span>
            </div>
            """

    st.html(f"""
<div class="hero-box" style="padding: 3rem; background: linear-gradient(135deg, {gradient_start} 0%, {gradient_end} 100%); min-height: 250px; margin-bottom: 20px;">
    <div style="display: flex; align-items: center; gap: 30px; position: relative; z-index: 2;">
        <div style="background: rgba(255,255,255,0.05); width: 80px; height: 80px; border-radius: 20px; display: flex; align-items: center; justify-content: center; backdrop-filter: blur(10px); border: 1px solid rgba(255,255,255,0.1);">
            <span class="material-icons-outlined" style="font-size: 2.5rem; color: {icon_color};">{icon}</span>
        </div>
        <div>
            <h1 style="margin:0; color:white; font-size: 2.5rem; font-weight:800; font-family: 'Outfit', sans-serif;">
                {title_main} <span style="background: {highlight_gradient}; -webkit-background-clip: text; -webkit-text-fill-color: transparent;">{title_highlight}</span>
            </h1>
            <p style="margin: 5px 0 0 0; color:rgba(255,255,255,0.7); font-size: 1.1rem; font-family: 'Outfit', sans-serif;">{subtitle}</p>
            <div style="display: flex; gap: 15px; margin-top: 15px;">
                {tag_html}
            </div>
        </div>
    </div>
</div>
""")

def render_dashboard_header():
    render_premium_header(
        title_main="Raportal",
        title_highlight="Vizyoneri",
        subtitle="Yapay Zeka Destekli Rapor ve Dashboard Mimarı",
        icon="auto_awesome",
        icon_color="#38bdf8",
        gradient_start="#1e1b4b",
        gradient_end="#4c1d95",
        highlight_gradient="linear-gradient(90deg, #d8b4fe 0%, #f472b6 100%)",
        tags=[
            ("architecture", "Layout", "#a78bfa"),
            ("query_stats", "Metrik", "#38bdf8")
        ]
    )

def render_dashboard():
    render_dashboard_header()
    
    if not st.session_state.api_key:
        st.info("👋 Sol menüden API anahtarını girerek devam edebilirsin.")
        st.stop()

    try:
        df, source_name = load_metadata()
        df, col_map = prepare_columns(df)
    except Exception as e:
        st.error(f"Dosya yüklenemedi: {e}")
        st.stop()

    tab_new, tab_history = st.tabs(["✨ Yeni Tasarım Yap", "🕛 Geçmiş Taleplerim"])
    
    with tab_new:
        st.html(f"""
<div style="background: white; border: 1px solid #e2e8f0; border-radius: 20px; padding: 1.5rem; display: flex; align-items: center; gap: 20px; margin-bottom: 20px;">
    <div style="background: #f8fafc; padding: 10px; border-radius: 12px; border: 1px solid #e2e8f0;">
        <span class="material-icons-outlined" style="color: #8c28e8;">auto_awesome</span>
    </div>
    <div>
        <div style="font-weight: 800; color: #1e1b4b; font-size: 1rem;">İhtiyacınızı tarif edin, yapay zeka tasarımına başlasın.</div>
        <div style="color: #64748b; font-size: 0.85rem;">Kariyer.net raporlama standartlarına uygun dashboard taslakları saniyeler içinde oluşturulur.</div>
    </div>
    <div style="margin-left: auto; opacity: 0.2;">
        <span class="material-icons-outlined" style="font-size: 4rem;">chat_bubble_outline</span>
    </div>
</div>
""")

        st.markdown('<div class="premium-form-box">', unsafe_allow_html=True)
        col_f1, col_f2 = st.columns([1, 1])
        with col_f1:
            st.html("""
<div style="display: flex; align-items: center; gap: 12px; margin-bottom: 10px;">
    <div style="background: #eef2ff; color: #6366f1; padding: 6px; border-radius: 8px;">
        <span class="material-icons-outlined" style="font-size: 1.2rem;">bolt</span>
    </div>
    <div>
        <div style="font-weight: 800; color: #1e1b4b; font-size: 1rem;">Hızlı Tasarım Tanımı</div>
        <div style="font-size: 0.75rem; color: #64748b;">Rapor veya dashboard ihtiyacınızı aşağıda açıklayın.</div>
    </div>
</div>
""")
        with col_f2:
            v_report_type = st.selectbox("Hedef Format", options=["Power BI", "SSRS", "Excel", "Otomatik"], key="v_format_pick")

        v_prompt = st.text_area("Prompt", placeholder="Örn: Satış performansını izleyen bir dashboard...", height=120, label_visibility="collapsed", key="v_prompt_final")
        
        # Options Toggles
        st.markdown("<div style='margin-top: 15px; margin-bottom: 20px;'>", unsafe_allow_html=True)
        t_col1, t_col2, t_col3, t_col4 = st.columns(4)
        with t_col1: st.checkbox("📝 Yönetici Özeti", value=True, key="opt_exec")
        with t_col2: st.checkbox("📈 Trend Vurgusu", value=False, key="opt_trend")
        with t_col3: st.checkbox("📊 Karşılaştırmalı KPI", value=True, key="opt_comp")
        with t_col4: st.button("+ Diğer Özellik Ekle", use_container_width=True, key="opt_add")
        st.markdown("</div>", unsafe_allow_html=True)

        design_clicked = st.button("✨ TASARIMI BAŞLAT", use_container_width=True, type="primary", key="btn_start_design_final")
        st.markdown('</div>', unsafe_allow_html=True)

        if design_clicked and v_prompt:
            # 1. Clear ALL relevant state
            st.session_state.visionary_result = None
            st.session_state.v_inspiration = []
            st.session_state.v_real_entities = None
            
            # communicate to UI that mockup is invalid
            if "visionary_mockup_valid" in st.session_state:
                st.session_state.visionary_mockup_valid = False
            
            # Step 1: Catalog Inspiration Search
            with st.spinner("🔍 Raportal kataloğu taranıyor..."):
                inspiration_matches = search_reports(df, v_prompt, col_map)
                st.session_state.v_inspiration = inspiration_matches[:3] if inspiration_matches else []
                ins_context = "\n".join([f"- {m['name']} (Path: {m['path']})" for m in inspiration_matches[:3]])
                ins_urls = [m['url'] for m in inspiration_matches if m.get('url')]
            
            # Step 2: Deep Visual Inspiration Scan
            ins_data = []
            if ins_urls:
                with st.spinner("💡 Bulunan benzer raporlar görsel olarak inceleniyor..."):
                    u, p, d = st.session_state.get("username", ""), st.session_state.get("password", ""), st.session_state.get("domain", NTLM_DOMAIN)
                    ins_data = run_visual_inspiration(ins_urls, u, p, d)
            
            # Step 3: Real-World Entity Fetch
            with st.spinner("📊 Kariyer.net dünyasından veriler çekiliyor (İKÇO, Bölümler)..."):
                real_ens = get_real_world_entities()
                st.session_state.v_real_entities = real_ens

            # Step 4: AI Design Generation
            with st.spinner("Raportal Vizyoneri tasarımı planlıyor..."):
                if st.session_state.get("sim_mode"):
                    time.sleep(1.5) # Simulate thinking
                    suggestion = {
                        "prompt": v_prompt,
                        "suggestion": f"""
### 📝 İhtiyaç Özeti
Bu tasarım, **"{v_prompt}"** talebi için Kariyer.net standartlarında kurgulanmıştır. Temel amaç, veriye dayalı hızlı aksiyon alınmasını sağlamaktır.

### Önerilen Dashboard Şablonu
**Premium Visionary Matrix v2.0** - Karmaşık verileri sadeleştiren ve trend odaklı hiyerarşik bir yapı sunar.

### Dashboard Tasarımı
- **Üst Panel:** 4 Ana KPI (Sayısal ve Oransal karşılaştırmalı).
- **Orta Panel:** Zaman bazlı trend grafiği (Alan grafiği).
- **Alt Panel:** Detaylı veri matrisi ve kırılım tabloları.

### KPI ve Boyutlar
- **KPI'lar:** Dönemsel Değişim, Toplam Hacim, Verimlilik Oranı.
- **Boyutlar:** Bölüm, İKÇO, Tarih, Yenileme Tipi.

### Tasarım Gerekçesi
Kariyer Moru (#8c28e8) vurgularıyla odak noktaları belirlenmiş, temiz beyaz alanlar (whitespace) ile okunabilirlik artırılmıştır.
                        """
                    }
                    st.session_state.visionary_result = suggestion
                else:
                    suggestion_text = suggest_report_template(
                        v_prompt, 
                        st.session_state.api_key, 
                        model_name=st.session_state.get("selected_model", GEMINI_MODEL),
                        inspiration_context=ins_context,
                        inspiration_data=ins_data,
                        real_entities=st.session_state.get("v_real_entities"),
                        report_type=v_report_type
                    )
                    suggestion = {
                        "prompt": v_prompt,
                        "suggestion": suggestion_text
                    }
                    st.session_state.visionary_result = suggestion
                
                mockup_path = os.path.join(os.path.dirname(__file__), "assets", "visionary_mockups", "latest_mockup.png")
                if os.path.exists(mockup_path):
                    from dashboard_agent.history_manager import save_visionary_request
                    save_visionary_request(v_prompt, suggestion, image_path=mockup_path)
                    try: os.remove(mockup_path)
                    except: pass
                else:
                    from dashboard_agent.history_manager import save_visionary_request
                    save_visionary_request(v_prompt, suggestion)
                
                st.session_state.visionary_mockup_valid = False
                st.rerun()

        if st.session_state.get("visionary_result"):
            res = st.session_state.visionary_result
            st.markdown('<div class="visionary-result">', unsafe_allow_html=True)
            
            # Data Badges
            tags_cols = st.columns([1,1,1,2])
            with tags_cols[0]: st.markdown("✨ **AI Architect**")
            if st.session_state.get("v_real_entities") and st.session_state.v_real_entities.get("bolumler"):
                with tags_cols[1]: st.markdown("🔗 **BIDB Connected**")
                with tags_cols[2]: st.markdown("🏢 **Real Entities**")
            
            st.markdown(f"### ✨ Tasarım Prototipi: {res['prompt'][:50]}...")
            
            # --- IMAGE DISPLAY LOGIC (SIMULATION AWARE) ---
            if st.session_state.get("sim_mode"):
                # Use local demo mockup to avoid network restrictions
                demo_path = os.path.join(os.path.dirname(__file__), "assets", "visionary_mockups", "demo_mockup.png")
                if os.path.exists(demo_path):
                    st.image(demo_path, caption="🚀 Simülasyon Modu: Premium Dashboard Örneği (Yerel Asset)", use_container_width=True)
                else:
                    st.warning("⚠️ Demo görseli bulunamadı. Lütfen varlıkları kontrol edin.")
            else:
                mockup_path = os.path.join(os.path.dirname(__file__), "assets", "visionary_mockups", "latest_mockup.png")
                if os.path.exists(mockup_path) and st.session_state.get("visionary_mockup_valid", False):
                    st.image(mockup_path, use_container_width=True)
                else:
                    st.info("🎨 Görsel prototip AI tarafından fırçalanıyor... Hazır olduğunda aşağıdaki butona basabilirsin.")
                    if st.button("Görsel Hazır mı? ✅", key="btn_validate_mockup_final"):
                        st.session_state.visionary_mockup_valid = True
                        st.rerun()

            st.markdown(res['suggestion'] if isinstance(res, dict) and 'suggestion' in res else str(res))
            
            c1, c2 = st.columns(2)
            with c1:
                if st.button("📊 Canlı Veriye Bağla", use_container_width=True):
                    st.session_state.live_dashboard_data = load_metadata()[0]
                    st.rerun()
            with c2: st.download_button("📥 Tasarımı İndir", data=str(res), file_name="raportal_tasarim.md", key="dl_f_v")
            
            if st.button("❌ Kapat", key="close_f_v"):
                st.session_state.visionary_result = None
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        if st.session_state.get("live_dashboard_data") is not None:
            df_live = st.session_state.live_dashboard_data
            st.markdown('<div class="visionary-result" style="border-left: 4px solid #00f2fe;">', unsafe_allow_html=True)
            st.markdown("### 🟢 CANLI VERİ EKRANI")
            st.metric("Toplam Satır", f"{len(df_live):,}")
            st.dataframe(df_live, use_container_width=True, height=300)
            if st.button("❌ Canlı Ekranı Kapat", key="close_live_f"):
                st.session_state.live_dashboard_data = None
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

    with tab_history:
        st.markdown("#### 🕛 Son Tasarım Talepleriniz")
        history = get_visionary_history(limit=50)
        if not history: st.info("Henüz geçmiş talebiniz yok.")
        else:
            for item in history:
                with st.expander(f"📌 {item['timestamp']} - {item['prompt'][:40]}..."):
                    col_ex_text, col_ex_del = st.columns([0.8, 0.2])
                    with col_ex_text:
                        st.write(f"**Talep:** {item['prompt']}")
                    with col_ex_del:
                        # Use id if exists, otherwise fallback to timestamp (for older entries)
                        req_id = item.get("id", item.get("timestamp"))
                        if st.button("🗑️ Kaydı Sil", key=f"del_h_{req_id}", use_container_width=True, help="Bu talebi geçmişten siler"):
                            delete_visionary_request(req_id)
                            st.rerun()
                    
                    st.info(f"**AI Özeti:** {item['summary']}")
                    
                    # Restore image display
                    img_path = item.get("image_path")
                    if img_path:
                        full_img_path = os.path.join(os.path.dirname(__file__), img_path)
                        if os.path.exists(full_img_path):
                            st.image(full_img_path, caption=f"Tasarım Prototipi ({item['timestamp']})", use_container_width=True)
                    
                    if st.button("Yükle", key=f"hist_f_{item['timestamp']}"):
                        st.session_state.sample_query = item['prompt']
                        st.rerun()



def send_outlook_mail(to, subject, body_html, attachment_bytes=None, attachment_name="Dosya.xlsx"):
    """Windows üzerinde yüklü Outlook'u kullanarak mail gönderir."""
    if os.name != 'nt':
        return False, "Outlook otomasyonu sadece Windows üzerinde çalışır."
    
    try:
        import win32com.client as win32
        import tempfile
        
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.Subject = subject
        mail.HTMLBody = body_html
        
        if attachment_bytes:
            tmp_path = os.path.join(tempfile.gettempdir(), attachment_name)
            with open(tmp_path, "wb") as tp:
                tp.write(attachment_bytes)
            mail.Attachments.Add(tmp_path)
            
        mail.Send()
        return True, "Başarılı"
    except Exception as e:
        return False, str(e)

def render_fix_sorgular():
    render_premium_header(
        title_main="Fix",
        title_highlight="Sorgular",
        subtitle="Otomatik SQL Sorgu Yönetimi ve Çalıştırma Merkezi",
        icon="code",
        icon_color="#38bdf8",
        gradient_start="#1e1b4b",
        gradient_end="#4c1d95",
        highlight_gradient="linear-gradient(90deg, #d8b4fe 0%, #f472b6 100%)",
        tags=[
            ("terminal", "Sorgu", "#a78bfa"),
            ("auto_mode", "Otomasyon", "#38bdf8")
        ]
    )
    
    templates = load_query_templates()
    if not templates:
        st.warning("Henüz yüklü sorgu şablonu bulunamadı.")
        return

    tab_run, tab_schedule = st.tabs(["🚀 Sorgu Çalıştır", "🕒 Otomatik Zamanlama"])

    with tab_run:

        # --- TOP CARD: Selection & Parameters ---
        st.markdown('<div class="custom-card">', unsafe_allow_html=True)
        
        # Row 1: Template Selection
        r1_col1, r1_col2 = st.columns([1, 1])
        with r1_col1:
            selected_tpl_key = st.selectbox("📝 Sorgu Şablonu", options=list(templates.keys()))
            tpl = templates[selected_tpl_key]
        with r1_col2:
            st.markdown(f"""
            <div style="background: #f0f7ff; padding: 12px; border-radius: 12px; border: 1px solid #dbeafe; min-height: 75px; display: flex; align-items: center; margin-top: 5px;">
                <div style="font-size: 0.85rem; color: #1e40af; line-height: 1.4;">
                    <strong>💡 Şablon Bilgisi:</strong><br>{tpl['description']}
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<div style='height: 15px;'></div>", unsafe_allow_html=True)
        
        # Row 2: Parameters & Run Button
        r2_col1, r2_col2, r2_col3 = st.columns([1, 1, 2])
        with r2_col1:
            donem_in = st.text_input("📅 Hedef Dönem", value="202603", help="Örn: 202603")
        with r2_col2:
            target_db = st.text_input("🗄️ Veritabanı", value="DWH", help="Varsayılan: DWH")
        with r2_col3:
            st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True) # Perfect alignment with inputs
            run_bidb = st.button("🚀 bidb Üzerinde Çalıştır (Canlı)", type="primary", use_container_width=True)

        # Row 3: Auto Mail
        st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
        m_auto_cols = st.columns([1, 1])
        with m_auto_cols[0]:
            auto_mail = st.checkbox("✉️ İşlem Sonrası Otomatik Mail Gönder", value=False)
        with m_auto_cols[1]:
            if auto_mail:
                auto_to = st.text_input("Alıcı E-Posta", placeholder="isminiz@kariyer.net", key="fix_auto_to", label_visibility="collapsed")
            else:
                auto_to = ""
                
        st.markdown('</div>', unsafe_allow_html=True)

        formatted_sql = tpl['sql_template'].replace("{{donem}}", donem_in)
        
        # --- SQL ACCORDION ---
        with st.expander("🔍 Güncel SQL Sorgusunu Görüntüle / Detaylar"):
            st.markdown(f"Bu sorgu şu anda **{target_db}** üzerinde çalışacak.")
            # Scrollable SQL box
            st.markdown(f'<div class="sql-scroll-box">{html.escape(formatted_sql)}</div>', unsafe_allow_html=True)
            st.caption("Bu sorguyu SSMS üzerinden manuel olarak da çalıştırabilirsiniz.")
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # --- RESULTS AREA ---
        if run_bidb:
            with st.spinner(f"{target_db} sunucusuna bağlanılıyor..."):
                try:
                    start_time = time.time()
                    df_live = run_query_on_bidb(formatted_sql, database=target_db)
                    end_time = time.time()
                    
                    st.success(f"Sorgu başarıyla tamamlandı! ({round(end_time - start_time, 2)} sn)")
                    
                    # Metrics card
                    st.markdown('<div class="custom-card">', unsafe_allow_html=True)
                    m1, m2, m3 = st.columns(3)
                    m1.metric("Toplam Satır", len(df_live))
                    if "FinalCheckGelir" in df_live.columns:
                        m2.metric("Toplam Gelir", f"{df_live['FinalCheckGelir'].sum():,.2f} TL")
                    if "TahsilatDurumu" in df_live.columns:
                        rate = (df_live['TahsilatDurumu'].sum() / len(df_live)) * 100 if len(df_live) > 0 else 0
                        m3.metric("Tahsilat Oranı", f"%{rate:,.1f}")
                    
                    st.dataframe(df_live.head(100), use_container_width=True, height=350)
                    
                    # Excel Download
                    excel_data = None
                    try:
                        excel_data = to_excel(df_live)
                        st.session_state['last_excel_data'] = excel_data
                        st.download_button(
                            label="📥 Sonucu Excel Olarak İndir (Segoe Font & Mor Başlık)",
                            data=excel_data,
                            file_name=f"FinalCheck_{donem_in}_{int(time.time())}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    except Exception as ex:
                        st.warning(f"Excel oluşturulamadı: {ex}")
                        csv = df_live.to_csv(index=False).encode('utf-8-sig')
                        st.session_state['last_excel_data'] = csv
                        st.download_button("📥 CSV İndir", data=csv, file_name=f"FinalCheck_{donem_in}.csv", mime="text/csv")
                        excel_data = csv
                    
                    # Save to session
                    st.session_state['last_df_live'] = df_live
                    st.session_state['last_donem'] = donem_in
                    st.session_state['last_tpl_key'] = selected_tpl_key
                    
                    # AUTOMATIC MAIL TRIGGER
                    if auto_mail and auto_to:
                        with st.spinner("📧 Otomatik mail gönderiliyor..."):
                            body = f"Merhaba,<br><br>{selected_tpl_key} sorgu sonuçları {donem_in} dönemi için ekte sunulmuştur.<br><br>İyi çalışmalar."
                            subject = f"{selected_tpl_key} Sonuçları - {donem_in}"
                            ok, msg = send_outlook_mail(auto_to, subject, body, excel_data, f"FinalCheck_{donem_in}.xlsx")
                            if ok:
                                st.success(f"✅ Sonuçlar {auto_to} adresine otomatik olarak gönderildi.")
                            else:
                                st.error(f"❌ Otomatik mail gönderilemedi: {msg}")

                    st.markdown('</div>', unsafe_allow_html=True)

                except Exception as e:
                    st.info("İpucu: ODBC Driver yanısıra VPN bağlantınızın açık olduğundan emin olun.")

    with tab_schedule:
        st.markdown("### 🕒 Otomatik Rapor Zamanlayıcı")
        st.info("""
🕒 **Otomatik Raporlama Rehberi**

💡 **Nasıl Çalışır?** Buradan eklediğiniz raporlar, her ayın 1'inde otomatik olarak **bir önceki ayın** verisiyle (N-1) çekilir ve gönderilir.

⚠️ **Önemli:** Tablo yenileme (refresh) saatlerinden (saat başı ve buçuklar) kaçınmak için gönderimi **09:15** veya **10:15** gibi ara saatlere kurmanız önerilir.
""")
        
        st.markdown('<div class="custom-card">', unsafe_allow_html=True)
        with st.form("schedule_form"):
            s_tpl = st.selectbox("Zamanlanacak Rapor", options=list(templates.keys()))
            s_mail = st.text_input("Alıcı E-Posta", placeholder="isminiz@kariyer.net")
            submit = st.form_submit_button("✅ Görevi Zamanla")
            
            if submit:
                if not s_mail:
                    st.error("Lütfen bir e-posta adresi girin.")
                else:
                    ok, msg = save_schedule(s_tpl, s_mail)
                    if ok: st.success(msg)
                    else: st.warning(msg)
        st.markdown('</div>', unsafe_allow_html=True)

        # List existing
        path = Path(__file__).resolve().parent / "config" / "schedules.json"
        if path.exists():
            try:
                with open(path, "r", encoding="utf-8") as f:
                    schedules = json.load(f)
            except: schedules = []
                
            if schedules:
                st.markdown("<br>#### Aktif Zamanlamalar", unsafe_allow_html=True)
                for i, s in enumerate(schedules):
                    with st.expander(f"📌 {s['template_name']} ➔ {s['recipient']}"):
                        c1, c2, c3 = st.columns([2, 2, 1])
                        c1.write(f"**Son Çalışma:** {s.get('last_run', 'N/A')}")
                        c2.write(f"**Durum:** {s.get('status', 'Bekliyor')}")
                        if c3.button("🗑️ Sil", key=f"del_sch_{i}"):
                            schedules.pop(i)
                            with open(path, "w", encoding="utf-8") as f:
                                json.dump(schedules, f, indent=4)
                            st.rerun()

        # LOG SECTION
        log_path = Path(__file__).resolve().parent / "logs" / "scheduler.log"
        if log_path.exists():
            st.markdown("<br>#### 📜 Gönderim Günlüğü (Son İşlemler)", unsafe_allow_html=True)
            try:
                with open(log_path, "r", encoding="utf-8") as lf:
                    log_lines = lf.readlines()
                # Show last 10 lines
                st.code("".join(log_lines[-15:]), language="text")
            except:
                st.info("Henüz bir işlem günlüğü oluşmadı.")



def render_pbix_analyzer():
    st.markdown('<div class="section-title">Power BI Analiz Fabrikası (Batch Engine)</div>', unsafe_allow_html=True)
    st.write("Birden fazla PBIX dosyasını otomatik PBIT'e dönüştürün ve rapor DNA'larını toplu olarak analiz edin.")
    
    # Session state for batch results
    if "batch_results" not in st.session_state:
        st.session_state.batch_results = {} # filename -> metadata_dict

    st.markdown('<div class="custom-card">', unsafe_allow_html=True)
    uploaded_files = st.file_uploader("Power BI Raporlarını Seçin (.pbix, .pbit, .bim)", 
                                     type=["pbix", "pbit", "bim"], 
                                     accept_multiple_files=True)
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_files:
        # Check for new files to clear old batch if needed or just add
        # We'll use a button to trigger the process to avoid heavy UI reruns
        st.markdown('<br>', unsafe_allow_html=True)
        col_btn, col_info = st.columns([1, 2])
        process_trigger = col_btn.button("🚀 Tümünü Dönüştür ve Analiz Et", type="primary", use_container_width=True)
        
        if process_trigger:
            groups = {} # base_name -> list of uploaded_files
            for f in uploaded_files:
                base = f.name.rsplit('.', 1)[0]
                if base not in groups: groups[base] = []
                groups[base].append(f)

            for base_name, files in groups.items():
                if base_name in st.session_state.batch_results: continue
                
                with st.spinner(f"Hibrit Analiz: {base_name}..."):
                    try:
                        master_metadata = None
                        temp_save_path = None
                        
                        # Process files
                        for f in files:
                            # Save to temp for potential Robot usage
                            local_path = os.path.join("temp_pbix", f.name)
                            with open(local_path, "wb") as tp:
                                tp.write(f.getbuffer())
                            
                            if f.name.lower().endswith(".pbit"):
                                master_metadata = parse_pbi_file(f.read(), f.name)
                                master_metadata["source_type"] = "Master (PBIT)"
                            elif f.name.lower().endswith(".pbix"):
                                temp_save_path = os.path.abspath(local_path)

                        pbix_file = next((f for f in files if f.name.lower().endswith(".pbix")), None)
                        if pbix_file and not (master_metadata and master_metadata.get("source_quality") == "High"):
                            # PBIX Scavenge
                            from pbi_parser import convert_to_pbit_bytes
                            pbit_payload = convert_to_pbit_bytes(pbix_file.read())
                            pbix_meta = parse_pbi_file(pbit_payload if pbit_payload else pbix_file.read(), pbix_file.name)
                            
                            if not master_metadata:
                                master_metadata = pbix_meta
                                master_metadata["source_type"] = "Hybrid (PBIX-Forensic)"
                            else:
                                if not master_metadata.get("pages") and pbix_meta.get("pages"):
                                    master_metadata["pages"] = pbix_meta["pages"]
                            
                            master_metadata["pbit_payload"] = pbit_payload
                            master_metadata["local_pbix_path"] = temp_save_path

                        if master_metadata:
                            st.session_state.batch_results[base_name] = master_metadata
                    except Exception as e:
                        st.session_state.batch_results[base_name] = {"status": "Hata", "error": str(e)}
            st.success("Hibrid tarama tamamlandı.")

        # --- HYBRID SUMMARY DASHBOARD ---
        if st.session_state.batch_results:
            st.markdown('<div class="section-title">Hibrid Rapor Envanteri</div>', unsafe_allow_html=True)
            summary_rows = []
            for name, meta in st.session_state.batch_results.items():
                if "error" in meta:
                    summary_rows.append({"Rapor": name, "Durum": "🔴 Hata", "PBIT Gerekli?": "NA"})
                    continue
                
                # Logic for PBIT Necessity
                pbit_req_badge = "✅ HAYIR" if not meta.get("pbit_required") else "⚠️ EVET"
                
                summary_rows.append({
                    "Rapor Adı": name,
                    "Tip": meta.get("source_type", meta.get("type")),
                    "Kalite": meta.get("source_quality", "N/A"),
                    "Tablo": len(meta.get("tables", [])),
                    "Ölçü": len(meta.get("measures", [])),
                    "İlişki": len(meta.get("relationships", [])),
                    "Veri Kaynağı": ", ".join(meta.get("data_sources", []))[:30] + "...",
                    "PBIT Gerekli mi?": pbit_req_badge
                })
            
            st.dataframe(pd.DataFrame(summary_rows), use_container_width=True)

            if st.button("🗑️ Tüm Analizleri Temizle"):
                st.session_state.batch_results = {}
                st.rerun()

            # --- INDIVIDUAL REPORT CARDS ---
            st.markdown('<div class="section-title">Bireysel Rapor DNA Kartları (Hibrid)</div>', unsafe_allow_html=True)
            for fname, metadata in st.session_state.batch_results.items():
                if "error" in metadata:
                    st.error(f"⚠️ {fname}: {metadata['error']}")
                    continue

                card_status = "✅ Tam Veri (DNA)" if metadata.get("source_quality") == "High" else "⚠️ Kısıtlı Veri (Robot Gerekebilir)"
                with st.expander(f"📊 {fname} — {card_status}", expanded=False):
                    # Robot and Download Bar
                    col_robot, col_dl = st.columns([1, 1])
                    
                    if metadata.get("pbit_required") and metadata.get("local_pbix_path"):
                        with col_robot:
                            if st.button(f"🤖 PBI Robotu Başlat (Derin Analiz)", key=f"robot_{fname}", type="primary"):
                                with st.spinner("PBI Robotu devreye giriyor... Power BI Desktop açılıyor..."):
                                    pbit_file = metadata["local_pbix_path"].replace(".pbix", ".pbit")
                                    success, msg = trigger_pbi_robot_export(metadata["local_pbix_path"], pbit_file)
                                    if success:
                                        st.success("PBIT Şablonu başarıyla üretildi! DNA analizi tazeleniyor...")
                                        # Parse the new PBIT
                                        if os.path.exists(pbit_file):
                                            with open(pbit_file, "rb") as pf:
                                                new_meta = parse_pbi_file(pf.read(), f"{fname}.pbit")
                                                # Update state
                                                new_meta["local_pbix_path"] = metadata["local_pbix_path"]
                                                st.session_state.batch_results[fname] = new_meta
                                                st.rerun()
                                    else:
                                        st.error(f"Robot Hatası: {msg}")

                    if metadata.get("pbit_payload"):
                        with col_dl:
                            st.download_button(f"📥 Otomatik Üretilen .pbit İndir", 
                                                 data=metadata["pbit_payload"], 
                                                 file_name=f"{fname}_raportal_template.pbit", 
                                                 mime="application/x-zip-compressed", key=f"dl_{fname}")
                    
                    m1, m2, m3, m4 = st.columns(4)
                    m1.metric("Tip", metadata.get("type", "Bilinmiyor"))
                    m2.metric("Tablo", len(metadata.get("tables", [])))
                    m3.metric("Ölçü", len(metadata.get("measures", [])))
                    m4.metric("Sayfa", len(metadata.get("pages", [])))

                    if metadata.get("forensic_matches"):
                        st.info("🔍 Adli Tarama: Binary bloklardan veri parçacıkları kurtarıldı.")

                    # DNA Tabs for this specific report
                    t_pages, t_tables, t_rels, t_measures, t_m, t_src = st.tabs([
                        "📄 Sayfalar", "📊 Tablalar", "🔗 İlişkiler", "📐 DAX Ölçüleri", "⚡ Power Query", "🔌 Kaynaklar"
                    ])
                    
                    with t_pages:
                        if metadata.get("pages"):
                            for p in metadata["pages"]: st.write(f"- {p['name']}")
                    
                    with t_tables:
                        if metadata.get("tables"):
                            for t in metadata["tables"]:
                                with st.expander(f"📁 {t['name']}"):
                                    st.write(", ".join(t['columns']))
                                    if t.get('partitions'):
                                        st.code(t['partitions'][0]['query'], language="powerquery")

                    with t_rels:
                        if metadata.get("relationships"):
                            st.table(pd.DataFrame(metadata["relationships"]))

                    with t_measures:
                        if metadata.get("measures"):
                            st.dataframe(pd.DataFrame(metadata["measures"]), use_container_width=True)
                        if metadata.get("forensic_matches"):
                            with st.expander("Ham Binary DAX"):
                                for match in metadata["forensic_matches"]: st.code(match)

                    with t_m:
                        if metadata.get("m_queries"):
                            for q_name, q_code in metadata["m_queries"].items():
                                with st.expander(f"⚡ {q_name}"): st.code(q_code, language="powerquery")

                    with t_src:
                        if metadata.get("data_sources"):
                            for ds in metadata["data_sources"]: st.code(ds)
    else:
        # Instructions
        st.markdown("""
        ### Fabrika Nasıl Çalışır?
        1. Listeye istediğiniz kadar `.pbix` dosyası ekleyin.
        2. **'Tümünü Analiz Et'** butonuyla süreci başlatın.
        3. Sistem her dosyayı PBIT'e çevirip DNA analizini yapar.
        4. Sonuçları özet tabloda ve rapor kartlarında inceleyin.
        """)



def _run_agent_in_thread(url: str, username: str, password: str, domain: str, api_key: str, headless: bool = False, model_name: str = None, max_pages: int = 5) -> dict:
    """Playwright async kodunu ayrı thread'de çalıştır ve yapılandırılmış sonuç döndür."""
    try:
        from dashboard_agent.browser_agent import RaportalBrowserAgent, AuthError, PageLoadError, PlaywrightNotInstalledError
        from dashboard_agent.dashboard_reader import DashboardReader
        from dashboard_agent.analyzer import analyze_dashboard as da_analyze
        from dashboard_agent.config import ANALYSIS_DIR
    except ImportError as e:
        return {
            "ok": False,
            "analysis": None,
            "data": None,
            "error": f"İçe aktarma hatası: {e}",
        }

    async def _inner():
        agent = RaportalBrowserAgent(username, password, domain, headless=headless)
        try:
            await agent.start()
            reader = DashboardReader(agent)
            data   = await reader.read_dashboard(url, max_pages=max_pages)
            # JSON kaydet
            safe_name = re.sub(r"[^\w\u00C0-\u017E\-]", "_", data.report_name).strip()[:80]
            json_path = ANALYSIS_DIR / f"{safe_name}.json"
            json_path.write_text(json.dumps(data.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
            analysis = da_analyze(data, api_key, model_name=model_name)
            md_path  = ANALYSIS_DIR / f"{safe_name}.md"
            md_path.write_text(analysis, encoding="utf-8")
            return {
                "ok": True,
                "analysis": analysis,
                "data": data,
                "strategy": getattr(agent, "startup_strategy", "unknown"),
                "error": None,
            }
        finally:
            await agent.close()

    result = {"value": None, "error": None}
    def thread_fn():
        try:
            if os.name == "nt":
                loop = asyncio.ProactorEventLoop()
                asyncio.set_event_loop(loop)
                try:
                    result["value"] = loop.run_until_complete(_inner())
                finally:
                    loop.close()
                    asyncio.set_event_loop(None)
            else:
                result["value"] = asyncio.run(_inner())
        except BaseException as e:
            # Bazı runtime/playwright hataları boş mesaj döndürebilir; traceback sakla.
            err_msg = str(e).strip() or repr(e)
            result["error"] = f"{err_msg}\n\n{traceback.format_exc()}"
    t = threading.Thread(target=thread_fn)
    t.start()
    t.join(timeout=480)
    if t.is_alive():
        return {
            "ok": False,
            "analysis": None,
            "data": None,
            "error": "Ajan zaman aşımına uğradı (480 sn). Browser süreci takılmış olabilir veya rapor çok yavaş yükleniyor.",
        }
    if result["error"] is not None:
        return {
            "ok": False,
            "analysis": None,
            "data": None,
            "error": result["error"],
        }
    if not result["value"]:
        return {
            "ok": False,
            "analysis": None,
            "data": None,
            "error": "Ajan sonuç üretemedi.",
        }
    return result["value"]


def render_dashboard_agent():
    st.markdown('<div class="section-title">🤖 Dashboard Ajan</div>', unsafe_allow_html=True)
    st.caption(
        "Playwright ile raportal.kariyer.net'teki herhangi bir raporun sayfasını gerçek anlamda açar, "
        "her sekmenin screenshot'ını alır, DOM'daki görünür metinleri çeker ve "
        "Gemini ile **Türkçe iş yorumu** üretir. Hiçbir değişiklik yapmaz (read-only)."
    )

    # ── Playwright kurulum kontrolü ────────────────────────────────────────
    try:
        import playwright  # noqa
        pw_ok = True
    except ImportError:
        pw_ok = False

    if not pw_ok:
        st.error(
            "**Playwright kurulu değil.** Terminalde şunu çalıştır:\n\n"
            "```\npip install playwright\npy -m playwright install chromium\n```"
        )
        return

    # ── Kimlik bilgileri ────────────────────────────────────────────────────
    with st.expander("🔐 Bağlantı Bilgileri", expanded=True):
        st.warning("Şifre yalnızca analiz sırasında kullanılır; dosyaya yazılmaz.")
        ca, cb, cc = st.columns([2, 2, 1])
        with ca:
            ag_user   = st.text_input("Kullanıcı adı", placeholder="esra.akinci", key="ag_user")
        with cb:
            ag_pass   = st.text_input("Şifre", type="password", key="ag_pass")
        with cc:
            ag_domain = st.text_input("Domain", value="KARIYER", key="ag_domain")
        
        ag_headless = st.checkbox("🕵️ Gizli Mod (Browser penceresini gizle)", value=True, key="ag_headless_mode")

    # ── Rapor URL girişi ───────────────────────────────────────────────────
    # Metadata'dan hızlı seçim veya manuel URL
    col_sel, col_or = st.columns([3, 1])
    with col_sel:
        try:
            df_meta, col_map = prepare_columns(load_metadata()[0])
            df_meta["URL"] = df_meta.apply(lambda r: get_raportal_url(r, col_map), axis=1)
            report_options = ["(manuel gir)"] + sorted(df_meta[col_map["name"]].astype(str).unique().tolist())
            selected_report = st.selectbox("Katalogdan seç", report_options, key="ag_report_sel")
        except Exception:
            selected_report = "(manuel gir)"
            df_meta = None

    manual_url = st.text_input(
        "Rapor URL'si",
        placeholder="https://raportal.kariyer.net/home/report/Model/...",
        key="ag_url",
    )

    # URL kaynağını belirle
    target_url = manual_url.strip()
    if selected_report and selected_report != "(manuel gir)" and not target_url:
        try:
            row = df_meta[df_meta[col_map["name"]].astype(str) == selected_report].iloc[0]
            target_url = row["URL"]
            st.caption(f"Hedef: {target_url}")
        except Exception:
            pass

    ag_question = st.text_input(
        "AI'ya özel soru (isteğe bağlı)",
        value="Bu raporda hangi KPI'lar var, hangi iş kararlarına destek veriyor?",
        key="ag_question",
    )

    run_btn = st.button("▶ Analizi Başlat", type="primary", key="ag_run")
    if not run_btn:
        st.info(
            "**Nasıl çalışır?**\n\n"
            "1. Kullanıcı adı, şifre ve domain gir\n"
            "2. Katalogdan rapor seç ya da URL'yi manuel yaz\n"
            "3. 'Analizi Başlat' düğmesine bas\n"
            "4. Agent gerçek bir browser açar, sayfayı görür, screenshot alır\n"
            "5. Gemini tüm sekme görsellerini ve DOM metnini okuyarak Türkçe iş yorumu üretir\n\n"
            "**Gereksinimler:** Playwright (`pip install playwright && py -m playwright install chromium`) "
            "ve geçerli Gemini API key (soldan bağlan)"
        )
        return

    if not ag_user or not ag_pass:
        st.error("Kullanıcı adı ve şifre zorunlu.")
        return
    if not target_url:
        st.error("Rapor URL'si girilmedi.")
        return
    if not st.session_state.get("api_key"):
        st.warning("AI yorumu için Gemini API key gerekli. Soldan bağlanın. Screenshot ve DOM verisi yine de alınacak.")

    api_key    = st.session_state.get("api_key", "")
    model_name = st.session_state.get("selected_model", GEMINI_MODEL)

    # ── Analiz çalıştır ────────────────────────────────────────────────────
    with st.status("🤖 Dashboard Ajan çalışıyor…", expanded=True) as status:
        st.write("Browser başlatılıyor ve NTLM kimlik doğrulaması yapılıyor…")
        try:
            # Buradaki akışta slider henüz yok ama default 5 veya sabit 1 eklenebilir
            # Dashboard Ajan (Eski Tab) için default 3 verelim
            outcome = _run_agent_in_thread(target_url, ag_user, ag_pass, ag_domain, api_key, headless=ag_headless, model_name=model_name, max_pages=3)
        except Exception as exc:
            st.error(f"Ajan çalışırken hata: {exc}")
            status.update(label="Hata oluştu", state="error")
            return

        if not isinstance(outcome, dict):
            st.error("Ajan beklenmeyen formatta sonuç döndürdü.")
            status.update(label="Hata oluştu", state="error")
            return

        if not outcome.get("ok"):
            st.error(outcome.get("error") or "Ajan çalışırken bilinmeyen hata oluştu.")
            status.update(label="Hata oluştu", state="error")
            return

        analysis_text = outcome.get("analysis")
        dashboard_data = outcome.get("data")
        if dashboard_data is None:
            st.error("Ajan dashboard verisini oluşturamadı.")
            status.update(label="Hata oluştu", state="error")
            return
        if outcome.get("strategy"):
            st.caption(f"Açılış stratejisi: {outcome.get('strategy')}")
        status.update(label="✅ Analiz tamamlandı!", state="complete")

    # ── Sonuçları göster ───────────────────────────────────────────────────
    if dashboard_data and dashboard_data.pages:
        st.markdown("---")
        st.subheader(f"📊 {dashboard_data.report_name}")
        st.caption(f"{len(dashboard_data.pages)} sayfa/sekme tarandı")

        # Sekme başına screenshot + DOM özeti
        tab_names = [p.tab_name for p in dashboard_data.pages]
        if len(tab_names) > 1:
            tabs_ui = st.tabs(tab_names)
        else:
            tabs_ui = [st.container()]

        for tab_ui, page in zip(tabs_ui, dashboard_data.pages):
            with tab_ui:
                img_path = Path(page.screenshot_path)
                if img_path.exists():
                    st.image(str(img_path), caption=f"Screenshot — {page.tab_name}", use_container_width=True)
                else:
                    st.warning("Screenshot alınamadı.")
                with st.expander("DOM Özeti (filtreler, KPI'lar, görsel başlıkları)"):
                    if page.filters:
                        st.write("**Filtreler:**", ", ".join(page.filters))
                    if page.kpi_values:
                        st.write("**KPI Değerleri:**", ", ".join(page.kpi_values))
                    if page.visual_titles:
                        st.write("**Görsel Başlıkları:**", ", ".join(page.visual_titles))
                    if page.visible_text:
                        with st.expander("Görünür Metin (DOM)"):
                            st.code(page.visible_text[:3000])

    # ── AI Analiz metni ────────────────────────────────────────────────────
    if analysis_text:
        st.markdown("---")
        st.markdown("## 🤖 AI Analiz Çıktısı")
        st.markdown(analysis_text)

        # İndirme butonu
        st.download_button(
            "📄 Analizi İndir (.md)",
            data=analysis_text.encode("utf-8"),
            file_name=f"{(dashboard_data.report_name if dashboard_data else 'analiz')}_analiz.md",
            mime="text/markdown",
        )


def render_pbit_downloader():
    st.markdown('<div class="section-title">📥 PBIT İndir — PBIRS Toplu İndirici</div>', unsafe_allow_html=True)
    st.caption(
        "raportal.kariyer.net üzerindeki tüm Power BI raporlarını "
        "PBIT (şablon) veya PBIX olarak indirir. "
        "NTLM domain kimlik bilgilerin gerekir."
    )

    # ── Kimlik bilgileri ─────────────────────────────────────────────────────
    with st.expander("🔐 Bağlantı Bilgileri", expanded=True):
        st.warning("Şifre yalnızca indirme isteği sırasında kullanılır; dosyaya kaydedilmez.")
        c1, c2, c3 = st.columns([2, 2, 1])
        with c1:
            dl_user = st.text_input("Kullanıcı adı", placeholder="esra.akinci", key="pbit_user")
        with c2:
            dl_pass = st.text_input("Şifre", type="password", key="pbit_pass")
        with c3:
            dl_domain = st.text_input("Domain", value="KARIYER", key="pbit_domain")

    # ── Seçenekler ───────────────────────────────────────────────────────────
    col_a, col_b, col_c = st.columns([2, 1, 1])
    with col_a:
        dl_limit = st.number_input("Maksimum rapor sayısı (0 = tümü)", min_value=0, value=0, step=5)
    with col_b:
        dry_run = st.checkbox("Kuru çalıştır (indirme yapma, listele)", value=True)
    with col_c:
        prefer_pbit = st.checkbox("PBIT tercih et (yoksa PBIX)", value=True)

    run = st.button("Bağlan ve İndir", type="primary")
    if not run:
        st.info(
            "**Nasıl çalışır?**  \n"
            "1. Kullanıcı adı + şifreni gir  \n"
            "2. Önce *Kuru çalıştır* ile kaç rapor bulunduğunu gör  \n"
            "3. Onay verdikten sonra işareti kaldır ve tekrar bas  \n"
            "4. Dosyalar `pbit_downloads/` klasörüne kaydedilir"
        )
        return

    if not dl_user or not dl_pass:
        st.error("Kullanıcı adı ve şifre zorunlu.")
        return

    # ── PBIRS oturumu aç ─────────────────────────────────────────────────────
    pbirs_session = create_pbirs_session(dl_user, dl_pass, dl_domain)
    PBIRS_API_BASE = "https://raportal.kariyer.net/powerbi/api/v2.0"

    with st.spinner("Sunucuya bağlanılıyor…"):
        try:
            test = pbirs_session.get(f"{PBIRS_API_BASE}/PowerBIReports?$top=1", timeout=15)
            if test.status_code == 401:
                st.error("Kimlik doğrulama başarısız (401). Kullanıcı adı / şifre kontrol edin.")
                return
            elif test.status_code != 200:
                st.error(f"Sunucu hatası: HTTP {test.status_code}")
                return
            st.success(f"✓ Bağlantı başarılı — KARIYER\\{dl_user}")
        except Exception as e:
            st.error(f"Bağlantı hatası: {e}")
            return

    # ── Rapor listesi ─────────────────────────────────────────────────────────
    with st.spinner("Power BI raporları listeleniyor…"):
        try:
            resp = pbirs_session.get(f"{PBIRS_API_BASE}/PowerBIReports?$top=2000", timeout=30)
            reports = resp.json().get("value", [])
        except Exception as e:
            st.error(f"Rapor listesi alınamadı: {e}")
            return

    if not reports:
        st.warning("Hiç Power BI raporu bulunamadı.")
        return

    st.metric("Bulunan Power BI raporu", len(reports))

    if dl_limit and int(dl_limit) > 0:
        reports = reports[: int(dl_limit)]
        st.caption(f"(ilk {len(reports)} rapor işlenecek)")

    # Listeyi tablo olarak göster
    preview_rows = [
        {
            "Rapor Adı": r.get("Name", "-"),
            "Path": r.get("Path", "-"),
            "ID": (r.get("Id") or "")[:8] + "…",
        }
        for r in reports
    ]
    st.dataframe(preview_rows, use_container_width=True)

    if dry_run:
        st.info("Kuru çalıştır modu: dosya indirilmedi. İşareti kaldırıp tekrar bas.")
        return

    # ── Gerçek indirme ────────────────────────────────────────────────────────
    import re as _re
    from pathlib import Path as _Path

    def _safe(name):
        return _re.sub(r'[<>:"/\\|?*]', "_", name).strip()

    out_root = _Path("pbit_downloads")
    out_root.mkdir(parents=True, exist_ok=True)

    progress_bar = st.progress(0, text="İndirme başlıyor…")
    log_area     = st.empty()
    logs         = []
    ok = skip = err = 0

    for i, report in enumerate(reports, 1):
        name  = report.get("Name")  or "unknown"
        path  = report.get("Path")  or ""
        rid   = report.get("Id")    or ""

        rel       = path.lstrip("/").replace("/", "\\") if path else name
        out_dir   = out_root / _Path(rel).parent
        ext       = "pbit" if prefer_pbit else "pbix"
        out_file  = out_dir / f"{_safe(name)}.{ext}"

        progress_bar.progress(i / len(reports), text=f"[{i}/{len(reports)}] {name}")

        if out_file.exists():
            logs.append(f"– {name}: zaten var")
            skip += 1
            log_area.code("\n".join(logs[-30:]))
            continue

        out_dir.mkdir(parents=True, exist_ok=True)
        downloaded = False

        if prefer_pbit:
            # ExportTo PBIT (async)
            try:
                er = pbirs_session.post(
                    f"{PBIRS_API_BASE}/PowerBIReports({rid})/ExportTo",
                    json={"format": "PBIT", "powerBIReportConfiguration": {"pages": []}},
                    timeout=30,
                )
                if er.status_code == 200 and len(er.content) > 100:
                    out_file.write_bytes(er.content)
                    downloaded = True
                elif er.status_code in (202,):
                    poll = er.headers.get("Operation-Location") or er.headers.get("Location")
                    if poll:
                        for _ in range(40):
                            time.sleep(3)
                            pr = pbirs_session.get(poll, timeout=30)
                            if pr.status_code == 200:
                                st_val = pr.json().get("status", "")
                                if st_val == "Succeeded":
                                    fr = pbirs_session.get(poll.rstrip("/") + "/files/0", timeout=120)
                                    if fr.status_code == 200:
                                        out_file.write_bytes(fr.content)
                                        downloaded = True
                                    break
                                elif st_val == "Failed":
                                    break
            except Exception:
                pass

        if not downloaded:
            # PBIX fallback
            pbix_file = out_dir / f"{_safe(name)}.pbix"
            try:
                cr = pbirs_session.get(f"{PBIRS_API_BASE}/PowerBIReports({rid})/Content", timeout=120)
                if cr.status_code == 200 and len(cr.content) > 100:
                    pbix_file.write_bytes(cr.content)
                    downloaded = True
                    out_file = pbix_file
            except Exception:
                pass

        if downloaded:
            ok += 1
            logs.append(f"✓ {name}")
        else:
            err += 1
            logs.append(f"✗ {name}: indirilemedi")

        log_area.code("\n".join(logs[-30:]))

    progress_bar.empty()
    st.success(f"Tamamlandı — ✓ {ok} indirildi · – {skip} atlandı · ✗ {err} hata")
    st.info(f"Dosya konumu: `pbit_downloads/` ({out_root.resolve()})")


# render_connection_management function removed. Connection is now handled in sidebar.

def render_about():
    st.markdown('<div class="section-title">Hakkında</div>', unsafe_allow_html=True)
    st.write("Raportal Agent v2.0 - Kariyer.net için özel olarak geliştirilmiştir.")



def render_insights_hub():
    render_premium_header(
        title_main="Raportal",
        title_highlight="Insights Hub",
        subtitle="Akıllı Rapor Analiz ve Yönetim Merkezi",
        icon="hub",
        icon_color="#38bdf8",
        gradient_start="#1e1b4b",
        gradient_end="#4c1d95",
        highlight_gradient="linear-gradient(90deg, #d8b4fe 0%, #f472b6 100%)",
        tags=[
            ("insights", "Analiz", "#a78bfa"),
            ("storage", "Katalog", "#38bdf8")
        ]
    )

    col_main, col_info = st.columns([2.5, 1])

    with col_main:
        # DUAL WORKFLOW TABS
        tab_link, tab_list = st.tabs(["🔗 Link ile Analiz Et", "📂 Katalogdan Seç & Analiz Et"])

        with tab_link:
            st.html(f"""
<div class="premium-form-box">
    <div style="font-weight: 800; color: #1e1b4b; font-size: 1.1rem; margin-bottom: 5px;">Hızlı URL Analizi</div>
    <div style="color: #64748b; font-size: 0.85rem; margin-bottom: 20px;">Herhangi bir Raportal rapor linkini buraya yapıştırarak yapay zeka analizini anında başlatabilirsiniz.</div>
</div>
""")
            
            default_url = st.session_state.get('target_report_url', "")
            report_url_input = st.text_input("Rapor Linki (URL)", placeholder="https://raportal.kariyer.net/home/...", value=default_url, key="hub_link_input")
            
            prompt_addon = st.text_area("Özel Analiz Notu (Opsiyonel)", placeholder="Örn: Sadece son 3 aylık trendi yorumla...", height=100, key="hub_addon_link")
            
            st.html("""
<div style='margin-top: 15px; margin-bottom: 20px;'>
    <div style='font-size: 0.85rem; font-weight: 700; color: #1e1b4b; margin-bottom: 10px;'>Analiz Türü</div>
</div>
""")
            sel_cols = st.columns(4)
            sel_cols[0].checkbox("✨ Genel Analiz", value=True, key="an_gen")
            sel_cols[1].checkbox("📈 Trend", value=False, key="an_trend")
            sel_cols[2].checkbox("⚠️ Anomali", value=False, key="an_ano")
            sel_cols[3].checkbox("📝 Yönetici Özeti", value=False, key="an_exec")

            if st.button("🚀 ANALİZİ BAŞLAT", type="primary", use_container_width=True, key="btn_run_link_final"):
                if not report_url_input or "raportal.kariyer.net" not in report_url_input:
                    st.error("Lütfen geçerli bir Raportal linki girin.")
                else:
                    _run_vision_logic(report_url_input, True, 3, prompt_addon)
            
            st.html("<div style='text-align: center; margin-top: 15px; color: #64748b; font-size: 0.8rem;'><span class='material-icons-outlined' style='font-size: 1rem; vertical-align: middle;'>security</span> Verileriniz güvenli şekilde işlenir.</div>")
        with tab_list:
            if "biportal_conn" not in st.session_state or not st.session_state.biportal_conn:
                st.markdown('<div class="premium-form-box" style="text-align: center; border: 2px dashed #e2e8f0; background: #f8fafc;">', unsafe_allow_html=True)
                st.markdown('<span class="material-icons-outlined" style="font-size: 4rem; color: #94a3b8; margin-bottom: 1rem;">storage</span>', unsafe_allow_html=True)
                st.info("Katalog listesine erişebilmek için önce SQL sunucusuna bağlanmalısınız.")
                st.markdown('</div>', unsafe_allow_html=True)
            else:
                if "prof_catalog_df" not in st.session_state:
                    st.html("""
<div class="premium-form-box" style="text-align: center; padding: 1.5rem 1rem;">
    <span class="material-icons-outlined" style="font-size: 3.5rem; color: #8c28e8; margin-bottom: 0.5rem;">cloud_download</span>
    <h3 style="margin-bottom: 10px;">Raportal Kataloğu Hazır</h3>
    <p style="color: #64748b; font-size: 0.95rem; margin-bottom: 20px;">SQL bağlantısı kuruldu. Binlerce rapor arasından seçim yapmak için listeyi şimdi çekebilirsin.</p>
</div>
""")
                    if st.button("📋 RAPOR LİSTESİNİ GETİR", type="primary"):
                        _load_catalog_data()
                else:
                    df = st.session_state.prof_catalog_df
                    
                    n_ssrs = len(df[df['Tip'] == 'SSRS'])
                    n_pbi  = len(df[df['Tip'] == 'Power BI'])
                    
                    m1, m2, m3 = st.columns(3)
                    
                    def centered_metric(label, value, color="#8c28e8"):
                        st.markdown(f"""
                            <div style="text-align: center; padding: 20px; background: #f8fafc; border-radius: 16px; border: 1px solid #e2e8f0;">
                                <div style="color: #64748b; font-size: 0.9rem; font-weight: 600; margin-bottom: 5px;">{label}</div>
                                <div style="color: {color}; font-size: 2.2rem; font-weight: 800; line-height: 1;">{value}</div>
                            </div>
                        """, unsafe_allow_html=True)

                    with m1: centered_metric("Toplam Rapor", len(df))
                    with m2: centered_metric("🗒️ SSRS Raporu", n_ssrs, "#6366f1")
                    with m3: centered_metric("📊 Power BI", n_pbi, "#f59e0b")

                    st.markdown("---")
                    
                    # ROW 1: Search and Selection
                    col1, col2 = st.columns([1, 1])
                    with col1:
                        search = st.text_input("🔍 Listede Rapor Ara", placeholder="Rapor adı, klasör...", key="hub_list_search")
                    
                    # --- FILTER LOGIC ---
                    display_df = df.copy()
                    if 'URL' not in display_df.columns:
                        def _make_url(row):
                            base_url = "https://raportal.kariyer.net/home/"
                            path = str(row['Path']).lstrip('/')
                            encoded_path = "/".join([quote(p) for p in path.split('/')])
                            return f"{base_url}{('report' if row['Tip']=='SSRS' else 'powerbi')}/{encoded_path}"
                        display_df['URL'] = display_df.apply(_make_url, axis=1)
                    
                    if search:
                        mask = display_df.apply(lambda r: r.astype(str).str.contains(search, case=False).any(), axis=1)
                        display_df = display_df[mask]
                    # --------------------

                    with col2:
                        selected_name = st.selectbox(
                            "Raporu Seçin:", 
                            options=display_df['Name'].unique(),
                            index=0 if not display_df.empty else None,
                            key="hub_report_pick"
                        )

                    # ROW 2: Filter and Action
                    col3, col4 = st.columns([1, 1])
                    with col3:
                        sub_f1, sub_f2 = st.columns([0.85, 0.15])
                        with sub_f1:
                            folders = sorted(df["Klasör1"].dropna().unique().tolist())
                            selected_folder = st.selectbox("📁 Klasöre Göre Filtrele", ["Tümü"] + folders, key="hub_folder_filter")
                        with sub_f2:
                            st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
                            def _clear_hub_filters():
                                st.session_state.hub_list_search = ""
                                st.session_state.hub_folder_filter = "Tümü"
                                
                            if st.button("🧹", help="Filtreleri Sıfırla", key="hub_clear_filters", on_click=_clear_hub_filters):
                                pass # Callback handles it
                        
                        if selected_folder != "Tümü":
                            display_df = display_df[display_df["Klasör1"] == selected_folder]

                    with col4:
                        st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
                        if st.button("⚡ Seçili Raporu Analize Gönder", type="primary", use_container_width=True):
                            if selected_name:
                                # Refetch the row based on potentially filtered data
                                try:
                                    row = display_df[display_df['Name'] == selected_name].iloc[0]
                                    _run_vision_logic(row['URL'], True, 3, "")
                                except:
                                    st.error("Rapor bilgisi alınamadı.")

                    st.markdown("#### 📋 Tüm Katalog Listesi")
                    st.dataframe(
                        display_df[['Name', 'Klasör1', 'Tip', 'Kullanim', 'URL', 'Yetki']], 
                        use_container_width=True, 
                        height=400,
                        column_config={
                            "URL": st.column_config.LinkColumn("Raportal Linki", display_text="🔗 Aç"),
                            "Yetki": st.column_config.TextColumn("Yetkili Kullanıcılar", help="Bu rapora erişimi olan kullanıcıların listesi.", width="large")
                        }
                    )

def _load_catalog_data():
    with st.spinner("biportal kataloğu taranıyor..."):
        try:
            query = """
            SET NOCOUNT ON;
            WITH YetkiBase AS (
                SELECT DISTINCT C.[ItemID], REPLACE([UserName], N'KARIYER\\', '') AS UserName
                FROM [Raportal].[dbo].[Catalog] C WITH (NOLOCK)
                JOIN PolicyUserRole PR WITH (NOLOCK) ON C.PolicyID = PR.PolicyID
                JOIN Users U WITH (NOLOCK) ON U.UserID = PR.UserID
                WHERE Type NOT IN (1, 3, 5, 8)
            ),
            YetkiSummary AS (
                SELECT ItemID, STRING_AGG(UserName, CHAR(13)) AS Yetki
                FROM YetkiBase GROUP BY ItemID
            ),
            RaporlarDetailed AS (
                SELECT DISTINCT C.[ItemID]
                  , SUBSTRING(RIGHT(Path, LEN([Path]) - 1), 1, CHARINDEX('/', RIGHT(Path, LEN([Path]) - 1))) AS Klasör1
                  , [Path], [Name], Y.Yetki
                  , CASE 
                        WHEN [Type] = 11 THEN N'KPI' 
                        WHEN [Type] = 13 THEN N'Power BI' 
                        WHEN [Type] = 2 THEN N'SSRS' 
                        ELSE N'Other'
                    END AS Tip
                FROM [Raportal].[dbo].[Catalog] C WITH (NOLOCK)
                JOIN YetkiSummary AS Y ON Y.ItemID = C.ItemID
                WHERE Type NOT IN (1, 3, 5, 8)
            ),
            Usage90 AS (
                SELECT B.[ItemID], COUNT(*) AS Kullanim, COUNT(DISTINCT A.UserName) AS KullaniciSayisi
                FROM [dbo].[Catalog] B WITH (NOLOCK) 
                LEFT JOIN [dbo].[ExecutionLogStorage] A WITH (NOLOCK) ON A.[ReportID] = B.[ItemID]
                    AND DATEDIFF(DAY, CONVERT(DATE, A.[TimeStart]), CONVERT(DATE, GETDATE())) <= 90 
                WHERE B.[Type] IN (2,11,13) GROUP BY B.[ItemID]
            )
            SELECT R.*, ISNULL(U.Kullanim, 0) as Kullanim, ISNULL(U.KullaniciSayisi, 0) as KullaniciSayisi
            FROM RaporlarDetailed R
            LEFT JOIN Usage90 U ON R.ItemID = U.ItemID
            """
            df = run_query_on_bidb(query, server=st.session_state.sb_sql_srv, database=st.session_state.sb_sql_db)
            
            # Gereksiz klasörleri filtrele (Rating, Silinen Raporlar, Test)
            exclude_folders = ["Rating", "Silinen Raporlar", "Test"]
            df = df[~df["Klasör1"].str.contains("|".join(exclude_folders), case=False, na=False)]
            
            # URL kolonunu kalıcı olarak ekle
            def _make_url(row):
                base_url = "https://raportal.kariyer.net/home/"
                path = str(row['Path']).lstrip('/')
                encoded_path = "/".join([quote(p) for p in path.split('/')])
                return f"{base_url}{('report' if row['Tip']=='SSRS' else 'powerbi')}/{encoded_path}"
            df['URL'] = df.apply(_make_url, axis=1)

            st.session_state.prof_catalog_df = df
            st.rerun()
        except Exception as e:
            st.error(f"Katalog verisi alınamadı: {e}")

def _run_vision_logic(url, headless, sheets, addon):
    current_key = st.session_state.get('api_key')
    
    # Model selection resolution
    available_models = st.session_state.get("available_models", [])
    model_name = st.session_state.get("selected_model")
    if not model_name and available_models: model_name = available_models[0]
    if not model_name: model_name = GEMINI_MODEL

    with st.status("🚀 Akıllı Analiz Süreci Başlatıldı", expanded=True) as status:
        st.write("🌐 Browser motoru hazırlanıyor (Playwright)...")
        try:
            # 1. Aşama: Veri Toplama
            outcome = _run_agent_in_thread(
                url, 
                st.session_state.sb_sql_user, 
                st.session_state.sb_sql_pass, 
                st.session_state.sb_sql_domain, 
                current_key,
                headless=headless,
                model_name=model_name,
                max_pages=sheets
            )
            
            if not outcome.get("ok"):
                status.update(label="❌ Analiz Başarısız", state="error")
                st.error(f"Hata: {outcome.get('error')}")
                return

            dashboard_data = outcome.get("data")
            analysis = outcome.get("analysis")
            
            # 2. Aşama: Görselleştirme ve Sunum
            status.update(label="🤖 Veriler yakalandı, AI profesyonel yorumu üretiyor...", state="running")
            st.write("📊 Rapor sayfaları analiz ediliyor ve sunum scripti hazırlanıyor...")
            
            # Archive
            if dashboard_data and dashboard_data.pages:
                with st.expander(f"📁 Yakalanan Rapor Sayfaları ({len(dashboard_data.pages)})", expanded=False):
                    page_names = [p.tab_name for p in dashboard_data.pages]
                    tabs = st.tabs(page_names)
                    for t, p in zip(tabs, dashboard_data.pages):
                        with t:
                            if Path(p.screenshot_path).exists():
                                st.image(p.screenshot_path)
                            st.write(f"**Bulgular:** {', '.join(p.kpi_values)}")

            if analysis:
                st.markdown("### 📝 AI Profesyonel Yorumu & Sunum Scripti")
                st.markdown(f'<div style="background-color: white; padding: 20px; border-radius: 12px; border-left: 5px solid #8c28e8; color: #1f2937; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);">{analysis}</div>', unsafe_allow_html=True)
            
            status.update(label="✅ Analiz Başarıyla Tamamlandı", state="complete", expanded=False)
        except Exception as e:
            st.error(f"Beklenmeyen hata: {e}")
            traceback.print_exc()


def render_about():
    st.markdown('<div class="section-title">Hakkında</div>', unsafe_allow_html=True)
    st.write("Raportal Agent v2.5 - Kariyer.net için özel olarak geliştirilmiştir.")

# --- MAIN PAGE ROUTING ---
if st.session_state.active_page == "Anasayfa":
    render_home()
elif st.session_state.active_page == "Dashboard":
    render_dashboard()
elif st.session_state.active_page == "Raportal Insights Hub":
    render_insights_hub()
elif st.session_state.active_page == "PBIT İndir":
    render_pbit_downloader()
elif st.session_state.active_page == "Fix Sorgular":
    render_fix_sorgular()
elif st.session_state.active_page == "PBIX Analizi":
    render_pbix_analyzer()
else:
    render_about()
