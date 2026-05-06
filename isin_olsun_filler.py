import openpyxl
from copy import copy
from isin_olsun_queries import IsinOlsunQueryEngine
import os
from concurrent.futures import ThreadPoolExecutor
from openpyxl.styles import Font, Fill, Border, Alignment, PatternFill

def get_turkish_month_name(yyyymm):
    months = {
        "01": "Ocak", "02": "Şubat", "03": "Mart", "04": "Nisan",
        "05": "Mayıs", "06": "Haziran", "07": "Temmuz", "08": "Ağustos",
        "09": "Eylül", "10": "Ekim", "11": "Kasım", "12": "Aralık"
    }
    year = yyyymm[:4][2:] # '2026' -> '26'
    month = yyyymm[4:]
    return f"{months.get(month, month)}'{year}"

def copy_style(src_cell, dst_cell):
    """Copies the style from one cell to another with high fidelity."""
    if src_cell.font:
        dst_cell.font = copy(src_cell.font)
    if src_cell.fill:
        dst_cell.fill = copy(src_cell.fill)
    if src_cell.border:
        dst_cell.border = copy(src_cell.border)
    if src_cell.alignment:
        dst_cell.alignment = copy(src_cell.alignment)
    if src_cell.number_format:
        dst_cell.number_format = copy(src_cell.number_format)

def fill_isin_olsun_excel(file_path, target_month=None):
    if not os.path.exists(file_path):
        print("Error: File not found")
        return
    
    engine = IsinOlsunQueryEngine()
    dates = engine.calculate_dates(target_month)
    yyyymm = dates['yyyymm']
    month_title = get_turkish_month_name(yyyymm)
    
    # Run Aday and Firma queries in parallel
    print("Fetching data from database (Parallel)...")
    with ThreadPoolExecutor(max_workers=2) as executor:
        future_aday = executor.submit(engine.run_aday_queries, dates)
        future_firma = executor.submit(engine.run_firma_queries, dates)
        
        aday_data = future_aday.result()
        firma_data = future_firma.result()
    
    print("Loading workbook...")
    wb = openpyxl.load_workbook(file_path)
    
    # --- ADAY SHEET ---
    if "Aday" in wb.sheetnames:
        ws = wb["Aday"]
        target_col = None
        # Search for existing column
        for col in range(2, ws.max_column + 1):
            if ws.cell(row=2, column=col).value == f"Aday Sayısı ({month_title})":
                target_col = col
                break
        
        if not target_col:
            target_col = 2
            while ws.cell(row=2, column=target_col).value is not None:
                target_col += 1
        
        # Always ensure header style and header text is correct
        header_cell = ws.cell(row=2, column=target_col)
        header_cell.value = f"Aday Sayısı ({month_title})"
        header_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        aday_mapping = {
            'SMS_İzinliAday': 3,
            'EMAIL_İzinliAday': 4,
            'EMAIL_DoluOlanAday': 5,
            'PUSH_İzinliAday': 6,
            'TCKN_OnaylıAday': 7,
            'IsTecrubesiDolu': 8,
            'SirketTakipEden': 9,
            'BasvuranAday': 12,
            'AdresDolu': 14,             # Ev lokasyonu ekleyen
            'SirketSikayetEden': 15,
            'PuanlayanAday': 16,
            'SehirDolu': 18,             # Şehir Bilgisi Dolu Olan Aday
            'IlanSikayetEden': 19
        }
        
        for metric, row in aday_mapping.items():
            if metric in aday_data:
                cell = ws.cell(row=row, column=target_col)
                cell.value = aday_data[metric]
                cell.number_format = '#,##0'
                # Copy style from previous column for each row
                copy_style(ws.cell(row=row, column=target_col-1), cell)
                # Ensure the number format and alignment are exactly what we want
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
        
        # Explicitly clear rows 13 and 17 as requested
        ws.cell(row=13, column=target_col).value = None
        ws.cell(row=17, column=target_col).value = None
        
        print("Aday sheet updated.")

    # --- İŞVEREN SHEET ---
    if "İşveren" in wb.sheetnames:
        ws = wb["İşveren"]
        target_col = None
        for col in range(2, ws.max_column + 1):
            header_val = ws.cell(row=2, column=col).value
            if header_val and (f"Firma Sayısı ({month_title})" in header_val or f"İşveren Sayısı ({month_title})" in header_val):
                target_col = col
                break
        
        if not target_col:
            target_col = 2
            while ws.cell(row=2, column=target_col).value is not None:
                target_col += 1
        
        # Always ensure header style and header text is correct (Now "Firma Sayısı")
        header_cell = ws.cell(row=2, column=target_col)
        header_cell.value = f"Firma Sayısı ({month_title})"
        header_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        firma_mapping = {
            'SMS_İzinliFirma': 3,
            'EMAIL_İzinliFirma': 4,
            'EMAIL_DoluOlanFirma': 5,
            'PUSH_İzinliFirma': 6,
            'TCKN_OnaylıFirma': 7,
            'VKKN_OnaylıFirma': 8,
            'EvrakOnayliFirmaTotal': 9,    # Evrak Doğrulamış (Snapshot Row 9)
            'OnayliIsverenRozeti': 10,     # Onaylı İşveren Rozeti Olan (Row 10)
            'RozetHakKazananMonthly': 11,  # Rozet Almaya Hak Kazanan Firma (Monthly Row 11)
            'ToplamRozetHakKazanan': 12,   # Toplam Rozet Almaya Hak Kazanmışlar (Row 12)
            'IlanYayinlayanFirmaTotal': 14, # İlan Yayınlayan Firma (Total Count Row 14)
            'EvrakOnaylayanMonthly': 15    # Evrak Onaylayan Firma (Monthly Row 15)
        }
        
        for metric, row in firma_mapping.items():
            if metric in firma_data:
                cell = ws.cell(row=row, column=target_col)
                cell.value = firma_data[metric]
                cell.number_format = '#,##0'
                copy_style(ws.cell(row=row, column=target_col-1), cell)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
                
                # Special logic for Row 14 (Ilan Yayınlayan)
                if row == 14:
                    active_count = firma_data.get('IlanYayinlayanFirmaActive', 0)
                    # Format the number with thousands separator (dot)
                    formatted_active = "{:,}".format(active_count).replace(",", ".")
                    
                    # Add complex comment to the right of current value (target_col + 1)
                    comment_cell = ws.cell(row=14, column=target_col+1)
                    comment_cell.value = f"{formatted_active} ilan ise silinmiş ilanlar hariçtir."
                    comment_cell.font = Font(italic=True, size=9)
        
        print("Isveren sheet updated.")

    try:
        wb.save(file_path)
        print("Success! Excel saved.")
    except PermissionError:
        raise RuntimeError("⚠️ Lütfen masaüstünde açık olan Excel dosyasını kapatın ve tekrar deneyin.")
